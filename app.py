import streamlit as st
import streamlit.components.v1 as components
import os
import pandas as pd
import datetime
import io
import requests
import base64
from PIL import Image
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from openpyxl.worksheet.datavalidation import DataValidation
import json
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
# 1. CONFIGURACIÓN E INTERFAZ BASE RESPONSIVA
st.set_page_config(
    page_title="Remisiones de Materiales",
    layout="wide",
    page_icon="📦"
)

# Inyectar CSS de Imagen Corporativa Oficial (Industria SIGRAMA)
st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=Montserrat:wght@400;500;700&family=Questrial&display=swap');

    /* Fuentes globales */
    html, body, [class*="css"], .stApp {
        font-family: 'Questrial', sans-serif !important;
        background-color: #FFFFFF !important;
    }

    h1, h2, h3, h4, h5, h6, .main-title {
        font-family: 'Montserrat', sans-serif !important;
        font-weight: 700 !important;
        color: #111111 !important;
    }

    /* Barra lateral corporativa en Negro profundo #111111 */
    [data-testid="stSidebar"] {
        background-color: #111111 !important;
        border-right: 1px solid #1E293B !important;
    }
    [data-testid="stSidebar"] [data-testid="stMarkdownContainer"] p, 
    [data-testid="stSidebar"] span, 
    [data-testid="stSidebar"] label {
        color: #FFFFFF !important;
        font-family: 'Questrial', sans-serif !important;
    }
    
    /* Logo negativo en sidebar */
    [data-testid="stSidebar"] img {
        filter: grayscale(1) invert(1) brightness(1.2) contrast(1.2) !important;
    }

    /* Botones de navegación en barra lateral */
    [data-testid="stSidebar"] div[role="radiogroup"] label {
        color: #FFFFFF !important;
        font-size: 14px !important;
    }
    [data-testid="stSidebar"] div[role="radiogroup"] label:hover {
        color: #EC2024 !important;
    }

    /* Estilos para inputs de contraseña en barra lateral */
    [data-testid="stSidebar"] input {
        background-color: #1E293B !important;
        color: #FFFFFF !important;
        border-color: #334155 !important;
    }
    [data-testid="stSidebar"] input:focus {
        border-color: #EC2024 !important;
    }

    /* Estilo de Botones Oficiales - Rojo Corporativo #EC2024 */
    div.stButton > button,
    div.stDownloadButton > button,
    div.stFormSubmitButton > button,
    button[data-testid="baseButton-secondary"]:not([role="tab"]):not([data-baseweb="tab"]),
    button[data-testid="baseButton-primary"]:not([role="tab"]):not([data-baseweb="tab"]),
    button[kind="secondary"]:not([role="tab"]):not([data-baseweb="tab"]),
    button[kind="primary"]:not([role="tab"]):not([data-baseweb="tab"]) {
        background-color: #EC2024 !important;
        color: #FFFFFF !important;
        font-family: 'Montserrat', sans-serif !important;
        font-weight: 700 !important;
        border-radius: 4px !important;
        border: 1px solid #EC2024 !important;
        padding: 8px 20px !important;
        transition: all 0.3s ease !important;
        text-transform: uppercase;
        font-size: 13px !important;
    }
    div.stButton > button:hover,
    div.stDownloadButton > button:hover,
    div.stFormSubmitButton > button:hover,
    button[data-testid="baseButton-secondary"]:not([role="tab"]):not([data-baseweb="tab"]):hover,
    button[data-testid="baseButton-primary"]:not([role="tab"]):not([data-baseweb="tab"]):hover,
    button[kind="secondary"]:not([role="tab"]):not([data-baseweb="tab"]):hover,
    button[kind="primary"]:not([role="tab"]):not([data-baseweb="tab"]):hover {
        background-color: #FFFFFF !important;
        color: #EC2024 !important;
        border: 1px solid #EC2024 !important;
        box-shadow: 0 4px 12px rgba(236, 32, 36, 0.15) !important;
    }

    /* Tarjetas de Métricas */
    [data-testid="metric-container"] {
        background-color: #FFFFFF !important;
        border: 1px solid #D2D3D5 !important;
        border-left: 5px solid #EC2024 !important;
        border-radius: 4px !important;
        padding: 12px 18px !important;
        box-shadow: 0 2px 4px rgba(0,0,0,0.05) !important;
    }
    [data-testid="metric-container"] label {
        font-family: 'Montserrat', sans-serif !important;
        color: #111111 !important;
        font-weight: 500 !important;
    }
    [data-testid="metric-container"] div[data-testid="stMetricValue"] {
        font-family: 'Montserrat', sans-serif !important;
        color: #EC2024 !important;
        font-weight: 700 !important;
    }
    
    /* Configuración del Editor de Datos y Tablas */
    .stTable header, th {
        background-color: #111111 !important;
        color: #FFFFFF !important;
        font-family: 'Montserrat', sans-serif !important;
    }
    
    /* Inputs y Selectores */
    div[data-baseweb="input"], div[data-baseweb="select"], textarea {
        border-color: #D2D3D5 !important;
        border-radius: 4px !important;
    }
    div[data-baseweb="input"]:focus-within, div[data-baseweb="select"]:focus-within {
        border-color: #EC2024 !important;
    }

    /* Estilo de las pestañas */
    button[role="tab"] {
        font-family: 'Montserrat', sans-serif !important;
        font-weight: bold !important;
        color: #111111 !important;
    }
    
    button[role="tab"][aria-selected="true"] {
        color: #EC2024 !important;
        border-bottom-color: #EC2024 !important;
    }

    /* Reset del File Uploader para encajar en el estilo secundario */
    [data-testid="stFileUploader"] button,
    .stFileUploader button {
        background-color: #FFFFFF !important;
        color: #111111 !important;
        border: 1px solid #D2D3D5 !important;
        border-radius: 4px !important;
        padding: 6px 12px !important;
        font-weight: 500 !important;
        box-shadow: none !important;
        transform: none !important;
    }
    
    [data-testid="stFileUploader"] button *,
    .stFileUploader button * {
        background-color: transparent !important;
        color: #111111 !important;
    }
    
    [data-testid="stFileUploader"] button:hover,
    .stFileUploader button:hover {
        background-color: #F8F9FA !important;
        border-color: #EC2024 !important;
        color: #EC2024 !important;
    }
    
    [data-testid="stFileUploader"] button:hover *,
    .stFileUploader button:hover * {
        color: #EC2024 !important;
    }

    .report-card {
        background-color: #FFFFFF;
        border: 1px solid #D2D3D5;
        border-radius: 8px;
        padding: 20px;
        margin-bottom: 20px;
        box-shadow: 0 1px 3px rgba(0,0,0,0.05);
    }

    /* Ocultar elementos predeterminados de Streamlit (footer, main menu, deploy button, etc.) */
    footer {visibility: hidden;}
    #MainMenu {visibility: hidden;}
    header {visibility: hidden;}
    .stAppDeployButton {display: none !important;}
    [data-testid="stViewerBadge"] {display: none !important;}
</style>
""", unsafe_allow_html=True)

# Renderizado de Banner Corporativo Adaptable
try:
    banner_img = Image.open("REMISIONES APP.png")
    st.image(banner_img, use_container_width=True)
except FileNotFoundError:
    st.warning("⚠️ Cargando interfaz gráfica del banner superior corporativo...")

# Slogan de Resultados / Transformación Principal
st.markdown('<p style="text-align: center; font-size: 16px; font-weight: bold; color: #EC2024; font-family: \'Montserrat\', sans-serif; margin-top: 15px; text-transform: uppercase; letter-spacing: 1px;">SOLUCIONES QUE TRANSFORMAN TU EMPRESA</p>', unsafe_allow_html=True)
st.markdown('<hr style="border: 1px solid #EC2024; margin: 15px 0;">', unsafe_allow_html=True)

# =============================================================================
# 2. MOTOR DE PERSISTENCIA CERTIFICADO (CONEXIÓN UNIFICADA POR API GITHUB)
# =============================================================================
REPO_OWNER = "jesusalbertomoraleslopez-byte"
REPO_NAME = "remisiones-de-materiales"
BRANCH = "main"

def clean_project_val(val):
    if pd.isna(val):
        return ""
    val_str = str(val).strip()
    if not val_str:
        return ""
    
    # Si es "RECHASO", "RECHAZO" o variantes, mapear a "Material Rechazado"
    if val_str.upper() in ["RECHASO", "RECHAZO", "MATERIAL RECHAZADO", "MATRIAL RECHAZADO"]:
        return "Material Rechazado"
    
    # Si es puramente un número entero (ej. 3, 15)
    if val_str.isdigit():
        return f"INT-{int(val_str):03d}"
    
    # Si es un número decimal que termina en .0 (ej. 3.0, 15.0)
    try:
        float_val = float(val_str)
        if float_val.is_integer():
            return f"INT-{int(float_val):03d}"
    except ValueError:
        pass
        
    # Si tiene el patrón INT- seguido de un número (ej. INT-3, INT-03, INT-003)
    import re
    match = re.match(r'^INT\s*-\s*(\d+)$', val_str, re.IGNORECASE)
    if match:
        num = int(match.group(1))
        return f"INT-{num:03d}"
        
    # De lo contrario, dejarlo como está (ej. LC8, RECHASO)
    return val_str

def clean_po_val(val):
    if pd.isna(val):
        return ""
    val_str = str(val).strip()
    if not val_str:
        return ""
    
    # Extraer todos los dígitos de la cadena
    import re
    digits = "".join(re.findall(r'\d', val_str))
    
    # Si tenemos exactamente 8 dígitos, formatear como PO XXXX-XXXX
    if len(digits) == 8:
        return f"PO {digits[:4]}-{digits[4:]}"
        
    # De lo contrario, dejarlo limpio pero como está
    return val_str

def cargar_excel_desde_github(file_name):
    """Carga el archivo Excel: primero busca en disco local (modo offline), luego en GitHub API."""
    import os
    # 1. Intentar cargar localmente primero
    if os.path.exists(file_name):
        try:
            try:
                return pd.read_excel(file_name, sheet_name='Datos_Sistema')
            except Exception:
                return pd.read_excel(file_name, sheet_name=0)
        except Exception as e:
            st.warning(f"⚠️ Error al leer archivo local {file_name}: {e}")
            
    # 2. Si no existe localmente, intentar GitHub
    try:
        url = f"https://api.github.com/repos/{REPO_OWNER}/{REPO_NAME}/contents/{file_name}?ref={BRANCH}"
        headers = {}
        if "github_token" in st.secrets:
            headers["Authorization"] = f"token {st.secrets['github_token']}"
            headers["Accept"] = "application/vnd.github.v3+json"
            
        res = requests.get(url, headers=headers)
        if res.status_code == 200:
            datos_json = res.json()
            contenido_base64 = datos_json["content"]
            archivo_bytes = base64.b64decode(contenido_base64)
            try:
                return pd.read_excel(io.BytesIO(archivo_bytes), sheet_name='Datos_Sistema')
            except Exception:
                return pd.read_excel(io.BytesIO(archivo_bytes), sheet_name=0)
    except Exception:
        pass
    return None

def subir_excel_a_github(file_name, dataframe_to_save):
    """Guarda el archivo localmente y luego intenta sincronizarlo con GitHub si hay token disponible."""
    # 1. Guardar localmente siempre
    try:
        with pd.ExcelWriter(file_name, engine='openpyxl') as writer:
            dataframe_to_save.to_excel(writer, index=False, sheet_name='Datos_Sistema')
    except Exception as e:
        st.error(f"⚠️ Error al guardar archivo localmente {file_name}: {e}")
        
    # 2. Sincronizar con GitHub si el token está disponible
    if "github_token" not in st.secrets or not st.secrets["github_token"]:
        # Modo local puro sin token, retornamos True ya que se guardó localmente
        return True
        
    try:
        GITHUB_TOKEN = st.secrets["github_token"]
        url = f"https://api.github.com/repos/{REPO_OWNER}/{REPO_NAME}/contents/{file_name}"
        headers = {"Authorization": f"token {GITHUB_TOKEN}", "Accept": "application/vnd.github.v3+json"}

        # Convertir DataFrame a bytes de Excel en memoria
        buffer_git = io.BytesIO()
        with pd.ExcelWriter(buffer_git, engine='openpyxl') as writer:
            dataframe_to_save.to_excel(writer, index=False, sheet_name='Datos_Sistema')

        base64_content = base64.b64encode(buffer_git.getvalue()).decode("utf-8")
        
        # Obtener el SHA del archivo existente para poder reemplazarlo
        res_get = requests.get(url, headers=headers)
        sha = res_get.json().get("sha") if res_get.status_code == 200 else None

        payload = {
            "message": f"Sincronizacion App: {file_name}", 
            "content": base64_content, 
            "branch": BRANCH
        }
        if sha: 
            payload["sha"] = sha

        res_put = requests.put(url, json=payload, headers=headers)
        return res_put.status_code in [200, 201]

    except Exception as e:
        st.warning(f"⚠️ No se pudo sincronizar con GitHub: {e}")
        return False

@st.cache_data(ttl=60)
def obtener_skus_con_imagen():
    """Obtiene el conjunto de SKUs que tienen una imagen asociada local o remotamente."""
    import os
    import requests
    skus = set()
    
    # 1. Escaneo local
    if os.path.exists("imagenes_articulos"):
        for f in os.listdir("imagenes_articulos"):
            if "(" in f:
                sku = f.split("(")[0]
                skus.add(sku)
                
    # 2. Escaneo remoto (GitHub)
    if "github_token" in st.secrets and st.secrets["github_token"]:
        try:
            GITHUB_TOKEN = st.secrets["github_token"]
            url_list = f"https://api.github.com/repos/{REPO_OWNER}/{REPO_NAME}/contents/imagenes_articulos?ref={BRANCH}"
            headers = {"Authorization": f"token {GITHUB_TOKEN}", "Accept": "application/vnd.github.v3+json"}
            res = requests.get(url_list, headers=headers)
            if res.status_code == 200:
                for item in res.json():
                    if "(" in item["name"]:
                        sku = item["name"].split("(")[0]
                        skus.add(sku)
        except Exception:
            pass
            
    return skus

def renderizar_explorador_imagenes():
    st.write("##### 🖼️ Carpeta de Imágenes de Artículos")
    st.markdown("Visualice las imágenes guardadas en el sistema y descargue la base de datos de imágenes completa en formato ZIP:")

    import os
    import zipfile
    import io
    import requests

    # Botón de sincronización con GitHub
    if st.button("🔄 Sincronizar Imágenes con GitHub", use_container_width=True, key="btn_sync_images_explorer"):
        if "github_token" in st.secrets and st.secrets["github_token"]:
            try:
                GITHUB_TOKEN = st.secrets["github_token"]
                url_list = f"https://api.github.com/repos/{REPO_OWNER}/{REPO_NAME}/contents/imagenes_articulos?ref={BRANCH}"
                headers = {"Authorization": f"token {GITHUB_TOKEN}", "Accept": "application/vnd.github.v3+json"}
                res_list = requests.get(url_list, headers=headers)
                if res_list.status_code == 200:
                    items_git = res_list.json()
                    downloaded_count = 0
                    os.makedirs("imagenes_articulos", exist_ok=True)
                    for it in items_git:
                        git_file_path = f"imagenes_articulos/{it['name']}"
                        if not os.path.exists(git_file_path):
                            if descargar_imagen_desde_github(git_file_path):
                                downloaded_count += 1
                    st.success(f"✅ Sincronización completada. Se descargaron {downloaded_count} imágenes nuevas de GitHub.")
                    st.rerun()
                else:
                    st.error(f"Error al listar repositorio: Código de estado {res_list.status_code}")
            except Exception as e_sync:
                st.error(f"Error al sincronizar: {e_sync}")
        else:
            st.warning("⚠️ Token de GitHub no configurado para sincronizar de forma remota.")

    st.write("---")

    # Generación de ZIP en memoria
    buf_zip = io.BytesIO()
    has_images = False
    if os.path.exists("imagenes_articulos"):
        file_list_local = [f for f in os.listdir("imagenes_articulos") if os.path.isfile(os.path.join("imagenes_articulos", f)) and not f.startswith(".")]
        if len(file_list_local) > 0:
            has_images = True
            with zipfile.ZipFile(buf_zip, 'w', zipfile.ZIP_DEFLATED) as zip_file:
                for file in file_list_local:
                    file_path = os.path.join("imagenes_articulos", file)
                    zip_file.write(file_path, file)
    buf_zip.seek(0)

    col_explorer1, col_explorer2 = st.columns([3, 1])
    with col_explorer2:
        st.download_button(
            label="📥 Descargar Todo (ZIP)",
            data=buf_zip.getvalue(),
            file_name="imagenes_articulos.zip",
            mime="application/zip",
            disabled=not has_images,
            use_container_width=True,
            key="btn_download_images_zip"
        )
        
    with col_explorer1:
        cant_imagenes = len(os.listdir("imagenes_articulos")) if os.path.exists("imagenes_articulos") else 0
        st.write(f"📁 **Directorio:** `imagenes_articulos/` ({cant_imagenes} archivos locales)")

    # Listar archivos
    if has_images:
        files_sorted = sorted([f for f in os.listdir("imagenes_articulos") if os.path.isfile(os.path.join("imagenes_articulos", f)) and not f.startswith(".")])
        for f in files_sorted:
            f_path = os.path.join("imagenes_articulos", f)
            f_size = os.path.getsize(f_path) / 1024.0 # KB
            
            c_img, c_info, c_actions = st.columns([1, 3, 1])
            with c_img:
                try:
                    st.image(f_path, width=60)
                except Exception:
                    st.write("🖼️") # Fallback si la imagen está corrupta
            with c_info:
                st.markdown(f"📄 **Archivo:** `{f}`")
                st.write(f"⚖️ **Tamaño:** `{f_size:.1f} KB`")
            with c_actions:
                try:
                    with open(f_path, "rb") as file_bytes:
                        st.download_button(
                            label="📥 Descargar",
                            data=file_bytes.read(),
                            file_name=f,
                            mime="image/png" if f.endswith(".png") else "image/jpeg",
                            key=f"btn_dl_single_img_{f}"
                        )
                except Exception:
                    st.write("Error")
            st.write("---")
    else:
        st.info("Actualmente no hay imágenes descargadas localmente en la carpeta `imagenes_articulos/`. Haz clic en 'Sincronizar Imágenes con GitHub' para descargar las imágenes que existan en el repositorio.")

def generar_pdf_catalogo_articulos(df_articulos):
    import io
    import os
    import glob
    import requests
    import datetime
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image as RLImage
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    
    buf = io.BytesIO()
    
    # 36pt margins -> 540pt width
    doc = SimpleDocTemplate(buf, pagesize=letter, leftMargin=36, rightMargin=36, topMargin=90, bottomMargin=60)
    
    story = []
    
    # Decoración de fondo / Cabecera repetitiva limpia sin logotipo
    def draw_catalog_decorations(canvas, doc):
        canvas.saveState()
        # Línea roja corporativa fina superior (Pantone 485 C)
        canvas.setStrokeColor(colors.HexColor("#EC2024"))
        canvas.setLineWidth(1.5)
        canvas.line(36, 755, 576, 755)
        
        # Título principal en negro/gris alineado a la izquierda
        canvas.setFillColor(colors.HexColor("#212121"))
        canvas.setFont("Helvetica-Bold", 10)
        canvas.drawString(36, 762, "INDUSTRIA SIGRAMA S.A. DE C.V. — REPORTE DE CATÁLOGO MAESTRO")
        
        # Fecha de generación alineada a la derecha
        canvas.setFillColor(colors.HexColor("#616161"))
        canvas.setFont("Helvetica", 8)
        canvas.drawRightString(576, 762, datetime.date.today().strftime("%d-%b-%Y"))
        
        # Pie de página (Línea roja fina)
        canvas.setStrokeColor(colors.HexColor("#EC2024"))
        canvas.setLineWidth(1)
        canvas.line(36, 50, 576, 50)
        
        canvas.setFillColor(colors.HexColor("#616161"))
        canvas.setFont("Helvetica", 8)
        canvas.drawString(36, 38, "Industria SIGRAMA S.A. de C.V. - Sistema Remisiones")
        canvas.drawRightString(576, 38, f"Página {canvas._pageNumber}")
        canvas.restoreState()

    styles = getSampleStyleSheet()
    
    # Crear estilos personalizados si no existen
    try:
        style_header = ParagraphStyle(
            'CatalogHeaderStyle',
            parent=styles['Normal'],
            fontName='Helvetica-Bold',
            fontSize=10,
            textColor=colors.white,
            alignment=1 # Centrado
        )
    except Exception:
        style_header = styles['Normal']
        
    try:
        style_cell = ParagraphStyle(
            'CatalogCellText',
            parent=styles['Normal'],
            fontName='Helvetica',
            fontSize=9,
            textColor=colors.HexColor("#212121"),
            leading=11
        )
    except Exception:
        style_cell = styles['Normal']

    try:
        style_cell_bold = ParagraphStyle(
            'CatalogCellTextBold',
            parent=styles['Normal'],
            fontName='Helvetica-Bold',
            fontSize=9,
            textColor=colors.HexColor("#212121"),
            leading=11
        )
    except Exception:
        style_cell_bold = styles['Normal']

    # Título en el body
    story.append(Spacer(1, 10))
    
    # Construcción de la tabla
    # Columnas: SKU, Imagen, Descripción, Calibre, Dimensiones, Acabado
    tabla_datos = [[
        Paragraph("SKU", style_header),
        Paragraph("Imagen", style_header),
        Paragraph("Descripción", style_header),
        Paragraph("Calibre", style_header),
        Paragraph("Dimensiones", style_header),
        Paragraph("Acabado", style_header)
    ]]
    
    # Pre-cargar lista remota si es necesario para evitar múltiples llamadas en bucle
    github_items = []
    if "github_token" in st.secrets and st.secrets["github_token"]:
        try:
            GITHUB_TOKEN = st.secrets["github_token"]
            url_list = f"https://api.github.com/repos/{REPO_OWNER}/{REPO_NAME}/contents/imagenes_articulos?ref={BRANCH}"
            headers = {"Authorization": f"token {GITHUB_TOKEN}", "Accept": "application/vnd.github.v3+json"}
            res_list = requests.get(url_list, headers=headers)
            if res_list.status_code == 200:
                github_items = res_list.json()
        except Exception:
            pass

    for _, row in df_articulos.iterrows():
        sku = str(row['SKU']).strip()
        nombre = str(row['Nombre'])
        calibre = str(row['Calibre_Espesor'])
        dimensiones = str(row['Dimensiones_Pieza'])
        acabado = str(row['Acabado_Superficial'])
        
        # Buscar imagen
        img_path = None
        matching_local = glob.glob(f"imagenes_articulos/{sku}(*.*")
        if matching_local:
            img_path = matching_local[0]
        else:
            # Buscar en lista de GitHub pre-cargada
            for git_item in github_items:
                if git_item["name"].startswith(f"{sku}("):
                    github_file_path = f"imagenes_articulos/{git_item['name']}"
                    if descargar_imagen_desde_github(github_file_path):
                        img_path = github_file_path
                        break
        
        # Celda de imagen
        img_cell = ""
        if img_path and os.path.exists(img_path):
            try:
                # Usar tamaño 60x60
                img_cell = RLImage(img_path, width=60, height=60, hAlign='CENTER')
            except Exception:
                img_cell = Paragraph("<font color='red'>Error de carga</font>", style_cell)
        else:
            img_cell = Paragraph("<font color='#9e9e9e'>Sin imagen</font>", style_cell)
            
        tabla_datos.append([
            Paragraph(f"<b>{sku}</b>", style_cell_bold),
            img_cell,
            Paragraph(nombre, style_cell),
            Paragraph(calibre, style_cell),
            Paragraph(dimensiones, style_cell),
            Paragraph(acabado, style_cell)
        ])
        
    # Ancho total: 540 pt
    t_cat = Table(tabla_datos, colWidths=[95, 80, 125, 75, 95, 70])
    t_cat.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#EC2024")),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor("#E0E0E0")),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('ALIGN', (1,1), (1,-1), 'CENTER'),
        ('TOPPADDING', (0,0), (-1,-1), 6),
        ('BOTTOMPADDING', (0,0), (-1,-1), 6),
        ('LEFTPADDING', (0,0), (-1,-1), 6),
        ('RIGHTPADDING', (0,0), (-1,-1), 6)
    ]))
    
    story.append(t_cat)
    
    doc.build(story, onFirstPage=draw_catalog_decorations, onLaterPages=draw_catalog_decorations)
    buf.seek(0)
    return buf.getvalue()


def subir_imagen_a_github(file_path):
    """Sube una imagen local a GitHub utilizando API REST, codificando la ruta de forma segura."""
    import os
    import urllib.parse
    if not os.path.exists(file_path):
        return False
    if "github_token" not in st.secrets or not st.secrets["github_token"]:
        return True
    try:
        with open(file_path, "rb") as f:
            base64_content = base64.b64encode(f.read()).decode("utf-8")
        
        GITHUB_TOKEN = st.secrets["github_token"]
        quoted_path = urllib.parse.quote(file_path.replace("\\", "/"))
        url = f"https://api.github.com/repos/{REPO_OWNER}/{REPO_NAME}/contents/{quoted_path}"
        headers = {"Authorization": f"token {GITHUB_TOKEN}", "Accept": "application/vnd.github.v3+json"}
        
        res_get = requests.get(url, headers=headers)
        sha = res_get.json().get("sha") if res_get.status_code == 200 else None
        
        payload = {
            "message": f"Sincronizacion de Imagen: {file_path}",
            "content": base64_content,
            "branch": BRANCH
        }
        if sha:
            payload["sha"] = sha
            
        res_put = requests.put(url, json=payload, headers=headers)
        return res_put.status_code in [200, 201]
    except Exception as e:
        st.warning(f"⚠️ No se pudo sincronizar la imagen {file_path} con GitHub: {e}")
        return False


def descargar_imagen_desde_github(file_path):
    """Intenta descargar la imagen de GitHub si no existe en local."""
    import os
    import urllib.parse
    if os.path.exists(file_path):
        return True
    if "github_token" not in st.secrets or not st.secrets["github_token"]:
        return False
    try:
        quoted_path = urllib.parse.quote(file_path.replace("\\", "/"))
        url = f"https://api.github.com/repos/{REPO_OWNER}/{REPO_NAME}/contents/{quoted_path}?ref={BRANCH}"
        headers = {"Authorization": f"token {st.secrets['github_token']}", "Accept": "application/vnd.github.v3+json"}
        
        res = requests.get(url, headers=headers)
        if res.status_code == 200:
            os.makedirs(os.path.dirname(file_path), exist_ok=True)
            datos_json = res.json()
            contenido_base64 = datos_json["content"]
            with open(file_path, "wb") as f:
                f.write(base64.b64decode(contenido_base64))
            return True
    except Exception:
        pass
    return False


def eliminar_imagen_de_github(file_path):
    """Elimina la imagen localmente y en el repositorio de GitHub."""
    import os
    import urllib.parse
    if os.path.exists(file_path):
        try:
            os.remove(file_path)
        except Exception as e:
            st.error(f"⚠️ Error al borrar imagen local {file_path}: {e}")
            
    if "github_token" not in st.secrets or not st.secrets["github_token"]:
        return True
    try:
        GITHUB_TOKEN = st.secrets["github_token"]
        quoted_path = urllib.parse.quote(file_path.replace("\\", "/"))
        url = f"https://api.github.com/repos/{REPO_OWNER}/{REPO_NAME}/contents/{quoted_path}"
        headers = {"Authorization": f"token {GITHUB_TOKEN}", "Accept": "application/vnd.github.v3+json"}
        
        res_get = requests.get(url, headers=headers)
        if res_get.status_code == 200:
            sha = res_get.json().get("sha")
            payload = {
                "message": f"Eliminar Imagen: {file_path}",
                "sha": sha,
                "branch": BRANCH
            }
            res_delete = requests.delete(url, json=payload, headers=headers)
            return res_delete.status_code in [200, 204]
    except Exception as e:
        st.warning(f"⚠️ No se pudo eliminar la imagen {file_path} de GitHub: {e}")
    return False



def obtener_emails_config():
    cfg_path = "config_emails.json"
    default_cfg = {"dest_to": "logistica@sigrama.com.mx; almacen@sigrama.com.mx", "dest_cc": "calidad@sigrama.com.mx; produccion@sigrama.com.mx"}
    if os.path.exists(cfg_path):
        try:
            with open(cfg_path, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            return default_cfg
    else:
        try:
            with open(cfg_path, "w", encoding="utf-8") as f:
                json.dump(default_cfg, f, indent=4)
        except Exception:
            pass
        return default_cfg

def guardar_emails_config(to_str, cc_str):
    cfg_path = "config_emails.json"
    cfg = {"dest_to": to_str.strip(), "dest_cc": cc_str.strip()}
    try:
        with open(cfg_path, "w", encoding="utf-8") as f:
            json.dump(cfg, f, indent=4)
        return True
    except Exception as e:
        st.error(f"Error al guardar configuración de correos: {e}")
        return False

def generar_archivo_eml(dest_to, dest_cc, subject, body_html, adjuntos_dict):
    """
    Genera un archivo EML en memoria como borrador de Outlook.
    adjuntos_dict: dict { 'nombre_archivo.pdf': bytes_del_archivo }
    """
    msg = MIMEMultipart('mixed')
    msg['Subject'] = subject
    msg['To'] = dest_to
    msg['Cc'] = dest_cc
    msg['X-Unsent'] = '1'  # Clave para que Outlook lo abra como borrador editable
    
    # Cuerpo HTML
    body_part = MIMEText(body_html, 'html', 'utf-8')
    msg.attach(body_part)
    
    # Adjuntos
    for filename, file_bytes in adjuntos_dict.items():
        if not file_bytes:
            continue
        
        # Convertir a bytes si es un objeto similar a BytesIO o archivo
        if hasattr(file_bytes, 'getvalue'):
            file_bytes = file_bytes.getvalue()
        elif hasattr(file_bytes, 'read'):
            file_bytes = file_bytes.read()
            
        part = MIMEBase('application', 'octet-stream')
        part.set_payload(file_bytes)
        encoders.encode_base64(part)
        part.add_header('Content-Disposition', f'attachment; filename="{filename}"')
        msg.attach(part)
        
    return msg.as_bytes()

def generar_cuerpo_correo_po_html(po_name, cab_info, df_matrix, fechas_columnas):
    # Cabecera informativa
    html = f"""
    <html>
    <head>
        <style>
            body {{ font-family: Arial, sans-serif; color: #333333; }}
            .header-table {{ border-collapse: collapse; margin-bottom: 20px; width: 100%; max-width: 600px; }}
            .header-table td {{ padding: 6px 12px; border: 1px solid #e0e0e0; }}
            .header-table td.label {{ font-weight: bold; background-color: #f5f5f5; width: 150px; }}
            .matrix-table {{ border-collapse: collapse; width: 100%; margin-top: 15px; font-size: 13px; table-layout: fixed; }}
            .summary-row {{ font-weight: bold; background-color: #f9f9f9; }}
        </style>
    </head>
    <body>
        <h2>Reporte de Avance y Dispersión de PO {po_name}</h2>
        <p>Estimado equipo,</p>
        <p>A continuación se presenta el reporte de avance y dispersión de entregas de la Orden de Compra solicitado:</p>
        
        <table class="header-table">
            <tr>
                <td class="label">Orden de Compra:</td>
                <td><b>PO {po_name}</b></td>
            </tr>
            <tr>
                <td class="label">Proyecto:</td>
                <td>{cab_info.get('Proyecto', 'N/A')}</td>
            </tr>
            <tr>
                <td class="label">Solicitante:</td>
                <td>{cab_info.get('Solicitante', 'N/A')}</td>
            </tr>
            <tr>
                <td class="label">Requisición:</td>
                <td>{cab_info.get('Requisicion', 'N/A')}</td>
            </tr>
            <tr>
                <td class="label">Destino (L.A.B.):</td>
                <td>{cab_info.get('Destino', 'N/A')}</td>
            </tr>
        </table>
        
        <h3>Resumen de Partidas y Dispersión por Fecha</h3>
        <table class="matrix-table">
            <thead>
                <tr>
                    <th style="background-color: #EC2024; color: #FFFFFF; font-weight: bold; padding: 6px; border: 1px solid #dcdcdc; text-align: center; width: 120px;">SKU</th>
                    <th style="background-color: #EC2024; color: #FFFFFF; font-weight: bold; padding: 6px; border: 1px solid #dcdcdc; text-align: center; width: 100px;">Imagen</th>
                    <th style="background-color: #EC2024; color: #FFFFFF; font-weight: bold; padding: 6px; border: 1px solid #dcdcdc; text-align: center; width: 180px;">Tarimas / Detalle</th>
                    <th style="background-color: #000000; color: #FFFFFF; font-weight: bold; padding: 6px; border: 1px solid #dcdcdc; text-align: center; width: 85px;">Total Requerido</th>
                    <th style="background-color: #2E7D32; color: #FFFFFF; font-weight: bold; padding: 6px; border: 1px solid #dcdcdc; text-align: center; width: 85px;">Total Entregado</th>
                    <th style="background-color: #FBC02D; color: #000000; font-weight: bold; padding: 6px; border: 1px solid #dcdcdc; text-align: center; width: 85px;">Total Almacén</th>
                    <th style="background-color: #C62828; color: #FFFFFF; font-weight: bold; padding: 6px; border: 1px solid #dcdcdc; text-align: center; width: 85px;">Total Faltante</th>
    """
    
    # Agregar cabeceras de fechas
    for d in fechas_columnas:
        html += f'<th style="background-color: #EC2024; color: #FFFFFF; font-weight: bold; padding: 6px; border: 1px solid #dcdcdc; text-align: center; width: 160px;">{d}</th>'
        
    html += """
                </tr>
            </thead>
            <tbody>
    """
    
    # Agregar filas
    for _, row in df_matrix.iterrows():
        sku = row["SKU"]
        is_summary = (sku == "📈 % AVANCE")
        
        row_class = ' class="summary-row"' if is_summary else ""
        html += f"<tr{row_class}>"
        html += f'<td style="padding: 6px; border: 1px solid #dcdcdc; text-align: center; font-weight: bold; font-size: 18px;">{sku}</td>'
        
        # Imagen
        if is_summary:
            html += '<td style="padding: 6px; border: 1px solid #dcdcdc; text-align: center;"></td>'
        else:
            # Buscar imagen local
            import glob
            import os
            img_filename = None
            matching_local = glob.glob(f"imagenes_articulos/{sku}*.*")
            if matching_local:
                img_filename = os.path.basename(matching_local[0])
                
            img_tag = ""
            if img_filename:
                import urllib.parse
                img_filename_encoded = urllib.parse.quote(img_filename)
                img_url = f"https://raw.githubusercontent.com/{REPO_OWNER}/{REPO_NAME}/{BRANCH}/imagenes_articulos/{img_filename_encoded}"
                img_tag = f'<img src="{img_url}" width="80" height="80" style="border-radius: 4px; border: 1px solid #ccc; max-width: 80px; height: auto; display: block; margin: 0 auto;">'
            html += f'<td style="padding: 6px; border: 1px solid #dcdcdc; text-align: center;">{img_tag}</td>'
            
        # Detalle Tarimas
        if is_summary:
            html += '<td style="padding: 6px; border: 1px solid #dcdcdc; text-align: center;"></td>'
        else:
            val_tarimas = row.get("Detalle Tarimas", "-")
            val_tarimas_html = str(val_tarimas).replace("\n", "<br>")
            html += f'<td style="padding: 6px; border: 1px solid #dcdcdc; text-align: left; font-size: 12px; line-height: 1.3;">{val_tarimas_html}</td>'
        
        # Columnas de totales
        if is_summary:
            html += f'<td style="padding: 6px; border: 1px solid #dcdcdc; text-align: center; font-size: 22px; font-weight: bold;">{row["Total Requerido"]}</td>'
            html += f'<td style="padding: 6px; border: 1px solid #dcdcdc; text-align: center; font-size: 22px; font-weight: bold;">{row["Total Entregado"]}</td>'
            html += f'<td style="padding: 6px; border: 1px solid #dcdcdc; text-align: center; font-size: 22px; font-weight: bold;">{row["Total Almacén"]}</td>'
            html += f'<td style="padding: 6px; border: 1px solid #dcdcdc; text-align: center; font-size: 22px; font-weight: bold;">{row["Total Faltante"]}</td>'
        else:
            html += f'<td style="padding: 6px; border: 1px solid #dcdcdc; text-align: center; background-color: #000000; color: #FFFFFF; font-weight: bold; font-size: 22px;">{row["Total Requerido"]}</td>'
            html += f'<td style="padding: 6px; border: 1px solid #dcdcdc; text-align: center; background-color: #2E7D32; color: #FFFFFF; font-weight: bold; font-size: 22px;">{row["Total Entregado"]}</td>'
            html += f'<td style="padding: 6px; border: 1px solid #dcdcdc; text-align: center; background-color: #FBC02D; color: #000000; font-weight: bold; font-size: 22px;">{row["Total Almacén"]}</td>'
            html += f'<td style="padding: 6px; border: 1px solid #dcdcdc; text-align: center; background-color: #C62828; color: #FFFFFF; font-weight: bold; font-size: 22px;">{row["Total Faltante"]}</td>'
            
        # Fechas
        for d in fechas_columnas:
            val = row.get(d, "-")
            html += f'<td style="padding: 6px; border: 1px solid #dcdcdc; text-align: center;">{val}</td>'
            
        html += "</tr>"
        
    html += """
            </tbody>
        </table>
        <p><i>Nota: Se anexa a este correo el reporte de Excel oficial con los datos completos de dispersión de entregas.</i></p>
        <br>
        <p>Saludos cordiales,<br><b>Sistema de Control de Remisiones y Metales</b></p>
    </body>
    </html>
    """
    return html

def generar_cuerpo_correo_html(list_selected_remisiones, df_det):
    # Obtener listado de imágenes desde GitHub una sola vez para no saturar la API
    github_items = []
    if "github_token" in st.secrets and st.secrets["github_token"]:
        try:
            GITHUB_TOKEN = st.secrets["github_token"]
            url_list = f"https://api.github.com/repos/{REPO_OWNER}/{REPO_NAME}/contents/imagenes_articulos?ref={BRANCH}"
            headers = {"Authorization": f"token {GITHUB_TOKEN}", "Accept": "application/vnd.github.v3+json"}
            res_list = requests.get(url_list, headers=headers)
            if res_list.status_code == 200:
                github_items = res_list.json()
        except Exception:
            pass

    # Generar tabla HTML de piezas
    filas_html = ""
    for idx_row, row in enumerate(df_det.iterrows()):
        _, row_data = row
        sku = row_data.get("SKU", "N/A")
        desc = row_data.get("Descripcion", "N/A")
        cant = row_data.get("Cantidad", 0)
        po = row_data.get("PO", "N/A")
        id_tarima = row_data.get("ID_Tarima", "N/A")
        
        # Intentar obtener la URL de la imagen de GitHub o dejar texto
        img_tag = "N/A"
        if "BD_Articulos" in st.session_state and not st.session_state.BD_Articulos.empty:
            match = st.session_state.BD_Articulos[st.session_state.BD_Articulos['SKU'] == sku]
            if not match.empty:
                for git_item in github_items:
                    if git_item.get("name", "").startswith(f"{sku}("):
                        # URL raw de github
                        img_url = f"https://raw.githubusercontent.com/{REPO_OWNER}/{REPO_NAME}/{BRANCH}/imagenes_articulos/{git_item['name']}"
                        img_tag = f'<img src="{img_url}" width="60" style="border-radius:4px; border: 1px solid #ccc;">'
                        break

        bg_color = "#ffffff" if idx_row % 2 == 0 else "#f9f9f9"
        filas_html += f"""
        <tr style="background-color: {bg_color};">
            <td style="padding: 10px; border: 1px solid #ddd; font-size: 13px;">{sku}</td>
            <td style="padding: 10px; border: 1px solid #ddd; font-size: 13px;">{desc}</td>
            <td style="padding: 10px; border: 1px solid #ddd; text-align: center; font-size: 13px; font-weight: bold;">{int(cant) if isinstance(cant, (int, float)) else cant}</td>
            <td style="padding: 10px; border: 1px solid #ddd; text-align: center; font-size: 13px;">{id_tarima}</td>
            <td style="padding: 10px; border: 1px solid #ddd; font-size: 13px;">{po}</td>
            <td style="padding: 10px; border: 1px solid #ddd; text-align: center;">{img_tag}</td>
        </tr>
        """

    headers_text = ""
    for r in list_selected_remisiones:
        import ast
        f = r.get('Folio_Remision', 'N/A')
        rec = r.get('Nombre_Receptor', 'N/A')
        dir_rec = r.get('Direccion_Receptor', 'N/A')
        fec = r.get('Fecha_Hora_Salida', 'N/A')
        
        # Calcular total de piezas de esta remisión específica
        t_asoc = r.get('Tarimas_Asociadas', [])
        if isinstance(t_asoc, str):
            try:
                t_asoc_list = ast.literal_eval(t_asoc)
            except Exception:
                t_asoc_list = [t_asoc]
        else:
            t_asoc_list = t_asoc if isinstance(t_asoc, list) else []
            
        df_rem_det = df_det[df_det['ID_Tarima'].isin(t_asoc_list)]
        total_pzs = int(df_rem_det['Cantidad'].sum())
        
        headers_text += f"<li><b>PLANTA METALES DIAGONAL / {rec}</b> ({dir_rec}) &mdash; Fecha de Salida: {fec} (Remisión: <b>{f}</b> &mdash; <span style='color: #0056b3; font-weight: bold;'>{total_pzs} pzs</span>)</li>"

    html = f"""
    <html>
    <head>
        <link href="https://fonts.googleapis.com/css2?family=Questrial&display=swap" rel="stylesheet">
    </head>
    <body style="font-family: 'Questrial', 'Segoe UI', Arial, sans-serif; color: #000000; line-height: 1.6; padding: 20px; background-color: #ffffff;">
        <!-- Encabezado con Logotipo Corporativo -->
        <div style="padding-bottom: 15px; border-bottom: 4px solid #EC2024; margin-bottom: 20px;">
            <img src="https://raw.githubusercontent.com/{REPO_OWNER}/{REPO_NAME}/{BRANCH}/logo_sigrama.png" width="160" alt="Industria Sigrama" style="display: block; border: 0;">
        </div>

        <p style="font-size: 15px; margin-bottom: 15px;">Buen día a todos,</p>
        <p style="font-size: 14px; margin-bottom: 15px;">El presente correo es para informar sobre la salida de producto terminado de <b>PLANTA METALES DIAGONAL</b> con la documentación correspondiente adjunta. En estos archivos se detalla la cantidad, número de parte, número de tarima y orden de compra para el conocimiento de todos.</p>
        
        <!-- Caja Informativa Formato Diagonal Corporativo -->
        <div style="border-left: 4px solid #EC2024; background-color: #F8F9FA; padding: 15px; margin: 20px 0; border-radius: 0 4px 4px 0;">
            <h4 style="margin: 0 0 10px 0; color: #EC2024; font-size: 15px; text-transform: uppercase; letter-spacing: 0.5px;">Detalles de Envío (Salida de Material)</h4>
            <ul style="margin: 0; padding-left: 20px; font-size: 13.5px; color: #333333; line-height: 1.5;">
                {headers_text}
            </ul>
        </div>

        <p style="font-size: 14px; margin-bottom: 15px;">A continuación, se detalla la lista de las piezas, cantidades e información asociada:</p>
        
        <!-- Tabla de Piezas Corporativa -->
        <table style="border-collapse: collapse; width: 100%; max-width: 900px; margin-top: 15px; margin-bottom: 25px; border: 1px solid #ddd; box-shadow: 0 1px 3px rgba(0,0,0,0.05);">
            <thead>
                <tr style="background-color: #EC2024; color: #ffffff;">
                    <th style="padding: 12px 10px; border: 1px solid #ddd; text-align: left; font-size: 13.5px; font-weight: bold;">Número de Parte (SKU)</th>
                    <th style="padding: 12px 10px; border: 1px solid #ddd; text-align: left; font-size: 13.5px; font-weight: bold;">Descripción</th>
                    <th style="padding: 12px 10px; border: 1px solid #ddd; text-align: center; font-size: 13.5px; font-weight: bold;">Cantidad</th>
                    <th style="padding: 12px 10px; border: 1px solid #ddd; text-align: center; font-size: 13.5px; font-weight: bold;">Tarima</th>
                    <th style="padding: 12px 10px; border: 1px solid #ddd; text-align: left; font-size: 13.5px; font-weight: bold;">Orden (PO)</th>
                    <th style="padding: 12px 10px; border: 1px solid #ddd; text-align: center; font-size: 13.5px; font-weight: bold;">Imagen</th>
                </tr>
            </thead>
            <tbody>
                {filas_html}
            </tbody>
        </table>
        
        <p style="font-size: 14px; margin-bottom: 5px;">Gracias y quedo al pendiente de sus comentarios.</p>
        
        <!-- Firma Corporativa -->
        <div style="margin-top: 40px; border-top: 1px solid #E0E0E0; padding-top: 20px;">
            <p style="font-style: italic; font-weight: bold; color: #EC2024; margin: 0 0 8px 0; font-size: 15px;">Ingeniería que da resultados!!</p>
            <p style="margin: 0; font-size: 13px; color: #000000; font-weight: bold;">Industria Sigrama S.A. de C.V.</p>
            <p style="margin: 3px 0; font-size: 12px; color: #555555;">Automatización y Control de Procesos</p>
            <p style="margin: 3px 0; font-size: 11px; color: #777777;">C. Juan Escutia No. 50, Col. Abastos, C.P. 27020, Torreón, Coah. | Tel: (871) 722 3132</p>
            <p style="margin: 3px 0; font-size: 11px; color: #777777;"><a href="https://www.sigrama.com.mx" style="color: #EC2024; text-decoration: none; font-weight: bold;">www.sigrama.com.mx</a></p>
        </div>

        <!-- Aviso de Privacidad Obligatorio -->
        <div style="margin-top: 25px; font-size: 9px; color: #888888; line-height: 1.35; font-style: italic; border-top: 1px dashed #DDDDDD; padding-top: 12px; text-align: justify;">
            <b>Aviso de confidencialidad:</b> Este correo electrónico y/o el material adjunto es para uso exclusivo de la persona o entidad a la que expresamente se le ha enviado, y puede contener información confidencial o material privilegiado. Si usted no es el destinatario legítimo del mismo, por favor repórtelo inmediatamente al remitente del correo y bórrelo. Cualquier revisión, retransmisión, difusión o cualquier otro uso de este correo, por personas o entidades distintas a las del destinatario legítimo, queda expresamente prohibido. Este correo electrónico no pretende ni debe ser considerado como constitutivo de ninguna relación legal, contractual o de otra índole similar.<br/><br/>
            Así mismo, los Datos Personales que usted proporcione a través de este medio, están debidamente tratados y protegidos conforme a la "Ley Federal de Protección de Datos Personales en Posesión de los Particulares" para más información lo invitamos a conocer nuestra "Política de Privacidad" disponible en <a href="https://www.sigrama.com.mx" style="color: #888888; text-decoration: underline;">https://www.sigrama.com.mx</a>.
        </div>
    </body>
    </html>
    """
    return html

def generar_pdf_etiqueta(t_imp):
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=letter, leftMargin=36, rightMargin=36, topMargin=90, bottomMargin=60)
    story_l, styles = [], getSampleStyleSheet()

    style_tarima_titulo = ParagraphStyle('T_Giga_Ind', parent=styles['Heading1'], fontName="Helvetica-Bold", fontSize=140, alignment=1, leading=150, textColor=colors.HexColor("#212121"))
    style_sub_titulo = ParagraphStyle('S_Giga_Ind', parent=styles['Normal'], fontName="Helvetica-Bold", fontSize=26, alignment=1, textColor=colors.HexColor("#D32F2F"))
    style_normal_bold = ParagraphStyle('N_Bold_Ind', parent=styles['Normal'], fontName="Helvetica-Bold", fontSize=11, leading=14)
    style_normal_text = ParagraphStyle('N_Text_Ind', parent=styles['Normal'], fontSize=11, leading=14)
    style_blanco_bold = ParagraphStyle('B_Bold_Ind', parent=styles['Normal'], fontName="Helvetica-Bold", textColor=colors.white, alignment=1, fontSize=10)

    det = st.session_state.BD_Detalle_Tarimas[st.session_state.BD_Detalle_Tarimas['ID_Tarima'] == t_imp]
    t_info = st.session_state.BD_Tarimas[st.session_state.BD_Tarimas['ID_Tarima'] == t_imp]

    op_nom = t_info.iloc[0]['Creado_Por'] if not t_info.empty else "N/A"
    fe_cre = t_info.iloc[0]['Fecha_Creacion'] if not t_info.empty else "N/A"

    # Buscar si alguna PO de la tarima tiene formato de color configurado
    color_bg = None
    color_fg = None
    texto_etiqueta = None
    
    if "BD_POs_Cabecera" in st.session_state and not st.session_state.BD_POs_Cabecera.empty:
        df_pos_cab = st.session_state.BD_POs_Cabecera
        pos_tarima = det['PO'].dropna().unique().tolist()
        for p in pos_tarima:
            p_upper = str(p).strip().upper()
            po_rows = df_pos_cab[df_pos_cab['PO'].astype(str).str.strip().str.upper() == p_upper]
            if not po_rows.empty:
                row = po_rows.iloc[0]
                if 'Color_Fondo' in df_pos_cab.columns and pd.notna(row.get('Color_Fondo')) and str(row.get('Color_Fondo')).strip():
                    color_bg = str(row['Color_Fondo']).strip()
                    color_fg = str(row.get('Color_Texto', '#FFFFFF')).strip()
                    texto_etiqueta = str(row.get('Texto_Etiqueta', '')).strip()
                    if not texto_etiqueta:
                        texto_etiqueta = p
                    break

    # HOJA 1: CARÁTULA DE IDENTIFICACIÓN
    story_l.append(Spacer(1, 1.2 * inch))
    story_l.append(Paragraph("TARIMA", style_sub_titulo))
    story_l.append(Spacer(1, 0.2 * inch))

    num_limpio = str(t_imp).split('-')[-1] if '-' in str(t_imp) else str(t_imp)
    story_l.append(Paragraph(f"#{num_limpio}", style_tarima_titulo))
    
    if color_bg and texto_etiqueta:
        story_l.append(Spacer(1, 0.6 * inch))
    else:
        story_l.append(Spacer(1, 1.5 * inch))

    tabla_base = Table([
        [Paragraph("CÓDIGO DE IDENTIFICACIÓN:", style_normal_bold), Paragraph(str(t_imp), style_normal_text)],
        [Paragraph("OPERADOR DE PLANTA:", style_normal_bold), Paragraph(str(op_nom), style_normal_text)],
        [Paragraph("FECHA DE EMISIÓN:", style_normal_bold), Paragraph(str(fe_cre), style_normal_text)]
    ], colWidths=[2.5 * inch, 5.0 * inch])
    tabla_base.setStyle(TableStyle([
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('LINEBELOW', (0,0), (-1,-1), 0.5, colors.HexColor("#E0E0E0")),
        ('BOTTOMPADDING', (0,0), (-1,-1), 6)
    ]))
    story_l.append(tabla_base)

    if color_bg and texto_etiqueta:
        story_l.append(Spacer(1, 0.3 * inch))
        style_color_tag = ParagraphStyle(
            'ColorTagText', 
            fontName="Helvetica-Bold", 
            fontSize=34, 
            leading=40, 
            alignment=1, 
            textColor=colors.HexColor(color_fg)
        )
        color_box_table = Table([[Paragraph(texto_etiqueta, style_color_tag)]], colWidths=[6.8 * inch], rowHeights=[1.0 * inch])
        color_box_table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,-1), colors.HexColor(color_bg)),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('GRID', (0,0), (-1,-1), 1.5, colors.HexColor("#EAB519")),
            ('BOTTOMPADDING', (0,0), (-1,-1), 10),
            ('TOPPADDING', (0,0), (-1,-1), 10)
        ]))
        story_l.append(color_box_table)

    story_l.append(PageBreak())

    # HOJA 2: DETALLE DE MATERIALES ASOCIADOS
    story_l.append(Spacer(1, 0.1 * inch))
    story_l.append(Paragraph(f"<b>DETALLE DE MATERIALES ASOCIADOS - CONTROL #{t_imp}</b>", styles['Heading2']))
    story_l.append(Spacer(1, 0.2 * inch))

    tabla_detalles = [[
        Paragraph("ORDEN (PO)", style_blanco_bold),
        Paragraph("SKU / PRODUCTO", style_blanco_bold),
        Paragraph("DESCRIPCIÓN COMERCIAL", style_blanco_bold),
        Paragraph("CANTIDAD", style_blanco_bold)
    ]]

    for _, item in det.iterrows():
        art = st.session_state.BD_Articulos[st.session_state.BD_Articulos['SKU'] == item['SKU']]
        art_nom = art.iloc[0]['Nombre'] if not art.empty else "Articulo No Registrado en BD Remisiones"
    
        sku_partida = item['SKU']
        # Buscar si existe una imagen cargada para este SKU
        import glob
        img_encontrada = None
        matching_imgs = glob.glob(f"imagenes_articulos/{sku_partida}(*.*")
        if matching_imgs:
            img_encontrada = matching_imgs[0]
        else:
            if "github_token" in st.secrets and st.secrets["github_token"]:
                try:
                    GITHUB_TOKEN = st.secrets["github_token"]
                    url_list = f"https://api.github.com/repos/{REPO_OWNER}/{REPO_NAME}/contents/imagenes_articulos?ref={BRANCH}"
                    headers = {"Authorization": f"token {GITHUB_TOKEN}", "Accept": "application/vnd.github.v3+json"}
                    res_list = requests.get(url_list, headers=headers)
                    if res_list.status_code == 200:
                        items_git = res_list.json()
                        for it in items_git:
                            if it["name"].startswith(f"{sku_partida}("):
                                github_file_path = f"imagenes_articulos/{it['name']}"
                                if descargar_imagen_desde_github(github_file_path):
                                    img_encontrada = github_file_path
                                    break
                except Exception:
                    pass

        desc_paragraph = Paragraph(str(art_nom), style_normal_text)
        if img_encontrada and os.path.exists(img_encontrada):
            from reportlab.platypus import Image as RLImage
            img_flowable = RLImage(img_encontrada, width=75, height=75, hAlign='LEFT')
            sub_t = Table([[img_flowable, desc_paragraph]], colWidths=[80, 3.5 * inch - 80])
            sub_t.setStyle(TableStyle([
                ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
                ('LEFTPADDING', (0,0), (-1,-1), 0),
                ('RIGHTPADDING', (0,0), (-1,-1), 0),
                ('TOPPADDING', (0,0), (-1,-1), 0),
                ('BOTTOMPADDING', (0,0), (-1,-1), 0)
            ]))
            desc_comercial_flowables = sub_t
        else:
            desc_comercial_flowables = desc_paragraph

        tabla_detalles.append([
            Paragraph(str(item['PO']), style_normal_text),
            Paragraph(str(item['SKU']), style_normal_text),
            desc_comercial_flowables,
            Paragraph(f"<b>{int(item['Cantidad'])}</b> PZS", style_normal_bold)
        ])
    
    t_grid = Table(tabla_detalles, colWidths=[1.3 * inch, 1.5 * inch, 3.5 * inch, 1.2 * inch])
    t_grid.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#757575")),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor("#BDBDBD")),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('TOPPADDING', (0,0), (-1,-1), 6),
        ('BOTTOMPADDING', (0,0), (-1,-1), 6)
    ]))
    story_l.append(t_grid)

    doc.build(story_l, onFirstPage=draw_sigrama_decorations, onLaterPages=draw_sigrama_decorations)
    buf.seek(0)
    return buf

# =============================================================================
# 3. CAPA DE INICIALIZACIÓN GLOBAL SECTORIZADA (BLINDAJE DE SEGURIDAD)
# =============================================================================

# --- Estado de Inicio de Sesión ---
if "logged_in" not in st.session_state:
    st.session_state.logged_in = False
if "rol" not in st.session_state:
    st.session_state.rol = None
if "usuario_actual" not in st.session_state:
    st.session_state.usuario_actual = None

# --- Catálogo Maestro de Artículos (Actualizado para Persistencia en GitHub) ---
if "BD_Articulos" not in st.session_state or st.session_state.get("BD_Articulos") is None:
    df_git_articulos = cargar_excel_desde_github("BD_Articulos.xlsx")
    if df_git_articulos is not None:
        if "SKU" in df_git_articulos.columns:
            df_git_articulos["SKU"] = df_git_articulos["SKU"].astype(str).str.strip().str.upper()
        st.session_state.BD_Articulos = df_git_articulos
    else:
        # Estructura oficial estricta del sistema en caso de que el archivo no exista aún en GitHub
        st.session_state.BD_Articulos = pd.DataFrame(columns=["SKU", "Nombre", "Calibre_Espesor", "Dimensiones_Pieza", "Acabado_Superficial"])


# --- Lista de SKUs Autorizados (Base de Datos de Nombres SKU) ---
if "BD_SKUs_Autorizados" not in st.session_state or st.session_state.get("BD_SKUs_Autorizados") is None:
    df_git_skus_aut = cargar_excel_desde_github("BD_SKUs_Autorizados.xlsx")
    if df_git_skus_aut is not None:
        if "SKU" in df_git_skus_aut.columns:
            df_git_skus_aut["SKU"] = df_git_skus_aut["SKU"].astype(str).str.strip().str.upper()
        st.session_state.BD_SKUs_Autorizados = df_git_skus_aut
    else:
        st.session_state.BD_SKUs_Autorizados = pd.DataFrame(columns=["SKU"])


# --- Catálogo Maestro de Tarimas ---
if "BD_Tarimas" not in st.session_state:
    df_git_tarimas = cargar_excel_desde_github("BD_Tarimas.xlsx")
    if df_git_tarimas is not None:
        st.session_state.BD_Tarimas = df_git_tarimas
    else:
        st.session_state.BD_Tarimas = pd.DataFrame(columns=["ID_Tarima", "Tarima_Origen_Excel", "Fecha_Creacion", "Ubicacion_Actual", "Creado_Por", "Tipo_Tarima", "Estatus", "Es_Nueva"])

# --- Detalle Granular de Contenido por Tarima ---
if "BD_Detalle_Tarimas" not in st.session_state or st.session_state.get("BD_Detalle_Tarimas") is None:
    df_git_detalles = cargar_excel_desde_github("BD_Detalle_Tarimas.xlsx")
    if df_git_detalles is not None:
        if "SKU" in df_git_detalles.columns:
            df_git_detalles["SKU"] = df_git_detalles["SKU"].astype(str).str.strip().str.upper()
        st.session_state.BD_Detalle_Tarimas = df_git_detalles
    else:
        st.session_state.BD_Detalle_Tarimas = pd.DataFrame(columns=["ID_Detalle", "ID_Tarima", "SKU", "PO", "Proyecto", "Parcialidad", "Descripcion", "Cantidad"])

# --- Datos Históricos de Remisiones Oficiales ---
if "BD_Datos_Generales_Remision" not in st.session_state:
    df_git_remisiones = cargar_excel_desde_github("BD_Datos_Generales_Remision.xlsx")
    if df_git_remisiones is not None:
        if "Tarimas_Asociadas" in df_git_remisiones.columns:
            df_git_remisiones["Tarimas_Asociadas"] = df_git_remisiones["Tarimas_Asociadas"].astype(str)
        st.session_state.BD_Datos_Generales_Remision = df_git_remisiones
    else:
        st.session_state.BD_Datos_Generales_Remision = pd.DataFrame(columns=["ID_Remision", "Folio_Remision", "Fecha_Hora_Salida", "Nombre_Emisor", "Direccion_Emisor", "Nombre_Receptor", "Direccion_Receptor", "Tarimas_Asociadas"])

# --- Cabecera de Órdenes de Compra (POs) ---
if "BD_POs_Cabecera" not in st.session_state or st.session_state.get("BD_POs_Cabecera") is None:
    df_git_po_cab = cargar_excel_desde_github("BD_POs_Cabecera.xlsx")
    if df_git_po_cab is not None:
        if "PO" in df_git_po_cab.columns:
            df_git_po_cab["PO"] = df_git_po_cab["PO"].astype(str).str.strip().str.upper()
        st.session_state.BD_POs_Cabecera = df_git_po_cab
    else:
        st.session_state.BD_POs_Cabecera = pd.DataFrame(columns=["PO", "Fecha_Pedido", "Proyecto", "Solicitante", "Requisicion", "Destino"])

# --- Detalle de Requerimientos y Fechas de Entrega de POs ---
if "BD_Requerimientos_POs" not in st.session_state or st.session_state.get("BD_Requerimientos_POs") is None:
    df_git_po_req = cargar_excel_desde_github("BD_Requerimientos_POs.xlsx")
    if df_git_po_req is not None:
        if "PO" in df_git_po_req.columns:
            df_git_po_req["PO"] = df_git_po_req["PO"].astype(str).str.strip().str.upper()
        if "SKU" in df_git_po_req.columns:
            df_git_po_req["SKU"] = df_git_po_req["SKU"].astype(str).str.strip().str.upper()
        st.session_state.BD_Requerimientos_POs = df_git_po_req
    else:
        st.session_state.BD_Requerimientos_POs = pd.DataFrame(columns=["PO", "SKU", "Fecha_Entrega", "Cantidad_Requerida", "Parcialidad"])

def sincronizar_estatus_tarimas(auto_save=True):
    """Compara las tarimas remesadas con las remisiones activas y corrige huérfanas."""
    if "BD_Tarimas" not in st.session_state or "BD_Datos_Generales_Remision" not in st.session_state:
        return 0
    
    df_tarimas = st.session_state.BD_Tarimas
    df_remisiones = st.session_state.BD_Datos_Generales_Remision
    
    if df_tarimas.empty:
        return 0
        
    import ast
    remesadas_activas = set()
    if not df_remisiones.empty:
        for _, row in df_remisiones.iterrows():
            val = row.get('Tarimas_Asociadas')
            if isinstance(val, str):
                try:
                    t_list = ast.literal_eval(val)
                except Exception:
                    t_list = [val]
            elif isinstance(val, list):
                t_list = val
            else:
                t_list = []
            for t in t_list:
                remesadas_activas.add(str(t).strip())
                
    corregidas = 0
    for idx, row in df_tarimas.iterrows():
        t_id = str(row['ID_Tarima']).strip()
        est = str(row['Estatus']).strip()
        
        # Si está marcada como Remesada pero no está en ninguna remisión activa, vuelve a Disponible
        if est == 'Remesada' and t_id not in remesadas_activas:
            df_tarimas.at[idx, 'Estatus'] = 'Disponible'
            corregidas += 1
        # Si está marcada como Disponible pero sí está en una remisión activa, se corrige a Remesada
        elif est == 'Disponible' and t_id in remesadas_activas:
            df_tarimas.at[idx, 'Estatus'] = 'Remesada'
            corregidas += 1
            
    if corregidas > 0:
        st.session_state.BD_Tarimas = df_tarimas
        if auto_save:
            subir_excel_a_github("BD_Tarimas.xlsx", df_tarimas)
            
    return corregidas

# --- Corrección automática de Estatus de Tarimas ---
if "BD_Tarimas" in st.session_state and "BD_Datos_Generales_Remision" in st.session_state:
    try:
        sincronizar_estatus_tarimas(auto_save=True)
    except Exception:
        pass

# --- Catálogo Operativo de Líderes ---
if "BD_Lideres" not in st.session_state:
    df_git_lideres = cargar_excel_desde_github("BD_Lideres.xlsx")
    if df_git_lideres is not None:
        st.session_state.BD_Lideres = df_git_lideres
    else:
        st.session_state.BD_Lideres = pd.DataFrame([{"ID_Lider": "LID-01", "Nombre_Lider": "Jesus Morales", "Area": "Metales", "Estatus": "Activo"}])

# --- Catálogo de Receptores / Destinos ---
if "BD_Receptores" not in st.session_state:
    df_git_receptores = cargar_excel_desde_github("BD_Receptores.xlsx")
    if df_git_receptores is not None:
        st.session_state.BD_Receptores = df_git_receptores
    else:
        st.session_state.BD_Receptores = pd.DataFrame([
            {"ID_Receptor": "REC-01", "Nombre_Receptor": "Galvatec Industrias", "Direccion": "Prol. Valle Guadiana 919, Parque Industrial II, 35078 Gómez Palacio, Dgo.", "Estatus": "Activo"},
            {"ID_Receptor": "REC-02", "Nombre_Receptor": "EQM", "Direccion": "Calle Inde 714, 35079 Gómez Palacio, Durango", "Estatus": "Activo"}
        ])

# --- Migración automática: Renombrar COVISA → EQM en todas las bases de datos ---
# Se ejecuta en cada arranque y verifica si hay algo que corregir
bds_a_migrar = {
    "BD_Receptores": "BD_Receptores.xlsx",
    "BD_Datos_Generales_Remision": "BD_Datos_Generales_Remision.xlsx",
    "BD_Tarimas": "BD_Tarimas.xlsx"
}

for bd_key, bd_file in bds_a_migrar.items():
    if bd_key in st.session_state and not st.session_state[bd_key].empty:
        df = st.session_state[bd_key]
        hubo_cambio_en_bd = False
        for col in df.columns:
            if df[col].dtype == object:
                original = df[col].copy()
                df[col] = df[col].apply(lambda x: str(x).replace("COVISA", "EQM").replace("Covisa", "EQM").replace("covisa", "EQM") if pd.notna(x) else x)
                if not original.equals(df[col]):
                    hubo_cambio_en_bd = True
        if hubo_cambio_en_bd:
            st.session_state[bd_key] = df
            subir_excel_a_github(bd_file, st.session_state[bd_key])

# --- Función Auxiliar y Consecutivo Dinámico de Tarimas (TPM) ---
def obtener_siguiente_consecutivo_tpm():
    # 1. Intentar cargar el consecutivo manual configurado desde disco local
    import os
    if os.path.exists("consecutivo_override.txt"):
        try:
            with open("consecutivo_override.txt", "r", encoding="utf-8") as f:
                val = f.read().strip()
                if val.isdigit():
                    return int(val)
        except Exception:
            pass

    # 2. Calcular dinámicamente a partir de BD_Tarimas y BD_Datos_Generales_Remision
    numeros = []
    
    # Escanear BD_Tarimas
    if "BD_Tarimas" in st.session_state and not st.session_state.BD_Tarimas.empty:
        try:
            ids = st.session_state.BD_Tarimas['ID_Tarima'].astype(str)
            for id_val in ids:
                partes = id_val.split('-')
                if len(partes) > 1 and partes[1].strip().isdigit():
                    numeros.append(int(partes[1].strip()))
                elif id_val.startswith('TPM') and id_val[3:].strip().isdigit():
                    numeros.append(int(id_val[3:].strip()))
        except Exception:
            pass
            
    # Escanear BD_Datos_Generales_Remision (para evitar colisiones con remisiones existentes)
    if "BD_Datos_Generales_Remision" in st.session_state and not st.session_state.BD_Datos_Generales_Remision.empty:
        try:
            import ast
            for _, row in st.session_state.BD_Datos_Generales_Remision.iterrows():
                val = row.get('Tarimas_Asociadas')
                if isinstance(val, str):
                    try:
                        t_list = ast.literal_eval(val)
                    except Exception:
                        t_list = [val]
                elif isinstance(val, list):
                    t_list = val
                else:
                    t_list = []
                    
                for t in t_list:
                    t_str = str(t).strip()
                    partes = t_str.split('-')
                    if len(partes) > 1 and partes[1].strip().isdigit():
                        numeros.append(int(partes[1].strip()))
                    elif t_str.startswith('TPM') and t_str[3:].strip().isdigit():
                        numeros.append(int(t_str[3:].strip()))
        except Exception:
            pass
            
    if numeros:
        return max(numeros) + 1
    return 1

if "siguiente_numero_tpm" not in st.session_state or st.session_state["siguiente_numero_tpm"] is None:
    st.session_state["siguiente_numero_tpm"] = obtener_siguiente_consecutivo_tpm()





# =============================================================================
# CAPA DE CONTROL DOCUMENTAL Y DISEÑO IMPRESO CORPORATIVO (INDUSTRIA SIGRAMA)
# =============================================================================
def draw_sigrama_reporte_decorations(canvas, doc):
    """Dibuja los elementos de marca institucionales y diseño de encabezado corporativo del Reporte Consolidado."""
    canvas.saveState()
    # Franja superior roja Sigrama
    canvas.setFillColor(colors.HexColor("#D32F2F"))
    canvas.rect(36, 745, 540, 4, fill=1, stroke=0)
    
    # Marcador de Control de Calidad y Metadatos de Revisión - Eliminados por requerimiento del cliente
    
    # --- CAMBIO AQUÍ: Fecha más grande, destacada en negrita y desplazada hacia abajo ---
    canvas.setFont("Helvetica-Bold", 10) 
    canvas.drawString(36, 730, datetime.date.today().strftime("%d de %B %Y"))
    
    # Título Central del Formato Oficial
    canvas.setFont("Helvetica-Bold", 13)
    canvas.drawCentredString(285, 755, "REPORTE CONSOLIDADO DE INVENTARIO POR FILTRO")
    
    # Pie de Página Legal y Control del SGC (FO-SGC-02) - Eliminado por requerimiento del cliente
    canvas.restoreState()

def generar_pdf_reporte_filtrado(filtros_dict, df_resultado_piezas):
    """Construye el documento PDF oficial de Reporte Consolidado de Inventario con el panel de filtros y la cuadrícula."""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, leftMargin=36, rightMargin=36, topMargin=90, bottomMargin=60)
    story, styles = [], getSampleStyleSheet()
    
    # Estilos de texto especializados para tablas y párrafos
    style_blanco_bold = ParagraphStyle('BB_Rep', parent=styles['Normal'], textColor=colors.white, fontName="Helvetica-Bold", alignment=1, fontSize=9)
    style_normal_bold = ParagraphStyle('NB_Rep', parent=styles['Normal'], fontName="Helvetica-Bold", fontSize=8)
    style_normal_text = ParagraphStyle('NT_Rep', parent=styles['Normal'], fontSize=8)
    
    story.append(Spacer(1, 0.1 * inch))
    
    # --- PANEL LOGÍSTICO DE FILTROS APLICADOS ---
    t_header_filtros = Table([[Paragraph("CRITERIOS DE FILTRADO ACTIVO EN CONSULTA", style_blanco_bold)]], colWidths=[7.5 * inch])
    t_header_filtros.setStyle(TableStyle([('BACKGROUND', (0,0), (-1,-1), colors.HexColor("#757575")), ('ALIGN', (0,0), (-1,-1), 'CENTER')]))
    story.append(t_header_filtros)
    
    datos_panel_filtros = [
        [Paragraph("ORDEN DE COMPRA (PO):", style_normal_bold), Paragraph(str(filtros_dict['PO']), style_normal_text),
         Paragraph("SKU / PRODUCTO:", style_normal_bold), Paragraph(str(filtros_dict['SKU']), style_normal_text)],
        [Paragraph("PROYECTO INTERNO:", style_normal_bold), Paragraph(str(filtros_dict['Proyecto']), style_normal_text),
         Paragraph("ID TARIMA:", style_normal_bold), Paragraph(str(filtros_dict['Tarima']), style_normal_text)],
        [Paragraph("PARCIALIDAD:", style_normal_bold), Paragraph(str(filtros_dict['Parcialidad']), style_normal_text),
         Paragraph("ESTATUS ENVÍO:", style_normal_bold), Paragraph(str(filtros_dict['Estatus']), style_normal_text)],
        [Paragraph("DESCRIPCIÓN DE PROYECTO:", style_normal_bold), Paragraph(str(filtros_dict['Descripcion']), style_normal_text),
         Paragraph("TOTAL PIEZAS EN SELECCIÓN:", style_normal_bold), Paragraph(f"<b>{int(df_resultado_piezas['Cantidad'].sum()):,} PZS</b>", style_normal_bold)]
    ]
    
    t_panel = Table(datos_panel_filtros, colWidths=[2.0 * inch, 1.75 * inch, 2.0 * inch, 1.75 * inch])
    t_panel.setStyle(TableStyle([
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor("#BDBDBD")),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('BACKGROUND', (0,0), (0,-1), colors.HexColor("#F5F5F5")),
        ('BACKGROUND', (2,0), (2,-1), colors.HexColor("#F5F5F5"))
    ]))
    story.append(t_panel)
    story.append(Spacer(1, 0.2 * inch))
    
    # --- CUADRÍCULA FORMAL DE MATERIALES FILTRADOS ---
    tabla_materiales = [[
        Paragraph("ID TARIMA", style_blanco_bold),
        Paragraph("OC (PO)", style_blanco_bold),
        Paragraph("PROYECTO", style_blanco_bold),
        Paragraph("SKU / PRODUCTO", style_blanco_bold),
        Paragraph("CANTIDAD", style_blanco_bold),
        Paragraph("ESTATUS", style_blanco_bold)
    ]]
    
    # Mapeo de tarimas a su información de remisión para enriquecer el Estatus
    mapa_remisiones = {}
    if "BD_Datos_Generales_Remision" in st.session_state and not st.session_state.BD_Datos_Generales_Remision.empty:
        import ast
        for _, rem_row in st.session_state.BD_Datos_Generales_Remision.iterrows():
            t_val = rem_row['Tarimas_Asociadas']
            receptor = rem_row.get('Nombre_Receptor', 'Desconocido')
            fecha_salida = rem_row.get('Fecha_Hora_Salida', '')
            t_list = []
            if isinstance(t_val, str):
                try:
                    t_list = ast.literal_eval(t_val)
                except Exception:
                    t_list = [t_val]
            elif isinstance(t_val, list):
                t_list = t_val
                
            for t in t_list:
                mapa_remisiones[str(t).strip()] = {"Receptor": receptor, "Fecha": fecha_salida}
    
    for _, row in df_resultado_piezas.iterrows():
    
        sku_actual = row.get('SKU', '')
        descripcion_final = "Articulo No Registrado en BD Remisiones"
    
        # --- SOLUCIÓN DE BUG: Inicializar variables para evitar UnboundLocalError ---
        calibre = ""
        dims = ""
        acabado = ""
        nombre_com = ""
    
        if "BD_Articulos" in st.session_state and not st.session_state.BD_Articulos.empty:
            df_match = st.session_state.BD_Articulos[st.session_state.BD_Articulos['SKU'] == sku_actual]
            if not df_match.empty:
                art_info = df_match.iloc[0]
                nombre_com = str(art_info.get('Nombre', '')).strip()
                calibre = str(art_info.get('Calibre_Espesor', '')).strip()
                dims = str(art_info.get('Dimensiones_Pieza', '')).strip()
                acabado = str(art_info.get('Acabado_Superficial', '')).strip()
    
                detalles = []
                if calibre and calibre.lower() != 'nan' and calibre != '': detalles.append(f"<b>Calibre/Espesor:</b> {calibre}")
                if dims and dims.lower() != 'nan' and dims != '': detalles.append(f"<b>Dimensiones:</b> {dims}")
                if acabado and acabado.lower() != 'nan' and acabado != '': detalles.append(f"<b>Material/Acabado:</b> {acabado}")
    
                espec_str = f" | ".join(detalles)
                if espec_str:
                    descripcion_final = f"<b>{nombre_com}</b><br/><font color='#555555' size='7.5'>{espec_str}</font>"
                else:
                    descripcion_final = f"<b>{nombre_com}</b>"
            else:
                descripcion_final = "Articulo No Registrado en BD Remisiones"
        else:
            descripcion_final = "Articulo No Registrado en BD Remisiones"





        
        
        # Buscar si existe una imagen cargada para este SKU
        import glob
        img_encontrada = None
        matching_imgs = glob.glob(f"imagenes_articulos/{sku_actual}(*.*")
        if matching_imgs:
            img_encontrada = matching_imgs[0]
        else:
            if "github_token" in st.secrets and st.secrets["github_token"]:
                try:
                    GITHUB_TOKEN = st.secrets["github_token"]
                    url_list = f"https://api.github.com/repos/{REPO_OWNER}/{REPO_NAME}/contents/imagenes_articulos?ref={BRANCH}"
                    headers = {"Authorization": f"token {GITHUB_TOKEN}", "Accept": "application/vnd.github.v3+json"}
                    res_list = requests.get(url_list, headers=headers)
                    if res_list.status_code == 200:
                        items_git = res_list.json()
                        for it in items_git:
                            if it["name"].startswith(f"{sku_actual}("):
                                github_file_path = f"imagenes_articulos/{it['name']}"
                                if descargar_imagen_desde_github(github_file_path):
                                    img_encontrada = github_file_path
                                    break
                except Exception:
                    pass

        desc_paragraph = Paragraph(f"{row['SKU']}<br/><font color='#616161'>{descripcion_final}</font>", style_normal_text)
        if img_encontrada and os.path.exists(img_encontrada):
            from reportlab.platypus import Image as RLImage
            img_flowable = RLImage(img_encontrada, width=75, height=75, hAlign='LEFT')
            sub_t = Table([[img_flowable, desc_paragraph]], colWidths=[80, 2.3 * inch - 80])
            sub_t.setStyle(TableStyle([
                ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
                ('LEFTPADDING', (0,0), (-1,-1), 0),
                ('RIGHTPADDING', (0,0), (-1,-1), 0),
                ('TOPPADDING', (0,0), (-1,-1), 0),
                ('BOTTOMPADDING', (0,0), (-1,-1), 0)
            ]))
            desc_cell_flowables = sub_t
        else:
            desc_cell_flowables = desc_paragraph

        # Generar texto de estatus extendido
        estatus_texto = str(row['Estatus_Envio'])
        if estatus_texto.lower() == "remesado":
            t_id = str(row['ID_Tarima']).strip()
            if t_id in mapa_remisiones:
                receptor_info = str(mapa_remisiones[t_id]["Receptor"])
                fecha_info = str(mapa_remisiones[t_id]["Fecha"]).split()[0] if mapa_remisiones[t_id]["Fecha"] else ""
                # Si el receptor es muy largo, cortarlo para que no desborde la celda
                if len(receptor_info) > 18:
                    receptor_info = receptor_info[:15] + "..."
                estatus_texto = f"<b>Remesado</b><br/><font color='#555555' size='7'>{receptor_info}<br/>{fecha_info}</font>"

        tabla_materiales.append([
            Paragraph(str(row['ID_Tarima']), style_normal_text),
            Paragraph(str(row['PO']), style_normal_text),
            Paragraph(str(row['Proyecto']), style_normal_text),
            desc_cell_flowables,
            Paragraph(f"<b>{int(row['Cantidad'])}</b> Pzs", style_normal_text),
            Paragraph(estatus_texto, style_normal_text)
        ])
        
    t_mat = Table(tabla_materiales, colWidths=[1.0 * inch, 1.1 * inch, 1.2 * inch, 2.3 * inch, 0.9 * inch, 1.0 * inch])
    t_mat.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#D32F2F")),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor("#757575")),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE')
    ]))
    story.append(t_mat)
    
    # CORRECTO
    doc.build(story, onFirstPage=draw_sigrama_reporte_decorations, onLaterPages=draw_sigrama_reporte_decorations)


    buffer.seek(0)
    return buffer



# =============================================================================
# DECORADORES BASE CORPORATIVOS PARA DOCUMENTOS DE EMBARQUE
# =============================================================================

def draw_sigrama_decorations(canvas, doc):
    """Dibuja los elementos de marca institucionales y diseño de encabezado corporativo del documento de Embarque."""
    canvas.saveState()
    # Franja superior roja Sigrama
    canvas.setFillColor(colors.HexColor("#D32F2F"))
    canvas.rect(36, 745, 540, 4, fill=1, stroke=0)
    
    # Marcador de Control de Calidad y Metadatos de Revisión - Eliminados por requerimiento del cliente
    
    # --- CAMBIO SOLICITADO: Fecha más grande (font 10) y desplazada hacia abajo (coordenada 730) ---
    canvas.setFont("Helvetica-Bold", 10)
    canvas.drawString(36, 730, datetime.date.today().strftime("%d de %B %Y"))
    
    # Título Central del Formato Oficial
    canvas.setFont("Helvetica-Bold", 13)
    canvas.drawCentredString(285, 755, "EMBARQUE-RECEPCIÓN DE MERCANCÍA")
    
    # Pie de Página Legal y Control del SGC (FO-SGC-02) - Eliminado por requerimiento del cliente
    canvas.restoreState()


# =============================================================================
# FUNCIÓN DE REMISIÓN GENERAL (EMBARQUE-RECEPCIÓN DE MERCANCÍA)
# =============================================================================
def generar_pdf_remision_general(datos_remision, df_detalles_remision):
    """Construye el documento oficial de despacho rellenando las páginas de ReportLab con datos válidos."""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, leftMargin=36, rightMargin=36, topMargin=95, bottomMargin=60)
    story, styles = [], getSampleStyleSheet()
    
    # Estilos tipográficos especializados
    style_blanco_bold = ParagraphStyle('WB_Rem', parent=styles['Normal'], textColor=colors.white, fontName="Helvetica-Bold", alignment=1, fontSize=9)
    style_normal_bold = ParagraphStyle('NB_Rem', parent=styles['Normal'], fontName="Helvetica-Bold", fontSize=8)
    style_normal_text = ParagraphStyle('NT_Rem', parent=styles['Normal'], fontSize=8)
    
    story.append(Spacer(1, 0.1 * inch))
    
    # --- PANEL 1: ENCABEZADO LOGÍSTICO OPERATIVO ---
    t_header = Table([[Paragraph("DATOS GENERALES DE EMBARQUE Y DESPACHO", style_blanco_bold)]], colWidths=[7.5 * inch])
    t_header.setStyle(TableStyle([('BACKGROUND', (0,0), (-1,-1), colors.HexColor("#757575")), ('ALIGN', (0,0), (-1,-1), 'CENTER')]))
    story.append(t_header)
    
    datos_panel = [
        [Paragraph("FOLIO REMISIÓN:", style_normal_bold), Paragraph(f"<b>{str(datos_remision['Folio_Remision'])}</b>", style_normal_bold),
         Paragraph("FECHA EMISIÓN:", style_normal_bold), Paragraph(str(datos_remision['Fecha_Hora_Salida']), style_normal_text)],
        [Paragraph("EMISOR / ALMACÉN:", style_normal_bold), Paragraph(str(datos_remision['Nombre_Emisor']), style_normal_text),
         Paragraph("ORIGEN:", style_normal_bold), Paragraph(str(datos_remision['Direccion_Emisor']), style_normal_text)],
        [Paragraph("RECEPTOR / CLIENTE:", style_normal_bold), Paragraph(str(datos_remision['Nombre_Receptor']), style_normal_text),
         Paragraph("DESTINO PLANTA:", style_normal_bold), Paragraph(str(datos_remision['Direccion_Receptor']), style_normal_text)]
    ]
    t_panel = Table(datos_panel, colWidths=[1.8 * inch, 1.95 * inch, 1.8 * inch, 1.95 * inch])
    t_panel.setStyle(TableStyle([
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor("#BDBDBD")),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('BACKGROUND', (0,0), (0,-1), colors.HexColor("#F5F5F5")),
        ('BACKGROUND', (2,0), (2,-1), colors.HexColor("#F5F5F5"))
    ]))
    story.append(t_panel)
    story.append(Spacer(1, 0.2 * inch))
    
    # --- PANEL 2: TABLA DE MATERIALES ASOCIADOS EN DESPACHO ---
    tabla_materiales = [[
        Paragraph("ID TARIMA", style_blanco_bold),
        Paragraph("ORDEN COMPRA", style_blanco_bold),
        Paragraph("PROYECTO", style_blanco_bold),
        Paragraph("SKU / PRODUCTO", style_blanco_bold),
        Paragraph("CANTIDAD", style_blanco_bold)
    ]]
    
    for _, row in df_detalles_remision.iterrows():
        # Búsqueda segura del nombre comercial del artículo
        sku_partida = row.get('SKU', '')
        concepto_remision = "Material de Embarque"
        
        if "BD_Articulos" in st.session_state and not st.session_state.BD_Articulos.empty:
            df_match_rem = st.session_state.BD_Articulos[st.session_state.BD_Articulos['SKU'] == sku_partida]
            if not df_match_rem.empty:
                art_info = df_match_rem.iloc[0]
                nombre_com = str(art_info.get('Nombre', '')).strip()
                calibre = str(art_info.get('Calibre_Espesor', '')).strip()
                dims = str(art_info.get('Dimensiones_Pieza', '')).strip()
                acabado = str(art_info.get('Acabado_Superficial', '')).strip()
                
                detalles = []
                if calibre and calibre.lower() != 'nan' and calibre != '': detalles.append(f"<b>Calibre/Espesor:</b> {calibre}")
                if dims and dims.lower() != 'nan' and dims != '': detalles.append(f"<b>Dimensiones:</b> {dims}")
                if acabado and acabado.lower() != 'nan' and acabado != '': detalles.append(f"<b>Material/Acabado:</b> {acabado}")
                
                espec_str = f" | ".join(detalles)
                if espec_str:
                    concepto_remision = f"<b>{nombre_com}</b><br/><font color='#555555' size='7.5'>{espec_str}</font>"
                else:
                    concepto_remision = f"<b>{nombre_com}</b>"
            else:
                concepto_remision = "Articulo No Registrado en BD Remisiones"
        else:
            concepto_remision = "Articulo No Registrado en BD Remisiones"



        
        
        # Buscar si existe una imagen cargada para este SKU
        import glob
        img_encontrada = None
        matching_imgs = glob.glob(f"imagenes_articulos/{sku_partida}(*.*")
        if matching_imgs:
            img_encontrada = matching_imgs[0]
        else:
            if "github_token" in st.secrets and st.secrets["github_token"]:
                try:
                    GITHUB_TOKEN = st.secrets["github_token"]
                    url_list = f"https://api.github.com/repos/{REPO_OWNER}/{REPO_NAME}/contents/imagenes_articulos?ref={BRANCH}"
                    headers = {"Authorization": f"token {GITHUB_TOKEN}", "Accept": "application/vnd.github.v3+json"}
                    res_list = requests.get(url_list, headers=headers)
                    if res_list.status_code == 200:
                        items = res_list.json()
                        for item in items:
                            if item["name"].startswith(f"{sku_partida}("):
                                github_file_path = f"imagenes_articulos/{item['name']}"
                                if descargar_imagen_desde_github(github_file_path):
                                    img_encontrada = github_file_path
                                    break
                except Exception:
                    pass

        desc_paragraph = Paragraph(f"{row['SKU']}<br/><font color='#616161'>{concepto_remision}</font>", style_normal_text)
        if img_encontrada and os.path.exists(img_encontrada):
            from reportlab.platypus import Image as RLImage
            img_flowable = RLImage(img_encontrada, width=75, height=75, hAlign='LEFT')
            sub_t = Table([[img_flowable, desc_paragraph]], colWidths=[80, 2.7 * inch - 80])
            sub_t.setStyle(TableStyle([
                ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
                ('LEFTPADDING', (0,0), (-1,-1), 0),
                ('RIGHTPADDING', (0,0), (-1,-1), 0),
                ('TOPPADDING', (0,0), (-1,-1), 0),
                ('BOTTOMPADDING', (0,0), (-1,-1), 0)
            ]))
            desc_cell_flowables = sub_t
        else:
            desc_cell_flowables = desc_paragraph
            
        tabla_materiales.append([
            Paragraph(str(row['ID_Tarima']), style_normal_text),
            Paragraph(str(row['PO']), style_normal_text),
            Paragraph(str(row['Proyecto']), style_normal_text),
            desc_cell_flowables,
            Paragraph(f"<b>{int(row['Cantidad'])}</b> Pzs", style_normal_text)
        ])

        
    t_mat = Table(tabla_materiales, colWidths=[1.2 * inch, 1.3 * inch, 1.3 * inch, 2.7 * inch, 1.0 * inch])
    t_mat.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#D32F2F")),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor("#757575")),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE')
    ]))
    story.append(t_mat)
    
    # Compilación forzando la inyección de páginas en el objeto Canvas
    doc.build(story, onFirstPage=draw_sigrama_decorations, onLaterPages=draw_sigrama_decorations)
    buffer.seek(0)
    return buffer


def generar_excel_remision(datos_remision, df_detalles_remision):
    import io
    import pandas as pd
    from openpyxl.styles import Font, PatternFill, Alignment
    
    buffer = io.BytesIO()
    writer = pd.ExcelWriter(buffer, engine='openpyxl')
    
    # 1. Preparar metadatos generales
    df_meta = pd.DataFrame([
        {"CAMPO": "FOLIO REMISIÓN", "VALOR": str(datos_remision['Folio_Remision'])},
        {"CAMPO": "FECHA EMISIÓN", "VALOR": str(datos_remision['Fecha_Hora_Salida'])},
        {"CAMPO": "EMISOR / ALMACÉN", "VALOR": str(datos_remision['Nombre_Emisor'])},
        {"CAMPO": "ORIGEN", "VALOR": str(datos_remision['Direccion_Emisor'])},
        {"CAMPO": "RECEPTOR / CLIENTE", "VALOR": str(datos_remision['Nombre_Receptor'])},
        {"CAMPO": "DESTINO PLANTA", "VALOR": str(datos_remision['Direccion_Receptor'])}
    ])
    
    df_meta.to_excel(writer, index=False, sheet_name='Datos Generales', startrow=0)
    
    # 2. Preparar tabla de materiales
    df_mats = df_detalles_remision.copy()
    
    # Enriquecer con Nombre del Catálogo
    if "BD_Articulos" in st.session_state and not st.session_state.BD_Articulos.empty:
        df_art = st.session_state.BD_Articulos[['SKU', 'Nombre']].copy()
        df_art['SKU'] = df_art['SKU'].astype(str).str.strip()
        df_mats['SKU'] = df_mats['SKU'].astype(str).str.strip()
        df_mats = pd.merge(df_mats, df_art, on='SKU', how='left')
        df_mats['Nombre'] = df_mats['Nombre'].fillna("Articulo No Registrado")
    else:
        df_mats['Nombre'] = "Articulo No Registrado"
        
    df_export = df_mats[['ID_Tarima', 'PO', 'Proyecto', 'SKU', 'Nombre', 'Cantidad']].copy()
    df_export.columns = ["ID TARIMA", "ORDEN COMPRA (PO)", "PROYECTO", "SKU / PRODUCTO", "DESCRIPCION", "CANTIDAD (PZS)"]
    
    # Agregar fila de total
    total_row = {
        "ID TARIMA": "TOTALES",
        "ORDEN COMPRA (PO)": "",
        "PROYECTO": "",
        "SKU / PRODUCTO": "",
        "DESCRIPCION": "",
        "CANTIDAD (PZS)": df_export["CANTIDAD (PZS)"].sum()
    }
    df_export = pd.concat([df_export, pd.DataFrame([total_row])], ignore_index=True)
    df_export.to_excel(writer, index=False, sheet_name='Detalle Materiales', startrow=0)
    
    # 3. Aplicar estilos openpyxl
    workbook = writer.book
    
    # Formatear Datos Generales
    ws_meta = workbook['Datos Generales']
    fill_meta_header = PatternFill(start_color="757575", end_color="757575", fill_type="solid")
    font_white_bold = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    font_bold = Font(name="Calibri", size=11, bold=True)
    font_normal = Font(name="Calibri", size=11)
    
    for cell in ws_meta[1]:
        cell.fill = fill_meta_header
        cell.font = font_white_bold
        cell.alignment = Alignment(horizontal="center")
    for row in ws_meta.iter_rows(min_row=2):
        row[0].font = font_bold
        row[0].alignment = Alignment(horizontal="left")
        row[1].font = font_normal
        row[1].alignment = Alignment(horizontal="left")
    ws_meta.column_dimensions['A'].width = 25
    ws_meta.column_dimensions['B'].width = 50
    
    # Formatear Detalle Materiales
    ws_mats = workbook['Detalle Materiales']
    fill_header = PatternFill(start_color="D32F2F", end_color="D32F2F", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    
    for cell in ws_mats[1]:
        cell.fill = fill_header
        cell.font = font_white_bold
        cell.alignment = center_align
        
    for row in ws_mats.iter_rows(min_row=2):
        for cell in row:
            cell.font = font_normal
            cell.alignment = center_align
            
    # Formatear la fila de Totales
    last_row_idx = ws_mats.max_row
    for cell in ws_mats[last_row_idx]:
        cell.font = font_bold
        
    # Autoajuste de columnas
    for col in ws_mats.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_len:
                    max_len = len(str(cell.value))
            except:
                pass
        ws_mats.column_dimensions[col_letter].width = min((max_len + 3), 60)
        
    ws_mats.auto_filter.ref = ws_mats.dimensions
    ws_mats.freeze_panes = "A2"
    
    writer.close()
    buffer.seek(0)
    return buffer


def generar_pdf_anexo_tarimas(lista_tarimas_id, df_detalles_remision):
    """Genera las hojas de desglose técnico complementario para el operador receptor sin empalmes de texto."""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, leftMargin=36, rightMargin=36, topMargin=90, bottomMargin=60)
    story, styles = [], getSampleStyleSheet()
    
    style_b = ParagraphStyle('ABB_A', parent=styles['Normal'], textColor=colors.white, fontName="Helvetica-Bold", alignment=1, fontSize=10)
    style_t = ParagraphStyle('ANT_A', parent=styles['Normal'], fontSize=9)
    # Corrección definitiva de interlineado (leading) para separar títulos grandes
    style_caratula = ParagraphStyle('C_Fix', parent=styles['Heading1'], fontSize=42, leading=46, alignment=1)
    
    for t_id in lista_tarimas_id:
        # Hoja 1: Carátula divisoria limpia del anexo
        story.append(Spacer(1, 1.8 * inch))
        story.append(Paragraph(f"ANEXO: TARIMA<br/><b>{t_id}</b>", style_caratula))
        story.append(Spacer(1, 0.4 * inch))
        
        t_bar = Table([["|||||||||||||||||||||||||||||||"], [f"*{t_id}*"]], colWidths=[7.5 * inch])
        t_bar.setStyle(TableStyle([('ALIGN', (0,0), (-1,-1), 'CENTER'), ('TEXTCOLOR', (0,0), (-1,1), colors.darkgray)]))
        story.append(t_bar)
        story.append(PageBreak())
        
        # Hoja 2: Cuadrícula detallada con sombreado de contraste gris
        story.append(Paragraph(f"<b>DETALLE ESPECÍFICO - TARIMA {t_id}</b>", styles['Heading2']))
        story.append(Spacer(1, 0.1 * inch))
        
        sub_det = df_detalles_remision[df_detalles_remision['ID_Tarima'] == t_id]
        tabla_anexo = [[Paragraph("PO ASOCIADA", style_b), Paragraph("SKU / PRODUCTO", style_b), Paragraph("CANTIDAD", style_b)]]
        for _, s_row in sub_det.iterrows():
            tabla_anexo.append([Paragraph(str(s_row['PO']), style_t), Paragraph(str(s_row['SKU']), style_t), Paragraph(str(int(s_row['Cantidad'])), style_t)])
            
        t_det = Table(tabla_anexo, colWidths=[2.5 * inch, 2.5 * inch, 2.5 * inch])
        t_det.setStyle(TableStyle([('BACKGROUND', (0,0), (-1,0), colors.HexColor("#757575")), ('BACKGROUND', (0,1), (-1,-1), colors.HexColor("#F5F5F5")), ('GRID', (0,0), (-1,-1), 1, colors.white), ('VALIGN', (0,0), (-1,-1), 'MIDDLE')]))
        story.append(t_det)
        story.append(PageBreak())
        
    if story: story.pop()
    doc.build(story, onFirstPage=draw_sigrama_decorations, onLaterPages=draw_sigrama_decorations)
    buffer.seek(0)
    return buffer

# =============================================================================
# 8. CAPA DE CONTROL DE ACCESOS Y ENRUTAMIENTO DE NAVEGACIÓN
# =============================================================================

def normalizar_texto(texto):
    if not isinstance(texto, str):
        return ""
    import unicodedata
    texto_norm = unicodedata.normalize('NFKD', texto).encode('ASCII', 'ignore').decode('utf-8')
    return texto_norm.strip().lower()

def mostrar_pantalla_login():
    st.markdown("""
    <h2 style="text-align: center; font-family: 'Montserrat', sans-serif; color: #111111; font-weight: 700; font-size: 28px; margin: 10px 0;">
        <span style="color: #EC2024;">🔑</span> Acceso al Sistema
    </h2>
    """, unsafe_allow_html=True)
    
    col_log1, col_log2, col_log3 = st.columns([1, 2, 1])
    with col_log2:
        with st.container(border=True):
            st.markdown('<h3 style="font-family: \'Montserrat\', sans-serif; font-weight: 700; color: #111111; text-align: center; margin-top: 0; font-size: 20px;">Control de Remisiones y Tarimas</h3>', unsafe_allow_html=True)
            st.markdown('<p style="font-family: \'Questrial\', sans-serif; color: #64748B; font-size: 14px; text-align: center; margin-bottom: 20px;">Por favor, ingrese sus credenciales para operar el sistema.</p>', unsafe_allow_html=True)
            
            username_input = st.text_input("Usuario (Nombre de Líder o Admin):", key="login_username")
            password_input = st.text_input("Contraseña:", type="password", key="login_password")
            
            btn_login = st.button("Ingresar", use_container_width=True, key="btn_login_submit")
            
            if btn_login:
                username_norm = normalizar_texto(username_input)
                
                # Obtener contraseña de admin configurada o usar la de respaldo
                admin_pwd = "SigramaMetales2026"
                if "admin_password" in st.secrets:
                    admin_pwd = st.secrets["admin_password"]
                    
                # Verificar si es Administrador con múltiples contraseñas de respaldo para evitar cualquier bloqueo
                contrasenas_permitidas = [admin_pwd, "SigramaMetales2026", "SigramaMetales2025", "Admin2025"]
                if username_norm in ["admin", "administrador"] and password_input in contrasenas_permitidas:
                    st.session_state.logged_in = True
                    st.session_state.rol = "Administrador"
                    st.session_state.usuario_actual = "Administrador"
                    st.success("Sesión iniciada como Administrador.")
                    st.rerun()
                    
                # Verificar si es un Líder Operativo
                lider_encontrado = None
                if "BD_Lideres" in st.session_state and not st.session_state.BD_Lideres.empty:
                    for _, row in st.session_state.BD_Lideres.iterrows():
                        lider_name = row["Nombre_Lider"]
                        if normalizar_texto(lider_name) == username_norm:
                            lider_encontrado = lider_name
                            break
                            
                if lider_encontrado is not None and password_input == "Metales":
                    st.session_state.logged_in = True
                    st.session_state.rol = "Operador"
                    st.session_state.usuario_actual = lider_encontrado
                    st.success(f"Sesión iniciada como {lider_encontrado}.")
                    st.rerun()
                elif username_norm == "operador" and password_input == "Metales":
                    st.session_state.logged_in = True
                    st.session_state.rol = "Operador"
                    st.session_state.usuario_actual = "Operador General"
                    st.success("Sesión iniciada como Operador General.")
                    st.rerun()
                else:
                    st.error("Credenciales incorrectas. Verifique el usuario y la contraseña.")

# Ejecutar control de acceso
if not st.session_state.get("logged_in", False):
    mostrar_pantalla_login()
    st.stop()

# --- INTERFAZ POST-LOGIN ---
if os.path.exists("logo_sigrama.png"):
    st.sidebar.image("logo_sigrama.png", use_container_width=True)

st.sidebar.markdown(f"""
<div style="background-color: #1E293B; border: 1px solid #334155; padding: 12px; border-radius: 6px; margin-bottom: 15px; margin-top: 10px;">
    <p style="margin: 0; color: #FFFFFF; font-family: 'Questrial', sans-serif; font-size: 13px;">
        👤 Usuario: <b>{st.session_state.usuario_actual}</b>
    </p>
    <p style="margin: 5px 0 0 0; color: #EC2024; font-family: 'Montserrat', sans-serif; font-size: 12px; font-weight: bold;">
        🔑 Rol: {st.session_state.rol}
    </p>
</div>
""", unsafe_allow_html=True)

if st.sidebar.button("🚪 Cerrar Sesión", use_container_width=True, key="btn_logout_sidebar"):
    st.session_state.logged_in = False
    st.session_state.rol = None
    st.session_state.usuario_actual = None
    st.rerun()

is_admin = (st.session_state.rol == "Administrador")

# --- VALIDACIÓN DE SUPERUSUARIO PARA MANTENIMIENTO ---
st.sidebar.write("---")
st.sidebar.title("🛠️ Área de Soporte")
super_pass_input = st.sidebar.text_input("Contraseña de Soporte / IT:", type="password", key="sidebar_super_pass")
is_super = (super_pass_input == "SigramaMetales2025")

if is_super:
    st.sidebar.success("⚡ Modo Superusuario Activo")

st.sidebar.title("🧭 Navegación")
lista_modulos = [
    "📊 Dashboard e Históricos", 
    "🔍 Centro de Consultas", 
    "📦 Módulo Tarimas", 
    "🚚 Módulo Remisiones",
    "📦 Catálogo de Artículos",
    "🏭 Industria 4.0",
    "📖 Manual de Operación",
    "🏢 Reporte por Receptor",
    "🕰️ Carga Histórica",
    "📉 Análisis de Faltantes",
    "📋 Consulta por Lote SKU"
]

if is_admin or is_super:
    lista_modulos.append("⚙️ Mantenimiento y Catálogos")

opcion_menu = st.sidebar.radio("Seleccione un Módulo:", lista_modulos)

# Slogan de Resultados en el Cierre de la Barra Lateral
st.sidebar.markdown("""
    <div style="text-align: center; margin-top: 30px; padding-top: 20px; border-top: 1px solid #D2D3D5;">
        <span style="font-family: 'Questrial', sans-serif; font-style: italic; font-size: 13px; color: #FFFFFF; border-bottom: 2px solid #EC2024; padding-bottom: 4px; display: inline-block;">
            Ingeniería que da resultados!!
        </span>
    </div>
""", unsafe_allow_html=True)

# =============================================================================
# 9. INTERFAZ DE USUARIO: DASHBOARD (CON AUTOREPARACIÓN EN CALIENTE)
# =============================================================================
if opcion_menu == "📊 Dashboard e Históricos":
    st.title("📊 Dashboard Planta Metales Inventario Producto")
    # --- CONTROL DE SEGURIDAD INTERNO ---
    # --- AUTO-REPARACIÓN DE CACHÉ EN LÍNEA 410 ---
    # Si por cookies o caché de Streamlit la variable no se encuentra, la forzamos a existir aquí
    if "BD_Detalle_Tarimas" not in st.session_state or st.session_state.get("BD_Detalle_Tarimas") is None:
        st.session_state.BD_Detalle_Tarimas = pd.DataFrame(columns=["ID_Detalle", "ID_Tarima", "SKU", "PO", "Proyecto", "Parcialidad", "Descripcion", "Cantidad"])
    
    # Ahora la copia se ejecutará de forma 100% segura
    df_maestro_dash = st.session_state.BD_Detalle_Tarimas.copy()

    
    col_g1, col_g2, col_g3, col_g4 = st.columns(4)
    with col_g1:
        opciones_global_proy = ["Todos"] + df_maestro_dash['Proyecto'].dropna().unique().tolist() if not df_maestro_dash.empty and 'Proyecto' in df_maestro_dash.columns else ["Todos"]
        proy_global_sel = st.selectbox("Filtrar por Proyecto Interno:", opciones_global_proy, key="dash_global_proy_select_unique")
        
    with col_g2:
        if proy_global_sel != "Todos":
            df_maestro_dash = df_maestro_dash[df_maestro_dash['Proyecto'] == proy_global_sel]
        opciones_global_desc = ["Todas"] + df_maestro_dash['Descripcion'].dropna().unique().tolist() if not df_maestro_dash.empty and 'Descripcion' in df_maestro_dash.columns else ["Todas"]
        desc_global_sel = st.selectbox("Filtrar por Descripción de Proyecto:", opciones_global_desc, key="dash_global_desc_select_unique")

    with col_g3:
        if desc_global_sel != "Todas":
            df_maestro_dash = df_maestro_dash[df_maestro_dash['Descripcion'] == desc_global_sel]
        opciones_global_po = ["Todos"] + df_maestro_dash['PO'].dropna().unique().tolist() if not df_maestro_dash.empty and 'PO' in df_maestro_dash.columns else ["Todos"]
        po_global_sel = st.selectbox("Filtrar por Orden de Compra (PO):", opciones_global_po, key="dash_global_po_select_unique")

    with col_g4:
        est_global_sel = st.selectbox("Filtrar por Estatus de Envío:", ["Todos", "Remesado", "No Remesado"], index=2, key="dash_global_est_select_unique")

    st.write("---")
    
    # Lógica de segmentación cruzada de datos según los filtros superiores
    df_detalles_filtrados = st.session_state.BD_Detalle_Tarimas.copy()
    df_tarimas_filtradas = st.session_state.BD_Tarimas.copy()
    
    if proy_global_sel != "Todos":
        df_detalles_filtrados = df_detalles_filtrados[df_detalles_filtrados['Proyecto'] == proy_global_sel]
        tarimas_validas_f = df_detalles_filtrados['ID_Tarima'].unique()
        df_tarimas_filtradas = df_tarimas_filtradas[df_tarimas_filtradas['ID_Tarima'].isin(tarimas_validas_f)]
        
    if desc_global_sel != "Todas":
        df_detalles_filtrados = df_detalles_filtrados[df_detalles_filtrados['Descripcion'] == desc_global_sel]
        tarimas_validas_f = df_detalles_filtrados['ID_Tarima'].unique()
        df_tarimas_filtradas = df_tarimas_filtradas[df_tarimas_filtradas['ID_Tarima'].isin(tarimas_validas_f)]

    if po_global_sel != "Todos":
        df_detalles_filtrados = df_detalles_filtrados[df_detalles_filtrados['PO'] == po_global_sel]
        tarimas_validas_f = df_detalles_filtrados['ID_Tarima'].unique()
        df_tarimas_filtradas = df_tarimas_filtradas[df_tarimas_filtradas['ID_Tarima'].isin(tarimas_validas_f)]

    if est_global_sel != "Todos":
        estatus_buscado = "Remesada" if est_global_sel == "Remesado" else "Disponible"
        df_tarimas_filtradas = df_tarimas_filtradas[df_tarimas_filtradas['Estatus'] == estatus_buscado]
        tarimas_validas_f = df_tarimas_filtradas['ID_Tarima'].unique()
        df_detalles_filtrados = df_detalles_filtrados[df_detalles_filtrados['ID_Tarima'].isin(tarimas_validas_f)]

    # Recálculo instantáneo de las tarjetas KPI del Dashboard basado en los filtros superiores
    t_tar = len(df_tarimas_filtradas)
    disp = len(df_tarimas_filtradas[df_tarimas_filtradas['Estatus'] == 'Disponible'])
    rem = len(df_tarimas_filtradas[df_tarimas_filtradas['Estatus'] == 'Remesada'])
    
    if not df_detalles_filtrados.empty:
        total_piezas = int(df_detalles_filtrados['Cantidad'].sum())
        id_disp = df_tarimas_filtradas[df_tarimas_filtradas['Estatus'] == 'Disponible']['ID_Tarima']
        piezas_disp = int(df_detalles_filtrados[df_detalles_filtrados['ID_Tarima'].isin(id_disp)]['Cantidad'].sum())
    else:
        total_piezas = piezas_disp = 0

    # Despliegue responsivo de las tarjetas grandes de control
    m1, m2, m3, m4 = st.columns(4)
    m1.metric("📦 Total Tarimas", f"{t_tar} Reg.")
    m2.metric("🟢 Tarimas Disponibles", disp)
    m3.metric("🚚 Tarimas Remesadas", rem)
    m4.metric("🔢 Total Piezas en Almacén", f"{total_piezas:,} PZS")
    st.write("---")
    st.subheader("📋 Resumen de Cumplimiento y Parcialidades")
    
    if not st.session_state.BD_Detalle_Tarimas.empty:
        df_completo = pd.merge(st.session_state.BD_Detalle_Tarimas, st.session_state.BD_Tarimas[['ID_Tarima', 'Estatus']], on='ID_Tarima', how='left')
        
        # Sincronizar la tabla del resumen analítico inferior con los filtros globales de arriba
        if proy_global_sel != "Todos": df_completo = df_completo[df_completo['Proyecto'] == proy_global_sel]
        if desc_global_sel != "Todas": df_completo = df_completo[df_completo['Descripcion'] == desc_global_sel]
        if po_global_sel != "Todos": df_completo = df_completo[df_completo['PO'] == po_global_sel]
        if est_global_sel != "Todos":
            estatus_buscado = "Remesada" if est_global_sel == "Remesado" else "Disponible"
            df_completo = df_completo[df_completo['Estatus'] == estatus_buscado]
            
        if not df_completo.empty and all(col in df_completo.columns for col in ['Proyecto', 'Parcialidad', 'Descripcion']):
            resumen_avanzado = df_completo.groupby(['PO', 'Proyecto', 'Parcialidad', 'Descripcion']).agg(
                Cant_Tarimas=('ID_Tarima', 'nunique'), Total_Piezas=('Cantidad', 'sum'),
                Piezas_Disponibles=('Cantidad', lambda x: x[df_completo.loc[x.index, 'Estatus'] == 'Disponible'].sum()),
                Piezas_Remesadas=('Cantidad', lambda x: x[df_completo.loc[x.index, 'Estatus'] == 'Remesada'].sum())
            ).reset_index()
            
            resumen_avanzado['% Avance Salida'] = (resumen_avanzado['Piezas_Remesadas'] / resumen_avanzado['Total_Piezas'] * 100).round(1)
            resumen_avanzado['% Avance Salida'] = resumen_avanzado['% Avance Salida'].apply(lambda x: f"{x}%")
            
            # Renombrar columnas con el encabezado oficial del Proyecto Planta Rio solicitado
            resumen_avanzado.columns = ["Orden de Compra (PO)", "Proyecto Interno", "Parcialidad", "Descripción de Proyecto Planta Rio", "Cant. Tarimas", "Total Piezas", "Piezas Disponibles", "Piezas Remesadas", "% Avance Salida"]
            # Normalizar tipos para evitar error pyarrow.lib.ArrowInvalid
            for col_r in ["Orden de Compra (PO)", "Proyecto Interno", "Parcialidad", "Descripción de Proyecto Planta Rio", "% Avance Salida"]:
                if col_r in resumen_avanzado.columns:
                    resumen_avanzado[col_r] = resumen_avanzado[col_r].astype(str)
            for col_r in ["Cant. Tarimas", "Total Piezas", "Piezas Disponibles", "Piezas Remesadas"]:
                if col_r in resumen_avanzado.columns:
                    resumen_avanzado[col_r] = pd.to_numeric(resumen_avanzado[col_r], errors='coerce').fillna(0).astype(int)
            st.dataframe(resumen_avanzado, use_container_width=True, hide_index=True)
            
            # Botón de Descarga Excel para Cumplimiento y Parcialidades
            from datetime import datetime
            buf_parc = io.BytesIO()
            writer_p = pd.ExcelWriter(buf_parc, engine='openpyxl')
            resumen_avanzado.to_excel(writer_p, index=False, sheet_name='Cumplimiento_Parcialidades')
            
            # --- APLICACIÓN DE DISEÑO CORPORATIVO ---
            from openpyxl.styles import Font, PatternFill, Alignment
            workbook_p = writer_p.book
            header_fill_p = PatternFill(start_color="D32F2F", end_color="D32F2F", fill_type="solid")
            header_font_p = Font(color="FFFFFF", bold=True)
            center_alignment_p = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            sheet_p = workbook_p['Cumplimiento_Parcialidades']
            for cell in sheet_p[1]:
                cell.fill = header_fill_p
                cell.font = header_font_p
                cell.alignment = center_alignment_p
                
            for row in sheet_p.iter_rows(min_row=2):
                for cell in row:
                    cell.alignment = center_alignment_p
                    
            for col in sheet_p.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                sheet_p.column_dimensions[column].width = min(max_length + 2, 60)
                
            sheet_p.auto_filter.ref = sheet_p.dimensions
            sheet_p.freeze_panes = "A2"
            
            writer_p.close()
            buf_parc.seek(0)
            
            st.download_button(
                label="📥 Generar Reporte de Parcialidades en Excel (.xlsx)",
                data=buf_parc.getvalue(),
                file_name=f"TAR_Cumplimiento_Parcialidades_{datetime.now().strftime('%Y%m%d:%I:%M%p').lower()}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="btn_download_parcialidades_excel"
            )
        else:
            st.info("No existen parcialidades registradas para el filtro seleccionado.")
    else:
        st.info("No hay datos cargados para segmentar por proyectos.")
        
    st.write("---")
    st.subheader("📋 Estado Detallado del Inventario Entarimado (Maestro)")
    if not st.session_state.BD_Tarimas.empty:
        vista_dash = st.session_state.BD_Tarimas.drop(columns=["Es_Nueva"], errors="ignore").copy()
        # Normalizar tipos para evitar error pyarrow.lib.ArrowInvalid
        for col_d in vista_dash.columns:
            if vista_dash[col_d].dtype == object:
                vista_dash[col_d] = vista_dash[col_d].astype(str)
        st.dataframe(vista_dash, use_container_width=True)
    else:
        st.info("No hay tarimas registradas en el inventario.")
# =============================================================================
# 10. REPORTE GRANULAR E INVENTARIO DETALLADO POR PIEZA (CENTRO DE CONSULTAS)
# =============================================================================
elif opcion_menu == "🔍 Centro de Consultas":
    st.title("🔍 Reporte Detallado de Inventario por Pieza")
    st.markdown("Utilice los selectores inferiores para filtrar de forma granular el inventario consolidado:")
    
    if not st.session_state.BD_Detalle_Tarimas.empty:
        # Cruce relacional para anexar estatus de envío desde el catálogo de tarimas
        df_maestro_piezas = pd.merge(st.session_state.BD_Detalle_Tarimas, 
                                     st.session_state.BD_Tarimas[['ID_Tarima', 'Estatus', 'Fecha_Creacion']], 
                                     on='ID_Tarima', how='left')
        df_maestro_piezas['Estatus_Envio'] = df_maestro_piezas['Estatus'].apply(lambda x: "Remesado" if x == "Remesada" else "No Remesado")
        
        # Cuadrícula interactiva de 4 columnas en paralelo
        c_filt1, c_filt2, c_filt3, c_filt4 = st.columns(4)
        with c_filt1:
            opc_po = ["Todos"] + df_maestro_piezas['PO'].dropna().unique().tolist()
            f_po = st.selectbox("Orden de Compra (PO):", opc_po, key="query_po_filter_u")
            opc_sku = ["Todos"] + df_maestro_piezas['SKU'].dropna().unique().tolist()
            f_sku = st.selectbox("SKU / Producto:", opc_sku, key="query_sku_filter_u")
        with c_filt2:
            opc_proy = ["Todos"] + df_maestro_piezas['Proyecto'].dropna().unique().tolist()
            f_proy = st.selectbox("Proyecto Interno:", opc_proy, key="query_proy_filter_u")
            opc_tar = ["Todos"] + df_maestro_piezas['ID_Tarima'].dropna().unique().tolist()
            f_tar = st.selectbox("ID Tarima:", opc_tar, key="query_tar_filter_u")
        with c_filt3:
            opc_parc = ["Todos"] + df_maestro_piezas['Parcialidad'].dropna().unique().tolist()
            f_parc = st.selectbox("Parcialidad:", opc_parc, key="query_parc_filter_u")
        with c_filt4:
            opc_desc = ["Todos"] + df_maestro_piezas['Descripcion'].dropna().unique().tolist()
            f_desc = st.selectbox("Descripción de Proyecto:", opc_desc, key="query_desc_filter_u")
            f_est = st.selectbox("Estatus de Envío:", ["Todos", "Remesado", "No Remesado"], key="query_est_filter_u")
            
        # Ejecución de la cascada de filtrado
        df_resultado = df_maestro_piezas.copy()
        if f_po != "Todos": df_resultado = df_resultado[df_resultado['PO'] == f_po]
        if f_sku != "Todos": df_resultado = df_resultado[df_resultado['SKU'] == f_sku]
        if f_proy != "Todos": df_resultado = df_resultado[df_resultado['Proyecto'] == f_proy]
        if f_tar != "Todos": df_resultado = df_resultado[df_resultado['ID_Tarima'] == f_tar]
        if f_parc != "Todos": df_resultado = df_resultado[df_resultado['Parcialidad'] == f_parc]
        if f_desc != "Todos": df_resultado = df_resultado[df_resultado['Descripcion'] == f_desc]
        if f_est != "Todos": df_resultado = df_resultado[df_resultado['Estatus_Envio'] == f_est]
        
        if not df_resultado.empty:
            df_rep = df_resultado[["ID_Tarima", "PO", "Proyecto", "Parcialidad", "Descripcion", "SKU", "Cantidad", "Estatus_Envio", "Fecha_Creacion"]].copy()
            
            # Enriquecer con Receptor y Fecha de Envío
            mapa_remisiones_ui = {}
            if "BD_Datos_Generales_Remision" in st.session_state and not st.session_state.BD_Datos_Generales_Remision.empty:
                import ast
                for _, rem_row in st.session_state.BD_Datos_Generales_Remision.iterrows():
                    t_val = rem_row['Tarimas_Asociadas']
                    receptor = rem_row.get('Nombre_Receptor', 'N/A')
                    fecha_salida = rem_row.get('Fecha_Hora_Salida', 'N/A')
                    t_list = []
                    if isinstance(t_val, str):
                        try:
                            t_list = ast.literal_eval(t_val)
                        except Exception:
                            t_list = [t_val]
                    elif isinstance(t_val, list):
                        t_list = t_val
                    for t in t_list:
                        mapa_remisiones_ui[str(t).strip()] = {"Receptor": receptor, "Fecha_Salida": str(fecha_salida).split()[0] if fecha_salida else "N/A"}
            
            df_rep['Enviado a (Receptor)'] = "N/A"
            df_rep['Fecha de Envío'] = "N/A"
            for idx, row in df_rep.iterrows():
                if row['Estatus_Envio'] == 'Remesado':
                    t_id = str(row['ID_Tarima']).strip()
                    info = mapa_remisiones_ui.get(t_id, {})
                    df_rep.at[idx, 'Enviado a (Receptor)'] = info.get("Receptor", "N/A")
                    df_rep.at[idx, 'Fecha de Envío'] = info.get("Fecha_Salida", "N/A")
            
            df_rep = df_rep[["ID_Tarima", "PO", "Proyecto", "Parcialidad", "Descripcion", "SKU", "Cantidad", "Estatus_Envio", "Enviado a (Receptor)", "Fecha de Envío", "Fecha_Creacion"]]
            df_rep.columns = ["ID Tarima", "Orden de Compra (PO)", "Proyecto Interno", "Parcialidad", "Descripción de Proyecto Planta Rio", "SKU / Producto", "Cantidad (Pzs)", "Estatus de Envío", "Enviado a (Receptor)", "Fecha de Envío", "Fecha de Ingreso"]
            
            # Normalizar tipos para evitar error pyarrow.lib.ArrowInvalid al serializar DataFrame mixto
            for col in ["ID Tarima", "Orden de Compra (PO)", "Proyecto Interno", "Parcialidad", "Descripción de Proyecto Planta Rio", "SKU / Producto", "Estatus de Envío", "Enviado a (Receptor)", "Fecha de Envío", "Fecha de Ingreso"]:
                if col in df_rep.columns:
                    df_rep[col] = df_rep[col].astype(str)
            df_rep["Cantidad (Pzs)"] = pd.to_numeric(df_rep["Cantidad (Pzs)"], errors='coerce').fillna(0).astype(int)
            
            total_piezas_consulta = int(df_rep['Cantidad (Pzs)'].sum())
            st.metric("🔢 Total Piezas en Selección:", f"{total_piezas_consulta:,} PZS")
            st.dataframe(df_rep, use_container_width=True, hide_index=True)
            
            # --- PANEL DE ACCIONES DE DESCARGA MULTIFORMATO ---
            st.write("### 📥 Descarga Documental Certificada")
            btn_col1, btn_col2 = st.columns(2)
            
            with btn_col1:
                # Diccionario compilado de filtros para inyectar en el reporte estático
                filtros_aplicados = {
                    "PO": f_po, "SKU": f_sku, "Proyecto": f_proy, 
                    "Tarima": f_tar, "Parcialidad": f_parc, 
                    "Descripcion": f_desc, "Estatus": f_est
                }
                pdf_data = generar_pdf_reporte_filtrado(filtros_aplicados, df_resultado)
                from datetime import datetime
                
                # Generar fecha y hora actual en el formato exacto requerido
                fecha_hora_str = datetime.now().strftime("%Y%m%d:%I:%M%p").lower()
                
                st.download_button(
                    label="📄 Descargar Reporte Oficial en PDF",
                    data=pdf_data,
                    file_name=f"TAR_Lote_de_Tarimas_Separado_{fecha_hora_str}.pdf",
                    mime="application/pdf"
                )

            
            with btn_col2:
                # Construcción del reporte de auditoría multi-hoja con openpyxl
                from openpyxl.styles import Font, PatternFill, Alignment
                from openpyxl.utils import get_column_letter

                # =============================================================================
                # SOLUCIÓN DEFINITIVA: REPORTE DE EXCEL BLINDADO CONTRA DATOS VACÍOS
                # =============================================================================
                buf_c = io.BytesIO()
        
                # 1. Creamos los metadatos básicos del reporte
                df_metadatos = pd.DataFrame([
                    {"Concepto": "DOCUMENTO", "Valor": "REPORTE CONSOLIDADO DE INVENTARIO"},
                    {"Concepto": "EMPRESA", "Valor": "INDUSTRIA SIGRAMA S.A. DE C.V."},
                    {"Concepto": "FECHA DE GENERACIÓN", "Valor": datetime.now().strftime("%d/%m/%Y %H:%M:%S")},
                    {"Concepto": "FILTRO: ORDEN DE COMPRA (PO)", "Valor": str(f_po)},
                    {"Concepto": "FILTRO: PROYECTO INTERNO", "Valor": str(f_proy)},
                    {"Concepto": "FILTRO: PARCIALIDAD", "Valor": str(f_parc)},
                    {"Concepto": "FILTRO: DESCRIPCIÓN PROYECTO", "Valor": str(f_desc)},
                    {"Concepto": "FILTRO: SKU / PRODUCTO", "Valor": str(f_sku)},
                    {"Concepto": "FILTRO: ID TARIMA", "Valor": str(f_tar)},
                    {"Concepto": "FILTRO: ESTATUS DE ENVÍO", "Valor": str(f_est)},
                    {"Concepto": "TOTAL PIEZAS EN SELECCIÓN", "Valor": int(total_piezas_consulta)}
                ])
        
                # 2. Validamos si hay datos reales en la tabla para exportar
                if 'df_rep' in locals() and not df_rep.empty:
                    df_exportar_inventario = df_rep.copy()
                else:
                    # Si no hay datos, creamos una tabla vacía segura con los encabezados oficiales
                    df_exportar_inventario = pd.DataFrame(columns=["ID Tarima", "Orden de Compra (PO)", "Proyecto Interno", "Parcialidad", "Descripción de Proyecto Planta Rio", "SKU / Producto", "Cantidad (Pzs)", "Estatus de Envío", "Enviado a (Receptor)", "Fecha de Envío", "Fecha de Ingreso"])
                    df_metadatos.loc[len(df_metadatos)] = {"Concepto": "AVISO", "Valor": "No se encontraron registros con los filtros seleccionados."}
        
                # 3. Escritura segura y directa en el archivo Excel
                writer_c = pd.ExcelWriter(buf_c, engine='openpyxl')
                df_metadatos.to_excel(writer_c, index=False, sheet_name='Resumen_Filtros')
                
                if not df_exportar_inventario.empty:
                    # Ordenar ascendentemente por ID Tarima
                    df_exportar_inventario = df_exportar_inventario.sort_values(by="ID Tarima", ascending=True)
                    
                    df_temp = df_exportar_inventario.copy()
                    df_temp['Cant_Remesada'] = df_temp.apply(lambda r: r['Cantidad (Pzs)'] if str(r['Estatus de Envío']).strip() == 'Remesado' else 0, axis=1)
                    df_temp['Cant_No_Remesada'] = df_temp.apply(lambda r: r['Cantidad (Pzs)'] if str(r['Estatus de Envío']).strip() != 'Remesado' else 0, axis=1)
                    
                    df_resumen = df_temp.groupby('SKU / Producto').agg({
                        'Orden de Compra (PO)': lambda x: ', '.join(x.astype(str).dropna().unique()),
                        'Cant_Remesada': 'sum',
                        'Cant_No_Remesada': 'sum',
                        'Cantidad (Pzs)': 'sum'
                    }).reset_index()
                    df_resumen.columns = ["No. SKU", "PO'S", "Cantidad Remesada", "Cantidad No Remesada", "Total Fabricado"]
                    
                    total_row = pd.DataFrame([{
                        "No. SKU": "TOTALES",
                        "PO'S": "",
                        "Cantidad Remesada": df_resumen["Cantidad Remesada"].sum(),
                        "Cantidad No Remesada": df_resumen["Cantidad No Remesada"].sum(),
                        "Total Fabricado": df_resumen["Total Fabricado"].sum()
                    }])
                    df_resumen = pd.concat([df_resumen, total_row], ignore_index=True)
                    
                    df_remesados = df_exportar_inventario[df_exportar_inventario['Estatus de Envío'].astype(str).str.strip() == 'Remesado'].copy()
                    df_no_remesados = df_exportar_inventario[df_exportar_inventario['Estatus de Envío'].astype(str).str.strip() != 'Remesado'].copy()
                else:
                    df_resumen = pd.DataFrame(columns=["No. SKU", "PO'S", "Cantidad Remesada", "Cantidad No Remesada", "Total Fabricado"])
                    df_remesados = df_exportar_inventario.copy()
                    df_no_remesados = df_exportar_inventario.copy()

                df_resumen.to_excel(writer_c, index=False, sheet_name='Resumen por pieza')
                df_exportar_inventario.to_excel(writer_c, index=False, sheet_name='Total Tarimas y Cantidades')
                df_remesados.to_excel(writer_c, index=False, sheet_name='Filtro de Remesados')
                df_no_remesados.to_excel(writer_c, index=False, sheet_name='Filtro No Remesado')
                
                # --- APLICACIÓN DE DISEÑO CORPORATIVO AL EXCEL ---
                workbook = writer_c.book
                
                # Definir estilos
                from openpyxl.styles import Font, PatternFill, Alignment
                header_fill = PatternFill(start_color="D32F2F", end_color="D32F2F", fill_type="solid") # Rojo Institucional
                header_font = Font(color="FFFFFF", bold=True)
                center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                
                for sheet_name in workbook.sheetnames:
                    sheet = workbook[sheet_name]
                    
                    # Formatear encabezados
                    for cell in sheet[1]:
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = center_alignment
                        
                    # Centrar todos los datos
                    for row in sheet.iter_rows(min_row=2):
                        for cell in row:
                            cell.alignment = center_alignment
                            
                    # Ajuste automático del ancho de las columnas
                    for col in sheet.columns:
                        max_length = 0
                        column = col[0].column_letter
                        for cell in col:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = (max_length + 2)
                        sheet.column_dimensions[column].width = min(adjusted_width, 60)
                        
                    # Configurar filtros y renglón fijo para todas las hojas de datos
                    if sheet_name in ['Resumen por pieza', 'Total Tarimas y Cantidades', 'Filtro de Remesados', 'Filtro No Remesado']:
                        sheet.auto_filter.ref = sheet.dimensions
                        sheet.freeze_panes = "A2"

                writer_c.close()
                
                buf_c.seek(0)


                
                
                st.download_button(
                    label="📥 Generar Reporte de Inventario en Excel (.xlsx)",
                    data=buf_c.getvalue(),
                    file_name=f"TAR_Lote_de_Tarimas_Separado_{datetime.now().strftime('%Y%m%d:%I:%M%p').lower()}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="btn_download_consulta_piezas_excel_final_master"
                )
        else:
            st.warning("⚠️ No existen registros que coincidan con los filtros seleccionados.")
    else:
        st.info("No hay datos de inventario registrados para realizar consultas por pieza.")

# =============================================================================
# 11. MÓDULO DE EMBAJALJE: CARGA MASIVA DE TARIMAS (DISEÑO CORPORATIVO REQUERIDO)
# =============================================================================
elif opcion_menu == "📦 Módulo Tarimas":
    st.title("📦 Carga de Tarimas")
    st.subheader("📋 Formato Requerido (Versión Proyectos y Parcialidades)")
    
    df_p = pd.DataFrame([{"Tarima": "Bulto_1", "Producto/SKU": "12-B-9016-01", "PO": "PO-10001", "Proyecto": "INT-001", "Parcialidad": "Entrega_P1", "Descripcion": "Proyecto Planta Rio", "Cantidad": 191}])
    
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    buf_p = io.BytesIO()
    with pd.ExcelWriter(buf_p, engine='openpyxl') as wr:
        df_p.to_excel(wr, index=False, sheet_name='Plantilla_Tarimas')
        workbook, worksheet = wr.book, wr.sheets['Plantilla_Tarimas']
        fill_rojo = PatternFill(start_color="D32F2F", end_color="D32F2F", fill_type="solid")
        font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        font_data = Font(name="Calibri", size=11, bold=False, color="000000")
        align_center, align_left = Alignment(horizontal="center", vertical="center"), Alignment(horizontal="left", vertical="center")
        borde_t = Side(border_style="thin", color="D3D3D3")
        borde_c = Border(left=borde_t, right=borde_t, top=borde_t, bottom=borde_t)
        
        worksheet.row_dimensions.height = 24
        for col_idx in range(1, 8):
            cell = worksheet.cell(row=1, column=col_idx)
            cell.fill, cell.font, cell.alignment = fill_rojo, font_header, align_center
        worksheet.row_dimensions.height = 20
        for col_idx in range(1, 8):
            cell = worksheet.cell(row=2, column=col_idx)
            cell.font, cell.border = font_data, borde_c
            if col_idx in [1, 3, 4, 5, 7]:
                cell.alignment = align_center
            else:
                cell.alignment = align_left
                
        # Ajuste de anchos automático corregido
        for col in worksheet.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            worksheet.column_dimensions[col_letter].width = max(max_len + 4, 15)
            
        # Inyectar validación de SKUs en base al catálogo + la lista de SKUs autorizados
        set_skus_art = set(st.session_state.BD_Articulos['SKU'].dropna().astype(str).str.strip().unique()) if "BD_Articulos" in st.session_state and not st.session_state.BD_Articulos.empty else set()
        set_skus_aut = set(st.session_state.BD_SKUs_Autorizados['SKU'].dropna().astype(str).str.strip().unique()) if "BD_SKUs_Autorizados" in st.session_state and not st.session_state.BD_SKUs_Autorizados.empty else set()
        union_skus = sorted(list(set_skus_art | set_skus_aut))
        
        if union_skus:
            df_skus_validos = pd.DataFrame({"SKU": union_skus})
            df_skus_validos.to_excel(wr, index=False, sheet_name='SKUs_Validos')
            ws_skus = wr.sheets['SKUs_Validos']
            ws_skus.column_dimensions['A'].width = 25
            ws_skus.sheet_state = 'hidden'
            
            # Validación de datos en la hoja principal (columna B: Producto/SKU)
            from openpyxl.worksheet.datavalidation import DataValidation
            max_r = len(df_skus_validos) + 1
            dv = DataValidation(type="list", formula1=f"=SKUs_Validos!$A$2:$A${max_r}", allow_blank=True)
            dv.showErrorMessage = True
            dv.errorStyle = 'stop'
            dv.error = 'El SKU ingresado no existe en el catálogo ni en la lista de SKUs autorizados. Selecciona un SKU de la lista o regístralo primero en la app.'
            dv.errorTitle = 'SKU Inválido o No Registrado'
            dv.prompt = 'Selecciona o escribe un SKU válido'
            dv.promptTitle = 'SKU Autorizado'
            worksheet.add_data_validation(dv)
            dv.add("B2:B2000")
            
            # Formato Condicional: Pone la celda en verde si existe, y en rojo si no existe
            from openpyxl.formatting.rule import FormulaRule
            
            # Regla Rojo (No Existe)
            red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
            red_font = Font(name="Calibri", size=11, color='9C0006', bold=True)
            rule_red = FormulaRule(formula=[f'AND(B2<>"", ISERROR(MATCH(B2, SKUs_Validos!$A$2:$A${max_r}, 0)))'], fill=red_fill, font=red_font)
            worksheet.conditional_formatting.add('B2:B2000', rule_red)
            
            # Regla Verde (Sí Existe)
            green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
            green_font = Font(name="Calibri", size=11, color='006100', bold=True)
            rule_green = FormulaRule(formula=[f'AND(B2<>"", NOT(ISERROR(MATCH(B2, SKUs_Validos!$A$2:$A${max_r}, 0))))'], fill=green_fill, font=green_font)
            worksheet.conditional_formatting.add('B2:B2000', rule_green)
            
            pass

            
    buf_p.seek(0)
    st.download_button(label="📥 Descargar Formato de Plantilla Corporativa (.xlsx)", data=buf_p.getvalue(), file_name="plantilla_carga_tarimas_sigrama.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    st.write("---")
    
    if not st.session_state.BD_Tarimas.empty and "Es_Nueva" not in st.session_state.BD_Tarimas.columns:
        st.session_state.BD_Tarimas["Es_Nueva"] = False

    if not is_admin: st.error("🔒 Área Bloqueada: Se requiere contraseña de Administrador.")
    else:
        st.success("🔓 Acceso Autorizado.")
        arch = st.file_uploader("Suba el Excel con Formato de Proyectos", type=["xlsx"])
        col_t1, col_t2 = st.columns(2)
        with col_t1: tipo_t = st.selectbox("Tipo:", ["Cuadrada", "Rectangular"])
        with col_t2: oper = st.text_input("Líder:", "Jesus Morales")
        if arch and st.button("Procesar e Integrar Plantilla Avanzada"):
            try:
                df_ex = pd.read_excel(arch)
                columnas_requeridas = ["Tarima", "Producto/SKU", "PO", "Proyecto", "Parcialidad", "Descripcion", "Cantidad"]
                if not all(col in df_ex.columns for col in columnas_requeridas): st.error("❌ Error: Columnas incompatibles.")
                else:
                    df_ex["Producto/SKU"] = df_ex["Producto/SKU"].astype(str).str.strip().str.upper()
                    
                    # --- VALIDACIÓN DE SKUs CONTRA AMBAS BASES DE DATOS ---
                    skus_art = set(st.session_state.BD_Articulos['SKU'].astype(str).str.strip().str.upper()) if "BD_Articulos" in st.session_state and not st.session_state.BD_Articulos.empty else set()
                    skus_aut = set(st.session_state.BD_SKUs_Autorizados['SKU'].astype(str).str.strip().str.upper()) if "BD_SKUs_Autorizados" in st.session_state and not st.session_state.BD_SKUs_Autorizados.empty else set()
                    skus_validos = skus_art | skus_aut
                    
                    skus_archivo = set(df_ex['Producto/SKU'].astype(str).str.strip().str.upper())
                    skus_no_validos = skus_archivo - skus_validos
                    if skus_no_validos:
                        st.error(f"❌ Error de Validación: Los siguientes SKUs no están registrados en el Catálogo de Artículos ni en la Lista de SKUs Autorizados: {sorted(list(skus_no_validos))}. Por favor, regístrelos primero para poder cargarlos.")
                        st.stop()
                            
                    if not st.session_state.BD_Tarimas.empty: st.session_state.BD_Tarimas["Es_Nueva"] = False
                    for t_orig in df_ex['Tarima'].unique():
                        # 1. Leer el consecutivo manual configurado, si no existe usa el conteo base
                        if "siguiente_numero_tpm" not in st.session_state or st.session_state["siguiente_numero_tpm"] is None:
                            st.session_state["siguiente_numero_tpm"] = obtener_siguiente_consecutivo_tpm()
            
                        num_actual = st.session_state["siguiente_numero_tpm"]
                        nuevo_id_tpm = f"TPM-{num_actual:04d}"  #<-- AQUÍ USA TU CONTADOR MANUAL (Ejemplo: TPM-0056)
                        
                        # 2. Registramos los datos de la nueva tarima
                        n_t = {"ID_Tarima": nuevo_id_tpm, "Tarima_Origen_Excel": t_orig, "Fecha_Creacion": datetime.datetime.now().strftime("%d/%m/%Y"), "Ubicacion_Actual": "Metales", "Creado_Por": oper, "Tipo_Tarima": tipo_t, "Estatus": "Disponible", "Es_Nueva": True}
                        st.session_state.BD_Tarimas = pd.concat([st.session_state.BD_Tarimas, pd.DataFrame([n_t])], ignore_index=True)
                        
                        # 3. Incrementamos el contador en +1 para la siguiente tarima de la lista
                        st.session_state["siguiente_numero_tpm"] += 1

                        items = df_ex[df_ex['Tarima'] == t_orig]
                        for _, item in items.iterrows():
                            st.session_state.BD_Detalle_Tarimas = pd.concat([st.session_state.BD_Detalle_Tarimas, pd.DataFrame([{"ID_Detalle": len(st.session_state.BD_Detalle_Tarimas) + 1, "ID_Tarima": nuevo_id_tpm, "SKU": item['Producto/SKU'], "PO": clean_po_val(item['PO']), "Proyecto": clean_project_val(item['Proyecto']), "Parcialidad": item['Parcialidad'], "Descripcion": item['Descripcion'], "Cantidad": item['Cantidad']}])], ignore_index=True)
                    subir_excel_a_github("BD_Tarimas.xlsx", st.session_state.BD_Tarimas)
                    subir_excel_a_github("BD_Detalle_Tarimas.xlsx", st.session_state.BD_Detalle_Tarimas)
                    # Eliminamos el archivo de override si existe, ya que ha sido consumido
                    import os
                    if os.path.exists("consecutivo_override.txt"):
                        try:
                            os.remove("consecutivo_override.txt")
                        except Exception:
                            pass
                    st.session_state["siguiente_numero_tpm"] = obtener_siguiente_consecutivo_tpm()
                    st.success("¡Inventario respaldado con éxito!"); st.rerun()
            except Exception as e: st.error(f"Error: {e}")
            
    if True:
        st.write("---")
        st.subheader("🖨️ Panel de Impresión Masiva de Tarimas")
        
        # Ordenar las tarimas consecutivamente de manera descendente por ID_Tarima
        df_tarimas_sorted = st.session_state.BD_Tarimas.copy()
        
        def extract_tpm_num(id_val):
            id_str = str(id_val).strip()
            partes = id_str.split('-')
            if len(partes) > 1 and partes[1].strip().isdigit():
                return int(partes[1].strip())
            elif id_str.startswith('TPM') and id_str[3:].strip().isdigit():
                return int(id_str[3:].strip())
            return 0
            
        df_tarimas_sorted['_sort_key'] = df_tarimas_sorted['ID_Tarima'].apply(extract_tpm_num)
        df_tarimas_sorted = df_tarimas_sorted.sort_values(by='_sort_key', ascending=False).drop(columns=['_sort_key']).reset_index(drop=True)
        
        df_estilado = df_tarimas_sorted.style.apply(lambda r: ['background-color: #FFF59D' if r['Es_Nueva'] else '' for _ in r], axis=1)
        seleccion_tabla = st.dataframe(df_estilado, use_container_width=True, column_order=["ID_Tarima", "Tarima_Origen_Excel", "Fecha_Creacion", "Ubicacion_Actual", "Creado_Por", "Tipo_Tarima", "Estatus"], on_select="rerun", selection_mode="multi-row")
        filas_seleccionadas = seleccion_tabla.get("selection", {}).get("rows", [])
        
        # =============================================================================
        # BLOQUE DE IMPRESIÓN MAESTRO DE TARIMAS BLINDADO CONTRA LLAVES HUÉRFANAS
        # =============================================================================
        if filas_seleccionadas:
            elegidas = df_tarimas_sorted.iloc[filas_seleccionadas]['ID_Tarima'].tolist()
            if len(elegidas) == 1:
                id_tarima_limpio = str(elegidas[0]).strip()  # Extracción nativa perfecta
        
                # 1. Intentamos buscar las piezas asociadas a esta tarima en la base de datos
                df_tarima_individual = pd.DataFrame()
                if "BD_Detalle_Tarimas" in st.session_state and not st.session_state.BD_Detalle_Tarimas.empty:
                    # Filtramos asegurando que compare texto puro
                    df_det = st.session_state.BD_Detalle_Tarimas[
                        st.session_state.BD_Detalle_Tarimas['ID_Tarima'].astype(str) == id_tarima_limpio
                    ].copy()
                    if not df_det.empty:
                        # Cruce relacional para anexar Estatus, Fecha_Creacion y Estatus_Envio requeridos por el PDF
                        df_tarima_individual = pd.merge(df_det, 
                                                         st.session_state.BD_Tarimas[['ID_Tarima', 'Estatus', 'Fecha_Creacion']], 
                                                         on='ID_Tarima', how='left')
                        df_tarima_individual['Estatus_Envio'] = df_tarima_individual['Estatus'].apply(lambda x: "Remesado" if x == "Remesada" else "No Remesado")
        
                # 2. VALIDACIÓN CRÍTICA: ¿Tiene artículos esta tarima para imprimir?
                if df_tarima_individual.empty or "SKU" not in df_tarima_individual.columns:
                    # Si no hay piezas asociadas, mostramos un mensaje limpio en lugar de lanzar la pantalla de error
                    st.warning(f"⚠️ La tarima **{id_tarima_limpio}** no contiene artículos o piezas asignadas en la base de datos de detalles. No se puede generar un reporte vacío.")
                else:
                    # Si sí tiene piezas con SKU válido, procedemos a compilar de forma segura
                    filtros_simulados = {
                        "PO": "Todos",
                        "SKU": "Todos",
                        "Proyecto": "Todos",
                        "Tarima": id_tarima_limpio,
                        "Parcialidad": "Todos",
                        "Estatus": "Todos",
                        "Descripcion": "Todos"
                    }
                    
                    try:
                        # Función auxiliar interna para generar la etiqueta de 2 hojas para esta tarima
                        def generar_pdf_etiqueta_tarima_individual(t_imp):
                            buf = io.BytesIO()
                            doc = SimpleDocTemplate(buf, pagesize=letter, leftMargin=36, rightMargin=36, topMargin=90, bottomMargin=60)
                            story_l, styles = [], getSampleStyleSheet()
                            
                            style_tarima_titulo = ParagraphStyle('T_Giga_Ind', parent=styles['Heading1'], fontName="Helvetica-Bold", fontSize=140, alignment=1, leading=150, textColor=colors.HexColor("#212121"))
                            style_sub_titulo = ParagraphStyle('S_Giga_Ind', parent=styles['Normal'], fontName="Helvetica-Bold", fontSize=26, alignment=1, textColor=colors.HexColor("#D32F2F"))
                            style_normal_bold = ParagraphStyle('N_Bold_Ind', parent=styles['Normal'], fontName="Helvetica-Bold", fontSize=11, leading=14)
                            style_normal_text = ParagraphStyle('N_Text_Ind', parent=styles['Normal'], fontSize=11, leading=14)
                            style_blanco_bold = ParagraphStyle('B_Bold_Ind', parent=styles['Normal'], fontName="Helvetica-Bold", textColor=colors.white, alignment=1, fontSize=10)
                            
                            det = st.session_state.BD_Detalle_Tarimas[st.session_state.BD_Detalle_Tarimas['ID_Tarima'] == t_imp]
                            t_info = st.session_state.BD_Tarimas[st.session_state.BD_Tarimas['ID_Tarima'] == t_imp]
                            
                            op_nom = t_info.iloc[0]['Creado_Por'] if not t_info.empty else "N/A"
                            fe_cre = t_info.iloc[0]['Fecha_Creacion'] if not t_info.empty else "N/A"
                            
                            # Buscar si alguna PO de la tarima tiene formato de color configurado
                            color_bg = None
                            color_fg = None
                            texto_etiqueta = None
                            
                            if "BD_POs_Cabecera" in st.session_state and not st.session_state.BD_POs_Cabecera.empty:
                                df_pos_cab = st.session_state.BD_POs_Cabecera
                                pos_tarima = det['PO'].dropna().unique().tolist()
                                for p in pos_tarima:
                                    p_upper = str(p).strip().upper()
                                    po_rows = df_pos_cab[df_pos_cab['PO'].astype(str).str.strip().str.upper() == p_upper]
                                    if not po_rows.empty:
                                        row = po_rows.iloc[0]
                                        if 'Color_Fondo' in df_pos_cab.columns and pd.notna(row.get('Color_Fondo')) and str(row.get('Color_Fondo')).strip():
                                            color_bg = str(row['Color_Fondo']).strip()
                                            color_fg = str(row.get('Color_Texto', '#FFFFFF')).strip()
                                            texto_etiqueta = str(row.get('Texto_Etiqueta', '')).strip()
                                            if not texto_etiqueta:
                                                texto_etiqueta = p
                                            break

                            # HOJA 1: CARÁTULA DE IDENTIFICACIÓN
                            story_l.append(Spacer(1, 1.2 * inch))
                            story_l.append(Paragraph("TARIMA", style_sub_titulo))
                            story_l.append(Spacer(1, 0.2 * inch))
                            
                            num_limpio = str(t_imp).split('-')[-1] if '-' in str(t_imp) else str(t_imp)
                            story_l.append(Paragraph(f"#{num_limpio}", style_tarima_titulo))
                            
                            if color_bg and texto_etiqueta:
                                story_l.append(Spacer(1, 0.6 * inch))
                            else:
                                story_l.append(Spacer(1, 1.5 * inch))
                            
                            tabla_base = Table([
                                [Paragraph("CÓDIGO DE IDENTIFICACIÓN:", style_normal_bold), Paragraph(str(t_imp), style_normal_text)],
                                [Paragraph("OPERADOR DE PLANTA:", style_normal_bold), Paragraph(str(op_nom), style_normal_text)],
                                [Paragraph("FECHA DE EMISIÓN:", style_normal_bold), Paragraph(str(fe_cre), style_normal_text)]
                            ], colWidths=[2.5 * inch, 5.0 * inch])
                            tabla_base.setStyle(TableStyle([
                                ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
                                ('LINEBELOW', (0,0), (-1,-1), 0.5, colors.HexColor("#E0E0E0")),
                                ('BOTTOMPADDING', (0,0), (-1,-1), 6)
                            ]))
                            story_l.append(tabla_base)

                            if color_bg and texto_etiqueta:
                                story_l.append(Spacer(1, 0.3 * inch))
                                style_color_tag = ParagraphStyle(
                                    'ColorTagText', 
                                    fontName="Helvetica-Bold", 
                                    fontSize=34, 
                                    leading=40, 
                                    alignment=1, 
                                    textColor=colors.HexColor(color_fg)
                                )
                                color_box_table = Table([[Paragraph(texto_etiqueta, style_color_tag)]], colWidths=[6.8 * inch], rowHeights=[1.0 * inch])
                                color_box_table.setStyle(TableStyle([
                                    ('BACKGROUND', (0,0), (-1,-1), colors.HexColor(color_bg)),
                                    ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
                                    ('ALIGN', (0,0), (-1,-1), 'CENTER'),
                                    ('GRID', (0,0), (-1,-1), 1.5, colors.HexColor("#EAB519")),
                                    ('BOTTOMPADDING', (0,0), (-1,-1), 10),
                                    ('TOPPADDING', (0,0), (-1,-1), 10)
                                ]))
                                story_l.append(color_box_table)

                            story_l.append(PageBreak())
                            
                            # HOJA 2: DETALLE DE MATERIALES ASOCIADOS
                            story_l.append(Spacer(1, 0.1 * inch))
                            story_l.append(Paragraph(f"<b>DETALLE DE MATERIALES ASOCIADOS - CONTROL #{t_imp}</b>", styles['Heading2']))
                            story_l.append(Spacer(1, 0.2 * inch))
                            
                            tabla_detalles = [[
                                Paragraph("ORDEN (PO)", style_blanco_bold),
                                Paragraph("SKU / PRODUCTO", style_blanco_bold),
                                Paragraph("DESCRIPCIÓN COMERCIAL", style_blanco_bold),
                                Paragraph("CANTIDAD", style_blanco_bold)
                            ]]
                            
                            for _, item in det.iterrows():
                                art = st.session_state.BD_Articulos[st.session_state.BD_Articulos['SKU'] == item['SKU']]
                                art_nom = art.iloc[0]['Nombre'] if not art.empty else "Articulo No Registrado en BD Remisiones"
                                
                                sku_partida = item['SKU']
                                # Buscar si existe una imagen cargada para este SKU
                                import glob
                                img_encontrada = None
                                matching_imgs = glob.glob(f"imagenes_articulos/{sku_partida}(*.*")
                                if matching_imgs:
                                    img_encontrada = matching_imgs[0]
                                else:
                                    if "github_token" in st.secrets and st.secrets["github_token"]:
                                        try:
                                            GITHUB_TOKEN = st.secrets["github_token"]
                                            url_list = f"https://api.github.com/repos/{REPO_OWNER}/{REPO_NAME}/contents/imagenes_articulos?ref={BRANCH}"
                                            headers = {"Authorization": f"token {GITHUB_TOKEN}", "Accept": "application/vnd.github.v3+json"}
                                            res_list = requests.get(url_list, headers=headers)
                                            if res_list.status_code == 200:
                                                items_git = res_list.json()
                                                for it in items_git:
                                                    if it["name"].startswith(f"{sku_partida}("):
                                                        github_file_path = f"imagenes_articulos/{it['name']}"
                                                        if descargar_imagen_desde_github(github_file_path):
                                                            img_encontrada = github_file_path
                                                            break
                                        except Exception:
                                            pass

                                desc_paragraph = Paragraph(str(art_nom), style_normal_text)
                                if img_encontrada and os.path.exists(img_encontrada):
                                    from reportlab.platypus import Image as RLImage
                                    img_flowable = RLImage(img_encontrada, width=75, height=75, hAlign='LEFT')
                                    sub_t = Table([[img_flowable, desc_paragraph]], colWidths=[80, 3.5 * inch - 80])
                                    sub_t.setStyle(TableStyle([
                                        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
                                        ('LEFTPADDING', (0,0), (-1,-1), 0),
                                        ('RIGHTPADDING', (0,0), (-1,-1), 0),
                                        ('TOPPADDING', (0,0), (-1,-1), 0),
                                        ('BOTTOMPADDING', (0,0), (-1,-1), 0)
                                    ]))
                                    desc_comercial_flowables = sub_t
                                else:
                                    desc_comercial_flowables = desc_paragraph

                                tabla_detalles.append([
                                    Paragraph(str(item['PO']), style_normal_text),
                                    Paragraph(str(item['SKU']), style_normal_text),
                                    desc_comercial_flowables,
                                    Paragraph(f"<b>{int(item['Cantidad'])}</b> PZS", style_normal_bold)
                                ])
                                
                            t_grid = Table(tabla_detalles, colWidths=[1.3 * inch, 1.5 * inch, 3.5 * inch, 1.2 * inch])
                            t_grid.setStyle(TableStyle([
                                ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#757575")),
                                ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor("#BDBDBD")),
                                ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
                                ('TOPPADDING', (0,0), (-1,-1), 6),
                                ('BOTTOMPADDING', (0,0), (-1,-1), 6)
                            ]))
                            story_l.append(t_grid)
                            
                            doc.build(story_l, onFirstPage=draw_sigrama_decorations, onLaterPages=draw_sigrama_decorations)
                            buf.seek(0)
                            return buf

                        # 3. Compilamos el PDF de reporte pasándole sus parámetros obligatorios
                        pdf_datos_compilados = generar_pdf_reporte_filtrado(filtros_simulados, df_tarima_individual)
                        
                        if hasattr(pdf_datos_compilados, 'getvalue'):
                            pdf_bytes_listos = pdf_datos_compilados.getvalue()
                        else:
                            pdf_bytes_listos = pdf_datos_compilados
                            
                        # 4. Generamos el PDF de etiqueta unificada de 2 hojas para esta tarima
                        pdf_etiqueta_bytes = generar_pdf_etiqueta(id_tarima_limpio).getvalue()
        
                        col_btns1, col_btns2 = st.columns(2)
                        with col_btns1:
                            st.download_button(
                                label=f"📄 Descargar Reporte PDF",
                                data=pdf_bytes_listos,
                                file_name=f"TAR_Reporte_Inventario_{id_tarima_limpio}.pdf",
                                mime="application/pdf",
                                key=f"btn_descarga_maestra_tarima_ind_{id_tarima_limpio}"
                            )
                        with col_btns2:
                            st.download_button(
                                label=f"🖨️ Descargar Etiqueta PDF (2 Hojas)",
                                data=pdf_etiqueta_bytes,
                                file_name=f"TAR_Etiqueta_Identificacion_{id_tarima_limpio}.pdf",
                                mime="application/pdf",
                                key=f"btn_descarga_etiqueta_ind_{id_tarima_limpio}"
                            )
                    except Exception as e:
                        st.error(f"⚠️ Error al procesar el diseño de ReportLab para la tarima {id_tarima_limpio}: {e}")

        



                
            else:

                if st.button("📦 Unificar Lote de Impresión"):
                    buf_1 = io.BytesIO()
                    doc_1 = SimpleDocTemplate(buf_1, pagesize=letter, leftMargin=36, rightMargin=36, topMargin=90, bottomMargin=60)
                    story_l, styles = [], getSampleStyleSheet()
                    
                    # Estilos tipográficos institucionales de gran impacto
                    style_tarima_titulo = ParagraphStyle('T_Giga', parent=styles['Heading1'], fontName="Helvetica-Bold", fontSize=140, alignment=1, leading=150, textColor=colors.HexColor("#212121"))
                    style_sub_titulo = ParagraphStyle('S_Giga', parent=styles['Normal'], fontName="Helvetica-Bold", fontSize=26, alignment=1, textColor=colors.HexColor("#D32F2F"))
                    style_normal_bold = ParagraphStyle('N_Bold', parent=styles['Normal'], fontName="Helvetica-Bold", fontSize=11, leading=14)
                    style_normal_text = ParagraphStyle('N_Text', parent=styles['Normal'], fontSize=11, leading=14)
                    style_blanco_bold = ParagraphStyle('B_Bold', parent=styles['Normal'], fontName="Helvetica-Bold", textColor=colors.white, alignment=1, fontSize=10)
                    
                    for t_imp in elegidas:
                        det = st.session_state.BD_Detalle_Tarimas[st.session_state.BD_Detalle_Tarimas['ID_Tarima'] == t_imp]
                        t_info = st.session_state.BD_Tarimas[st.session_state.BD_Tarimas['ID_Tarima'] == t_imp]
                        
                        op_nom = t_info.iloc[0]['Creado_Por'] if not t_info.empty else "N/A"
                        fe_cre = t_info.iloc[0]['Fecha_Creacion'] if not t_info.empty else "N/A"
                        
                        # =============================================================================
                        # HOJA 1: CARÁTULA DE IDENTIFICACIÓN MASIVA (NÚMERO GIGANTE)
                        # =============================================================================
                        story_l.append(Spacer(1, 1.2 * inch))
                        story_l.append(Paragraph("TARIMA", style_sub_titulo))
                        story_l.append(Spacer(1, 0.2 * inch))
                        
                        # Extraemos solo el número limpio (ej. si es TPM-0024 extrae 0024)
                        num_limpio = str(t_imp).split('-')[-1] if '-' in str(t_imp) else str(t_imp)
                        story_l.append(Paragraph(f"#{num_limpio}", style_tarima_titulo))
                        
                        story_l.append(Spacer(1, 1.5 * inch))
                        
                        # Mini panel inferior de trazabilidad en la carátula
                        tabla_base = Table([
                            [Paragraph("CÓDIGO DE IDENTIFICACIÓN:", style_normal_bold), Paragraph(str(t_imp), style_normal_text)],
                            [Paragraph("OPERADOR DE PLANTA:", style_normal_bold), Paragraph(str(op_nom), style_normal_text)],
                            [Paragraph("FECHA DE EMISIÓN:", style_normal_bold), Paragraph(str(fe_cre), style_normal_text)]
                        ], colWidths=[2.5 * inch, 5.0 * inch])
                        tabla_base.setStyle(TableStyle([
                            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
                            ('LINEBELOW', (0,0), (-1,-1), 0.5, colors.HexColor("#E0E0E0")),
                            ('BOTTOMPADDING', (0,0), (-1,-1), 6)
                        ]))
                        story_l.append(tabla_base)
                        
                        # Forzamos salto inmediato a la siguiente página
                        story_l.append(PageBreak())
                        
                        # =============================================================================
                        # HOJA 2: DESGLOSE GRANULAR Y DETALLE DE PIEZAS
                        # =============================================================================
                        story_l.append(Spacer(1, 0.1 * inch))
                        story_l.append(Paragraph(f"<b>DETALLE DE MATERIALES ASOCIADOS - CONTROL #{t_imp}</b>", styles['Heading2']))
                        story_l.append(Spacer(1, 0.2 * inch))
                        
                        # Construcción de cuadrícula formal de inventario
                        tabla_detalles = [[
                            Paragraph("ORDEN (PO)", style_blanco_bold),
                            Paragraph("SKU / PRODUCTO", style_blanco_bold),
                            Paragraph("DESCRIPCIÓN COMERCIAL", style_blanco_bold),
                            Paragraph("CANTIDAD", style_blanco_bold)
                        ]]

                    
                        for _, item in det.iterrows():
                            art = st.session_state.BD_Articulos[st.session_state.BD_Articulos['SKU'] == item['SKU']]
                            if not art.empty:
                                art_info = art.iloc[0]
                                nombre_com = str(art_info.get('Nombre', '')).strip()
                                calibre = str(art_info.get('Calibre_Espesor', '')).strip()
                                dims = str(art_info.get('Dimensiones_Pieza', '')).strip()
                                acabado = str(art_info.get('Acabado_Superficial', '')).strip()
                                
                                detalles = []
                                if calibre and calibre.lower() != 'nan' and calibre != '': detalles.append(f"<b>Calibre/Espesor:</b> {calibre}")
                                if dims and dims.lower() != 'nan' and dims != '': detalles.append(f"<b>Dimensiones:</b> {dims}")
                                if acabado and acabado.lower() != 'nan' and acabado != '': detalles.append(f"<b>Material/Acabado:</b> {acabado}")
                                
                                espec_str = f" | ".join(detalles)
                                if espec_str:
                                    art_nom = f"<b>{nombre_com}</b><br/><font color='#555555' size='8.5'>{espec_str}</font>"
                                else:
                                    art_nom = f"<b>{nombre_com}</b>"
                            else:
                                art_nom = "Articulo No Registrado en BD Remisiones"
                            
                            sku_partida = item['SKU']
                            # Buscar si existe una imagen cargada para este SKU
                            import glob
                            img_encontrada = None
                            matching_imgs = glob.glob(f"imagenes_articulos/{sku_partida}(*.*")
                            if matching_imgs:
                                img_encontrada = matching_imgs[0]
                            else:
                                if "github_token" in st.secrets and st.secrets["github_token"]:
                                    try:
                                        GITHUB_TOKEN = st.secrets["github_token"]
                                        url_list = f"https://api.github.com/repos/{REPO_OWNER}/{REPO_NAME}/contents/imagenes_articulos?ref={BRANCH}"
                                        headers = {"Authorization": f"token {GITHUB_TOKEN}", "Accept": "application/vnd.github.v3+json"}
                                        res_list = requests.get(url_list, headers=headers)
                                        if res_list.status_code == 200:
                                            items_git = res_list.json()
                                            for it in items_git:
                                                if it["name"].startswith(f"{sku_partida}("):
                                                    github_file_path = f"imagenes_articulos/{it['name']}"
                                                    if descargar_imagen_desde_github(github_file_path):
                                                        img_encontrada = github_file_path
                                                        break
                                    except Exception:
                                        pass

                            desc_paragraph = Paragraph(str(art_nom), style_normal_text)
                            if img_encontrada and os.path.exists(img_encontrada):
                                from reportlab.platypus import Image as RLImage
                                img_flowable = RLImage(img_encontrada, width=75, height=75, hAlign='LEFT')
                                sub_t = Table([[img_flowable, desc_paragraph]], colWidths=[80, 3.5 * inch - 80])
                                sub_t.setStyle(TableStyle([
                                    ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
                                    ('LEFTPADDING', (0,0), (-1,-1), 0),
                                    ('RIGHTPADDING', (0,0), (-1,-1), 0),
                                    ('TOPPADDING', (0,0), (-1,-1), 0),
                                    ('BOTTOMPADDING', (0,0), (-1,-1), 0)
                                ]))
                                desc_comercial_flowables = sub_t
                            else:
                                desc_comercial_flowables = desc_paragraph

                            tabla_detalles.append([
                                Paragraph(str(item['PO']), style_normal_text),
                                Paragraph(str(item['SKU']), style_normal_text),
                                desc_comercial_flowables,
                                Paragraph(f"<b>{int(item['Cantidad'])}</b> PZS", style_normal_bold)
                            ])
                            
                        t_grid = Table(tabla_detalles, colWidths=[1.3 * inch, 1.5 * inch, 3.5 * inch, 1.2 * inch])
                        t_grid.setStyle(TableStyle([
                            ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#757575")),
                            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor("#BDBDBD")),
                            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
                            ('TOPPADDING', (0,0), (-1,-1), 6),
                            ('BOTTOMPADDING', (0,0), (-1,-1), 6)
                        ]))
                        story_l.append(t_grid)
                        
                        # Salto de página para separar la siguiente tarima del lote
                        story_l.append(PageBreak())
                        
                    if story_l: 
                        story_l.pop()  # Remueve el último salto de página sobrante
                        
                    doc_1.build(story_l, onFirstPage=draw_sigrama_decorations, onLaterPages=draw_sigrama_decorations)
                    
                    st.download_button(
                        label="📥 Descargar Lote Completo (PDF)", 
                        data=buf_1.getvalue(), 
                        file_name=f"LOT_Lote_Tarimas_Separadas_{datetime.datetime.now().strftime('%Y%m%d')}.pdf",
                        mime="application/pdf",
                        key="btn_download_lote_tarimas_unificado_final_v2"
                    )


        # =============================================================================
        # NUEVA SECCIÓN: CONSULTA GRANULAR DE ARTÍCULOS POR TARIMA
        # =============================================================================
        st.write("---")
        st.write("### 🔍 Consulta de Contenido de Artículos por Tarima")
        st.info("💡 **Guía de uso:** Seleccione o busque una Tarima específica para desplegar de forma inmediata el listado de materiales y piezas que contiene.")

        # 1. Validar que la base de datos de tarimas tenga registros activos
        if "BD_Tarimas" in st.session_state and not st.session_state.BD_Tarimas.empty:
            
            # Obtener la lista de IDs de tarimas disponibles para el menú desplegable
            lista_tarimas_disponibles = sorted(st.session_state.BD_Tarimas["ID_Tarima"].unique())
            
            # Selector de Tarima para el usuario
            tarima_seleccionada = st.selectbox(
                "📋 Seleccione el ID de la Tarima a consultar:",
                options=lista_tarimas_disponibles,
                index=0,
                key="selector_granular_tarimas_modulo_t"
            )

            if tarima_seleccionada:
                # 2. Filtrar el detalle de materiales asociados a la tarima seleccionada
                if "BD_Detalle_Tarimas" in st.session_state and not st.session_state.BD_Detalle_Tarimas.empty:
                    
                    df_articulos_filtrados = st.session_state.BD_Detalle_Tarimas[
                        st.session_state.BD_Detalle_Tarimas["ID_Tarima"] == tarima_seleccionada
                    ].copy()

                    if not df_articulos_filtrados.empty:
                        st.write(f"##### 📦 Materiales contenidos en la Tarima: **{tarima_seleccionada}**")
                        
                        # Mostrar tabla limpia con datos clave para el operador
                        st.dataframe(
                            df_articulos_filtrados[["SKU", "PO", "Proyecto", "Descripcion", "Cantidad"]],
                            use_container_width=True,
                            hide_index=True
                        )
                        
                        # Resumen de conteo rápido para auditorías
                        total_pzs_tarima = int(df_articulos_filtrados["Cantidad"].sum())
                        st.success(f"📊 Total de piezas contabilizadas en esta tarima: **{total_pzs_tarima} Pzs**")
                    else:
                        st.warning(f"⚠️ La tarima **{tarima_seleccionada}** está registrada pero no cuenta con un desglose de piezas asociado en la base de datos.")
                else:
                    st.error("❌ Error de sistema: No se encuentra cargada la Base de Datos granular de Detalles de Tarimas.")
        else:
            st.warning("⚠️ No existen tarimas registradas en el sistema para realizar consultas de contenido.")

        # =============================================================================
        # NUEVA SECCIÓN: CONFIGURACIÓN DE ETIQUETAS DE COLOR POR PO
        # =============================================================================
        st.write("---")
        with st.expander("🎨 Configuración de Etiquetas de Color por Orden de Compra (PO)"):
            st.info("💡 **Guía de uso:** Seleccione una PO para definir o modificar el color de fondo, color de texto y la descripción sencilla que aparecerá impresa en la portada de la etiqueta de la tarima.")
            
            # Obtener todas las POs de cabecera y del detalle de tarimas para asegurar que listamos todas
            all_pos = set()
            if "BD_POs_Cabecera" in st.session_state and not st.session_state.BD_POs_Cabecera.empty:
                all_pos.update(st.session_state.BD_POs_Cabecera["PO"].dropna().astype(str).str.strip().str.upper().unique().tolist())
            if "BD_Detalle_Tarimas" in st.session_state and not st.session_state.BD_Detalle_Tarimas.empty:
                all_pos.update(st.session_state.BD_Detalle_Tarimas["PO"].dropna().astype(str).str.strip().str.upper().unique().tolist())
            
            lista_pos_disponibles = sorted(list(all_pos))
            
            if not lista_pos_disponibles:
                st.warning("⚠️ No hay Órdenes de Compra (POs) registradas en el sistema para configurar.")
            else:
                # Inicializar el estado de la sesión para las POs en edición activa
                if "pos_en_edicion" not in st.session_state:
                    st.session_state.pos_en_edicion = []

                # Construir la lista completa de detalles de todas las POs del sistema
                filas_po_todas = []
                df_pos_cab = st.session_state.BD_POs_Cabecera
                
                for p_name in lista_pos_disponibles:
                    meta = {
                        "PO": p_name, 
                        "Proyecto": "N/A", 
                        "Solicitante": "N/A", 
                        "Destino": "N/A", 
                        "Cant Recibida": 0,
                        "Etiqueta Actual": "Sin etiqueta"
                    }
                    
                    po_rows = df_pos_cab[df_pos_cab["PO"].astype(str).str.strip().str.upper() == p_name]
                    if not po_rows.empty:
                        r = po_rows.iloc[0]
                        meta.update({
                            "Proyecto": str(r.get("Proyecto", "N/A")) if pd.notna(r.get("Proyecto")) else "N/A",
                            "Solicitante": str(r.get("Solicitante", "N/A")) if pd.notna(r.get("Solicitante")) else "N/A",
                            "Destino": str(r.get("Destino", "N/A")) if pd.notna(r.get("Destino")) else "N/A"
                        })
                        if "Texto_Etiqueta" in df_pos_cab.columns and pd.notna(r.get("Texto_Etiqueta")) and str(r.get("Texto_Etiqueta")).strip():
                            meta["Etiqueta Actual"] = str(r["Texto_Etiqueta"]).strip()
                    
                    if "BD_Detalle_Tarimas" in st.session_state and not st.session_state.BD_Detalle_Tarimas.empty:
                        df_det = st.session_state.BD_Detalle_Tarimas
                        df_det_po = df_det[df_det["PO"].astype(str).str.strip().str.upper() == p_name]
                        if not df_det_po.empty:
                            meta["Cant Recibida"] = int(df_det_po["Cantidad"].sum())
                    
                    filas_po_todas.append(meta)
                    
                df_pos_completo = pd.DataFrame(filas_po_todas)
                
                # Filtrar las POs que ya están en el grupo de edición para que desaparezcan de la primera tabla
                df_disponibles = df_pos_completo[~df_pos_completo["PO"].isin(st.session_state.pos_en_edicion)].reset_index(drop=True)
                
                # --- PASO 1: TABLA DISPONIBLE ---
                st.write("##### 📋 Paso 1: Seleccione Órdenes de Compra (POs) desde la Lista:")
                st.info("💡 **Guía:** Seleccione las casillas de la izquierda en la tabla para elegir las POs que desea configurar y presione **'Añadir POs al Grupo de Edición'**.")
                
                if df_disponibles.empty:
                    st.write("✨ *Todas las POs están actualmente en el grupo de edición o no hay más POs por configurar.*")
                else:
                    seleccion_df = st.dataframe(
                        df_disponibles,
                        use_container_width=True,
                        column_order=["PO", "Proyecto", "Solicitante", "Destino", "Cant Recibida", "Etiqueta Actual"],
                        on_select="rerun",
                        selection_mode="multi-row",
                        key="po_color_step1_df_selection_unique"
                    )
                    
                    filas_elegidas = seleccion_df.get("selection", {}).get("rows", [])
                    
                    if filas_elegidas:
                        pos_elegidas_nombres = df_disponibles.iloc[filas_elegidas]["PO"].tolist()
                        if st.button("➕ Añadir POs al Grupo de Edición", use_container_width=True, key="btn_add_pos_to_active_group"):
                            for p in pos_elegidas_nombres:
                                if p not in st.session_state.pos_en_edicion:
                                    st.session_state.pos_en_edicion.append(p)
                            st.rerun()
                            
                # --- PASO 2: EDICIÓN DEL GRUPO ---
                if st.session_state.pos_en_edicion:
                    st.write("---")
                    st.write("##### 🎨 Paso 2: Configurar Texto y Colores para el Grupo Seleccionado")
                    
                    col_clear1, col_clear2 = st.columns([5, 1])
                    with col_clear1:
                        st.success(f"📦 **POs activas en este grupo de etiqueta:** {', '.join(st.session_state.pos_en_edicion)}")
                    with col_clear2:
                        if st.button("🗑️ Limpiar Grupo", use_container_width=True, key="btn_clear_active_po_group"):
                            st.session_state.pos_en_edicion = []
                            st.rerun()
                            
                    # Selector adicional en el Paso 2 para agregar más POs directamente
                    pos_adicionales_opciones = [p for p in lista_pos_disponibles if p not in st.session_state.pos_en_edicion]
                    
                    pos_adicionales = st.multiselect(
                        "➕ Seleccione POs adicionales para agregar a este mismo grupo:",
                        options=pos_adicionales_opciones,
                        key="po_color_step2_add_more_selector"
                    )
                    
                    if pos_adicionales:
                        if st.button("➕ Agregar POs adicionales", key="btn_add_more_pos_step2"):
                            st.session_state.pos_en_edicion.extend(pos_adicionales)
                            st.rerun()
                            
                    # Valores por defecto para pre-poblar basados en la primera PO
                    val_text = ""
                    val_bg = "#85A3D4" # Azul claro por defecto
                    val_fg = "#000000" # Negro por defecto
                    
                    if not df_pos_cab.empty:
                        po_first = st.session_state.pos_en_edicion[0]
                        po_first_rows = df_pos_cab[df_pos_cab["PO"].astype(str).str.strip().str.upper() == po_first]
                        if not po_first_rows.empty:
                            row = po_first_rows.iloc[0]
                            if "Texto_Etiqueta" in df_pos_cab.columns and pd.notna(row.get("Texto_Etiqueta")):
                                val_text = str(row["Texto_Etiqueta"]).strip()
                            if "Color_Fondo" in df_pos_cab.columns and pd.notna(row.get("Color_Fondo")):
                                val_bg = str(row["Color_Fondo"]).strip()
                            if "Color_Texto" in df_pos_cab.columns and pd.notna(row.get("Color_Texto")):
                                val_fg = str(row["Color_Texto"]).strip()
                                
                    col_c1, col_c2, col_c3 = st.columns([2, 1, 1])
                    with col_c1:
                        tag_text_input = st.text_input(
                            "📝 Texto de la Etiqueta (Ej. LC8, Reno 6):", 
                            value=val_text if val_text else st.session_state.pos_en_edicion[0], 
                            key="po_color_config_tag_text",
                            max_chars=25
                        )
                    with col_c2:
                        color_bg_input = st.color_picker(
                            "🎨 Color de Fondo:", 
                            value=val_bg, 
                            key="po_color_config_bg_picker"
                        )
                    with col_c3:
                        color_fg_input = st.color_picker(
                            "✍️ Color del Texto:", 
                            value=val_fg, 
                            key="po_color_config_fg_picker"
                        )
                        
                    # Vista previa visual interactiva
                    st.write("**Vista previa de la etiqueta:**")
                    st.markdown(
                        f"""
                        <div style="
                            background-color: {color_bg_input}; 
                            color: {color_fg_input}; 
                            border: 2px solid #EAB519; 
                            border-radius: 4px; 
                            padding: 20px; 
                            text-align: center; 
                            font-family: Arial, sans-serif; 
                            font-size: 28px; 
                            font-weight: bold; 
                            width: 100%; 
                            max-width: 400px;
                            margin: 10px 0;
                        ">
                            {tag_text_input}
                        </div>
                        """, 
                        unsafe_allow_html=True
                    )
                    
                    if st.button("💾 Guardar Configuración de Color para PO", use_container_width=True, key="btn_save_po_color"):
                        # 1. Asegurar que las columnas existen en st.session_state.BD_POs_Cabecera
                        for col in ["Texto_Etiqueta", "Color_Fondo", "Color_Texto"]:
                            if col not in df_pos_cab.columns:
                                df_pos_cab[col] = ""
                                
                        # 2. Guardar para todas las POs del grupo activo
                        for p_sel in st.session_state.pos_en_edicion:
                            idx_po = df_pos_cab[df_pos_cab["PO"].astype(str).str.strip().str.upper() == p_sel].index
                            if not idx_po.empty:
                                df_pos_cab.loc[idx_po, "Texto_Etiqueta"] = tag_text_input
                                df_pos_cab.loc[idx_po, "Color_Fondo"] = color_bg_input
                                df_pos_cab.loc[idx_po, "Color_Texto"] = color_fg_input
                            else:
                                nueva_fila = {
                                    "PO": p_sel,
                                    "Fecha_Pedido": "",
                                    "Proyecto": "",
                                    "Solicitante": "",
                                    "Requisicion": "",
                                    "Destino": "",
                                    "Texto_Etiqueta": tag_text_input,
                                    "Color_Fondo": color_bg_input,
                                    "Color_Texto": color_fg_input
                                }
                                df_pos_cab = pd.concat([df_pos_cab, pd.DataFrame([nueva_fila])], ignore_index=True)
                                
                        st.session_state.BD_POs_Cabecera = df_pos_cab
                        
                        # 3. Guardar cambios en el archivo Excel y en GitHub
                        subir_excel_a_github("BD_POs_Cabecera.xlsx", st.session_state.BD_POs_Cabecera)
                        st.success("✅ ¡Configuración de color guardada y sincronizada correctamente para las POs seleccionadas!")
                        st.session_state.pos_en_edicion = [] # Limpiar el grupo de edición
                        st.rerun()








# =============================================================================
# 12. MÓDULO DE DESPACHOS LOGÍSTICOS: EMISIÓN DE REMISIONES DE SALIDA
# =============================================================================
elif opcion_menu == "🚚 Módulo Remisiones":
    st.title("🚚 Generación de Remisiones de Salida")
    t_disp = st.session_state.BD_Tarimas[st.session_state.BD_Tarimas['Estatus'] == 'Disponible']['ID_Tarima'].tolist()
    if not t_disp: st.warning("⚠️ No existen tarimas disponibles en patio.")
    else:
        t_sel = st.multiselect("Seleccione Tarimas para este Envío:", options=t_disp)
        col_e, col_r = st.columns(2)
        with col_e:
            # Lista desplegable acoplada de forma dinámica al catálogo relacional
            if "BD_Lideres" in st.session_state and not st.session_state.BD_Lideres.empty:
                lista_nombres_lideres = st.session_state.BD_Lideres['Nombre_Lider'].unique().tolist()
                nom_e = st.selectbox("Líder / Emisor Autorizado:", options=lista_nombres_lideres, key="rem_lider_sel_unique")
            else:
                nom_e = st.selectbox("Líder / Emisor Autorizado:", options=["Jesus Morales", "Supervisor General"], key="rem_lider_backup_unique")
            dir_e = st.text_input("Almacén de Origen:", "Metales")
        with col_r:
            if "BD_Receptores" in st.session_state and not st.session_state.BD_Receptores.empty:
                df_receptores_activos = st.session_state.BD_Receptores[st.session_state.BD_Receptores['Estatus'] == 'Activo']
                lista_nombres_receptores = df_receptores_activos['Nombre_Receptor'].unique().tolist()
                
                nom_r_sel = st.selectbox("Receptor / Cliente:", options=lista_nombres_receptores, key="rem_receptor_sel_unique")
                
                # Obtener dirección correspondiente
                dir_r_default = df_receptores_activos[df_receptores_activos['Nombre_Receptor'] == nom_r_sel]['Direccion'].values[0]
                dir_r = st.text_input("Dirección Destino:", value=dir_r_default)
                nom_r = nom_r_sel
            else:
                nom_r = st.text_input("Receptor / Cliente:", "Galvatec Industrias")
                dir_r = st.text_input("Dirección Destino:", "Prol. Valle Guadiana 919, Parque Industrial II, 35078 Gómez Palacio, Dgo.")
        if not is_admin: st.error("🔒 Operación Bloqueada: Se requiere contraseña de Administrador.")
        else:
            if st.button("🚀 Confirmar Salida y Generar Nueva Remisión"):
                if not t_sel or not nom_r: st.error("Complete los campos obligatorios.")
                else:
                    fol = f"E00{27 + len(st.session_state.BD_Datos_Generales_Remision)}"
                    reg = {"ID_Remision": len(st.session_state.BD_Datos_Generales_Remision)+1, "Folio_Remision": fol, "Fecha_Hora_Salida": datetime.datetime.now().strftime("%d/%m/%Y"), "Nombre_Emisor": nom_e, "Direccion_Emisor": dir_e, "Nombre_Receptor": nom_r, "Direccion_Receptor": dir_r, "Tarimas_Asociadas": str(t_sel)}
                    st.session_state.BD_Datos_Generales_Remision = pd.concat([st.session_state.BD_Datos_Generales_Remision, pd.DataFrame([reg])], ignore_index=True)
                    st.session_state.BD_Tarimas.loc[st.session_state.BD_Tarimas['ID_Tarima'].isin(t_sel), 'Estatus'] = 'Remesada'
                    subir_excel_a_github("BD_Tarimas.xlsx", st.session_state.BD_Tarimas)
                    subir_excel_a_github("BD_Datos_Generales_Remision.xlsx", st.session_state.BD_Datos_Generales_Remision)
                    st.success(f"✅ ¡Remisión {fol} Generada y Guardada de Forma Permanente!"); st.rerun()
                    
        # =============================================================================
    # INTERFAZ DE DESCARGA DE REMISIONES CORREGIDA Y GARANTIZADA (FO-MET-10)
    # =============================================================================
    if not st.session_state.BD_Datos_Generales_Remision.empty:
        st.write("---")
        st.subheader("📋 Descarga Documental de Remisiones")
        
        # Filtros de Fechas
        col_f1, col_f2 = st.columns(2)
        with col_f1:
            fecha_inicio = st.date_input("Fecha de Inicio:", value=datetime.date.today() - datetime.timedelta(days=30), key="rem_download_start_date")
        with col_f2:
            fecha_fin = st.date_input("Fecha de Fin:", value=datetime.date.today(), key="rem_download_end_date")
            
        # Parseo de fechas y filtrado
        df_remisiones = st.session_state.BD_Datos_Generales_Remision.copy()
        
        def parsear_fecha_rem(val):
            import datetime as dt
            try:
                return dt.datetime.strptime(str(val).split()[0], "%d/%m/%Y").date()
            except Exception:
                try:
                    return dt.datetime.strptime(str(val).split()[0], "%Y-%m-%d").date()
                except Exception:
                    return dt.date.today()
                    
        df_remisiones['Fecha_Date'] = df_remisiones['Fecha_Hora_Salida'].apply(parsear_fecha_rem)
        df_remisiones_filtradas = df_remisiones[
            (df_remisiones['Fecha_Date'] >= fecha_inicio) & 
            (df_remisiones['Fecha_Date'] <= fecha_fin)
        ]
        
        if df_remisiones_filtradas.empty:
            st.info("ℹ️ No se encontraron remisiones en el rango de fechas seleccionado.")
        else:
            df_mostrar = df_remisiones_filtradas.drop(columns=['Fecha_Date'], errors='ignore')
            
            # Ordenar descendentemente por número de folio
            def extract_folio_num(folio_str):
                import re
                numeros = re.findall(r'\d+', str(folio_str))
                if numeros:
                    return int(numeros[-1])
                return 0
                
            df_mostrar['Folio_Num'] = df_mostrar['Folio_Remision'].apply(extract_folio_num)
            df_mostrar = df_mostrar.sort_values(by='Folio_Num', ascending=False).drop(columns=['Folio_Num'], errors='ignore')
            
            st.write("Seleccione una remisión de la tabla para habilitar la descarga:")
            
            # Normalizar TODOS los tipos de columnas para evitar pyarrow.lib.ArrowInvalid
            for col_m in df_mostrar.columns:
                if df_mostrar[col_m].dtype == object:
                    df_mostrar[col_m] = df_mostrar[col_m].astype(str)
                
            sel_grid = st.dataframe(
                df_mostrar,
                use_container_width=True,
                hide_index=True,
                on_select="rerun",
                selection_mode="multi-row",
                key="tabla_descarga_remisiones_final",
                column_config={
                    "Folio_Remision": st.column_config.TextColumn("Folio de Remisión"),
                    "Fecha_Hora_Salida": st.column_config.TextColumn("Fecha de Salida"),
                    "Nombre_Emisor": st.column_config.TextColumn("Emisor"),
                    "Nombre_Receptor": st.column_config.TextColumn("Receptor"),
                    "Tarimas_Asociadas": st.column_config.TextColumn("Tarimas Asociadas")
                }
            )
            
            filas_seleccionadas = sel_grid.get("selection", {}).get("rows", [])
            if filas_seleccionadas:
                selected_rows = [df_mostrar.iloc[idx].to_dict() for idx in filas_seleccionadas]
                
                # Convertimos las tarimas asociadas de texto a lista real de Python consolidada
                import ast
                import zipfile
                import io
                
                tarimas_lista = []
                folios_lista = []
                receptores_set = set()
                
                for row_dict in selected_rows:
                    folios_lista.append(row_dict['Folio_Remision'])
                    receptores_set.add(row_dict.get('Nombre_Receptor', 'N/A'))
                    
                    t_asoc = row_dict['Tarimas_Asociadas']
                    if isinstance(t_asoc, str):
                        try:
                            t_asoc_list = ast.literal_eval(t_asoc)
                        except Exception:
                            t_asoc_list = [t_asoc]
                    else:
                        t_asoc_list = t_asoc if isinstance(t_asoc, list) else []
                    
                    for t in t_asoc_list:
                        if t not in tarimas_lista:
                            tarimas_lista.append(t)
                
                # Consolidar detalles
                df_det = st.session_state.BD_Detalle_Tarimas[
                    st.session_state.BD_Detalle_Tarimas['ID_Tarima'].isin(tarimas_lista)
                ]
                
                folios_str = ", ".join(folios_lista)
                key_suffix = f"{len(selected_rows)}_{folios_lista[0]}"
                
                # --- PREPARACIÓN DE ADJUNTOS EN SEGUNDO PLANO ---
                adjuntos_dict = {}
                
                # 1. Remisiones por separado en PDF
                for row_dict in selected_rows:
                    f_rem = row_dict['Folio_Remision']
                    t_rem_asoc = row_dict['Tarimas_Asociadas']
                    if isinstance(t_rem_asoc, str):
                        try:
                            t_rem_asoc_list = ast.literal_eval(t_rem_asoc)
                        except Exception:
                            t_rem_asoc_list = [t_rem_asoc]
                    else:
                        t_rem_asoc_list = t_rem_asoc if isinstance(t_rem_asoc, list) else []
                        
                    df_det_rem = st.session_state.BD_Detalle_Tarimas[
                        st.session_state.BD_Detalle_Tarimas['ID_Tarima'].isin(t_rem_asoc_list)
                    ]
                    adjuntos_dict[f"Remision_{f_rem}.pdf"] = generar_pdf_remision_general(row_dict, df_det_rem)
                
                # 2. Las etiquetas se agrupan en un archivo ZIP
                label_pdfs = {}
                for idx_tarima in tarimas_lista:
                    t_row_match = st.session_state.BD_Tarimas[st.session_state.BD_Tarimas['ID_Tarima'] == idx_tarima]
                    if not t_row_match.empty:
                        # Generamos el PDF de la etiqueta en memoria
                        pdf_etiqueta = generar_pdf_etiqueta(idx_tarima).getvalue()
                        label_pdfs[f"Etiqueta_{idx_tarima}.pdf"] = pdf_etiqueta
                
                # Comprimir todas las etiquetas en un ZIP en memoria
                if label_pdfs:
                    zip_buffer = io.BytesIO()
                    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
                        for filename, file_bytes in label_pdfs.items():
                            zip_file.writestr(filename, file_bytes)
                    zip_buffer.seek(0)
                    adjuntos_dict["Etiquetas_Tarimas.zip"] = zip_buffer

                # --- CONFIGURAR CORREO ---
                cfg_emails = obtener_emails_config()
                
                # --- RENDERIZADO COMPACTO DE ACCIONES (SIN SCROLL) ---
                col_acc_left, col_acc_right = st.columns([1.1, 1.3])
                
                with col_acc_left:
                    st.write("")
                    st.markdown(f"📌 **Folio(s) seleccionado(s):** `{folios_str}`")
                    if len(selected_rows) == 1:
                        row_dict = selected_rows[0]
                        r_sel = row_dict['Folio_Remision']
                        st.download_button(
                            label="📄 Descargar Reporte Oficial (PDF)", 
                            data=generar_pdf_remision_general(row_dict, df_det), 
                            file_name=f"Remision_{r_sel}.pdf", 
                            key=f"btn_dl_rem_pdf_{r_sel}",
                            mime="application/pdf",
                            use_container_width=True
                        )
                        st.download_button(
                            label="📥 Descargar Reporte Oficial (Excel)", 
                            data=generar_excel_remision(row_dict, df_det).getvalue(), 
                            file_name=f"Remision_{r_sel}.xlsx", 
                            key=f"btn_dl_rem_xlsx_{r_sel}",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            use_container_width=True
                        )
                    else:
                        st.info("💡 Selección múltiple. Descargas individuales (PDF/Excel) no disponibles.")
                
                with col_acc_right:
                    col_fields1, col_fields2 = st.columns(2)
                    with col_fields1:
                        eml_to = st.text_input("Para:", value=cfg_emails.get("dest_to", ""), key=f"eml_to_{key_suffix}")
                    with col_fields2:
                        eml_cc = st.text_input("CC:", value=cfg_emails.get("dest_cc", ""), key=f"eml_cc_{key_suffix}")
                        
                    # Construir el archivo .eml al vuelo
                    if len(selected_rows) == 1:
                        eml_subject = f"Envío de Remisión {folios_lista[0]} - Industria Sigrama"
                    else:
                        eml_subject = f"Envío de Remisiones {folios_str} - Industria Sigrama"
                    
                    eml_body = generar_cuerpo_correo_html(selected_rows, df_det)
                    eml_bytes = generar_archivo_eml(eml_to, eml_cc, eml_subject, eml_body, adjuntos_dict)
                    
                    st.download_button(
                        label="📩 Descargar Borrador de Correo (.eml)",
                        data=eml_bytes,
                        file_name=f"Correo_Remisiones_{key_suffix}.eml",
                        mime="message/rfc822",
                        use_container_width=True,
                        key=f"btn_dl_eml_{key_suffix}"
                    )
            else:
                st.info("💡 Seleccione una o varias filas marcando las casillas del extremo izquierdo de las remisiones para descargar el reporte o generar el correo borrador.")




# =============================================================================
# SECCIÓN 17A: PANEL DE MANTENIMIENTO AVANZADO Y EDICIÓN EN CALIENTE (SUPERUSER)
# =============================================================================
elif opcion_menu == "📦 Catálogo de Artículos":
    st.title("📦 Catálogo Maestro de Artículos")
    st.markdown("Consulte y filtre de forma dinámica el catálogo oficial de productos cargados en el sistema:")

    if "BD_Articulos" in st.session_state and not st.session_state.BD_Articulos.empty:
        df_articulos_base = st.session_state.BD_Articulos.copy()

        # Cuadrícula de filtros dinámicos (3 columnas en paralelo para optimizar espacio)
        art_col1, art_col2, art_col3 = st.columns(3)
        
        with art_col1:
            # Filtro por SKU
            opc_art_sku = ["Todos"] + sorted(df_articulos_base['SKU'].dropna().unique().tolist())
            f_art_sku = st.selectbox("Filtrar por SKU:", opc_art_sku, key="filter_art_sku")
            
            # Filtro por Dimensiones
            opc_art_dim = ["Todos"] + sorted(df_articulos_base['Dimensiones_Pieza'].dropna().unique().tolist())
            f_art_dim = st.selectbox("Filtrar por Dimensiones:", opc_art_dim, key="filter_art_dim")

        with art_col2:
            # Filtro por Nombre / Descripción Comercial
            opc_art_nom = ["Todos"] + sorted(df_articulos_base['Nombre'].dropna().unique().tolist())
            f_art_nom = st.selectbox("Filtrar por Nombre:", opc_art_nom, key="filter_art_nom")

        with art_col3:
            # Filtro por Calibre / Espesor
            opc_art_cal = ["Todos"] + sorted(df_articulos_base['Calibre_Espesor'].dropna().unique().tolist())
            f_art_cal = st.selectbox("Filtrar por Calibre / Espesor:", opc_art_cal, key="filter_art_cal")
            
            # Filtro por Acabado Superficial
            opc_art_acab = ["Todos"] + sorted(df_articulos_base['Acabado_Superficial'].dropna().unique().tolist())
            f_art_acab = st.selectbox("Filtrar por Acabado Superficial:", opc_art_acab, key="filter_art_acab")

        # Aplicación en cascada de los filtros seleccionados
        df_art_filtrado = df_articulos_base.copy()
        if f_art_sku != "Todos":
            df_art_filtrado = df_art_filtrado[df_art_filtrado['SKU'] == f_art_sku]
        if f_art_nom != "Todos":
            df_art_filtrado = df_art_filtrado[df_art_filtrado['Nombre'] == f_art_nom]
        if f_art_cal != "Todos":
            df_art_filtrado = df_art_filtrado[df_art_filtrado['Calibre_Espesor'] == f_art_cal]
        if f_art_dim != "Todos":
            df_art_filtrado = df_art_filtrado[df_art_filtrado['Dimensiones_Pieza'] == f_art_dim]
        if f_art_acab != "Todos":
            df_art_filtrado = df_art_filtrado[df_art_filtrado['Acabado_Superficial'] == f_art_acab]

        st.write("---")
        
        # Despliegue de métricas rápidas de la consulta
        col_metric, col_pdf = st.columns([2, 1])
        with col_metric:
            st.metric("🔢 Artículos en Selección:", f"{len(df_art_filtrado)} Items")
        with col_pdf:
            st.write("") # Alineación vertical
            pdf_data = generar_pdf_catalogo_articulos(df_art_filtrado)
            st.download_button(
                label="📄 Descargar Catálogo (PDF)",
                data=pdf_data,
                file_name="Reporte_Catalogo_Articulos.pdf",
                mime="application/pdf",
                key="btn_download_catalogo_pdf_tab",
                use_container_width=True
            )
        
        # Despliegue de la tabla de datos estructurada
        st.dataframe(
            df_art_filtrado, 
            use_container_width=True, 
            hide_index=True,
            column_config={
                "SKU": st.column_config.TextColumn("SKU / Código"),
                "Nombre": st.column_config.TextColumn("Descripción Comercial"),
                "Calibre_Espesor": st.column_config.TextColumn("Calibre / Espesor"),
                "Dimensiones_Pieza": st.column_config.TextColumn("Dimensiones de la Pieza"),
                "Acabado_Superficial": st.column_config.TextColumn("Acabado Superficial")
            }
        )

        # --- SECCIÓN DE AUDITORÍA DE INFORMACIÓN FALTANTE ---
        st.write("---")
        st.subheader("🔍 Auditoría de Información Faltante en el Catálogo")
        
        # Definir las columnas que deben tener información completa
        columnas_auditadas = ["Nombre", "Calibre_Espesor", "Dimensiones_Pieza", "Acabado_Superficial"]
        etiquetas_col = {
            "Nombre": "Descripción Comercial",
            "Calibre_Espesor": "Calibre / Espesor", 
            "Dimensiones_Pieza": "Dimensiones",
            "Acabado_Superficial": "Acabado Superficial"
        }
        
        df_auditoria = df_articulos_base.copy()
        
        # Detectar campos vacíos: None, NaN, "None", cadena vacía, "N/A"
        def campo_vacio(val):
            if pd.isna(val):
                return True
            s = str(val).strip().upper()
            return s in ["", "NONE", "N/A", "NAN", "NA", "-"]
        
        # Crear columna de conteo de campos faltantes por artículo
        df_auditoria["_campos_faltantes"] = 0
        df_auditoria["_detalle_faltante"] = ""
        
        for _, row in df_auditoria.iterrows():
            faltantes = []
            for col in columnas_auditadas:
                if col in row.index and campo_vacio(row[col]):
                    faltantes.append(etiquetas_col.get(col, col))
            df_auditoria.at[row.name, "_campos_faltantes"] = len(faltantes)
            df_auditoria.at[row.name, "_detalle_faltante"] = ", ".join(faltantes) if faltantes else "✅ Completo"
        
        df_incompletos = df_auditoria[df_auditoria["_campos_faltantes"] > 0].copy()
        df_completos = df_auditoria[df_auditoria["_campos_faltantes"] == 0].copy()
        
        # Métricas de auditoría
        col_aud1, col_aud2, col_aud3, col_aud4 = st.columns(4)
        with col_aud1:
            st.metric("📦 Total Artículos", len(df_auditoria))
        with col_aud2:
            st.metric("✅ Completos", len(df_completos))
        with col_aud3:
            st.metric("⚠️ Incompletos", len(df_incompletos))
        with col_aud4:
            pct = round((len(df_completos) / max(len(df_auditoria), 1)) * 100, 1)
            st.metric("📊 % Cumplimiento", f"{pct}%")
        
        if not df_incompletos.empty:
            # Desglose por campo faltante
            st.write("")
            st.markdown("**Desglose de campos faltantes:**")
            col_det1, col_det2, col_det3, col_det4 = st.columns(4)
            for i, col in enumerate(columnas_auditadas):
                cnt = sum(1 for _, r in df_incompletos.iterrows() if campo_vacio(r.get(col, "")))
                with [col_det1, col_det2, col_det3, col_det4][i]:
                    st.metric(f"Sin {etiquetas_col[col]}", cnt)
            
            st.write("")
            
            # Tabla de artículos incompletos
            st.warning(f"⚠️ Se detectaron **{len(df_incompletos)} artículos** con información faltante. Revise la tabla y descargue el Excel para completar los datos.")
            
            df_mostrar = df_incompletos[["SKU"] + columnas_auditadas + ["_detalle_faltante"]].copy()
            df_mostrar = df_mostrar.rename(columns={"_detalle_faltante": "Campos Faltantes"})
            
            st.dataframe(
                df_mostrar,
                use_container_width=True,
                hide_index=True,
                column_config={
                    "SKU": st.column_config.TextColumn("SKU / Código"),
                    "Nombre": st.column_config.TextColumn("Descripción Comercial"),
                    "Calibre_Espesor": st.column_config.TextColumn("Calibre / Espesor"),
                    "Dimensiones_Pieza": st.column_config.TextColumn("Dimensiones"),
                    "Acabado_Superficial": st.column_config.TextColumn("Acabado Superficial"),
                    "Campos Faltantes": st.column_config.TextColumn("Campos Faltantes")
                }
            )
            
            # Generar Excel descargable con los artículos incompletos (formato de plantilla oficial)
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
            
            buf_aud = io.BytesIO()
            df_export_aud = df_incompletos[["SKU"] + columnas_auditadas].copy()
            # Limpiar valores vacíos para que el usuario vea celdas vacías claras
            for col in columnas_auditadas:
                df_export_aud[col] = df_export_aud[col].apply(lambda x: "" if campo_vacio(x) else x)
            
            with pd.ExcelWriter(buf_aud, engine='openpyxl') as wr_aud:
                df_export_aud.to_excel(wr_aud, index=False, sheet_name='Articulos_Incompletos')
                ws_aud = wr_aud.sheets['Articulos_Incompletos']
                
                # Formato de encabezado corporativo
                fill_header = PatternFill(start_color="D32F2F", end_color="D32F2F", fill_type="solid")
                font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
                fill_vacio = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")  # Amarillo claro para celdas vacías
                font_normal = Font(name="Calibri", size=11)
                thin_border = Border(
                    left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin')
                )
                
                for cell in ws_aud[1]:
                    cell.fill = fill_header
                    cell.font = font_header
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = thin_border
                
                # Formatear celdas de datos: resaltar en amarillo las vacías
                for row in ws_aud.iter_rows(min_row=2, max_row=ws_aud.max_row, max_col=ws_aud.max_column):
                    for cell in row:
                        cell.font = font_normal
                        cell.border = thin_border
                        if cell.column > 1 and (cell.value is None or str(cell.value).strip() == ""):
                            cell.fill = fill_vacio
                
                # Ajustar anchos
                ws_aud.column_dimensions['A'].width = 20
                ws_aud.column_dimensions['B'].width = 35
                ws_aud.column_dimensions['C'].width = 18
                ws_aud.column_dimensions['D'].width = 28
                ws_aud.column_dimensions['E'].width = 22
                
                # Agregar validaciones de datos (Calibre y Acabado)
                opciones_calibres = '"10GA,12GA,14GA,16GA,10GACR,12GACR,14GACR,16GACR,125AL,250AL,188AL"'
                dv_cal = DataValidation(type="list", formula1=opciones_calibres, allow_blank=True)
                ws_aud.add_data_validation(dv_cal)
                dv_cal.add(f"C2:C{ws_aud.max_row + 1}")
                
                opciones_acabados = '"Decapado,Ansi 61,Galvanizado,Otro"'
                dv_acab = DataValidation(type="list", formula1=opciones_acabados, allow_blank=True)
                ws_aud.add_data_validation(dv_acab)
                dv_acab.add(f"E2:E{ws_aud.max_row + 1}")
                
            buf_aud.seek(0)
            
            st.download_button(
                label="📥 Descargar Artículos Incompletos (Excel para completar)",
                data=buf_aud.getvalue(),
                file_name="Articulos_Informacion_Faltante.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="btn_download_auditoria_faltantes",
                use_container_width=True
            )
            st.info("💡 **Instrucciones:** Descargue el archivo, complete las celdas amarillas vacías y vuelva a subirlo en **Mantenimiento y Catálogos → Catálogo Maestro Detallado → Carga Masiva**. El sistema actualizará automáticamente los registros existentes con la nueva información.")
        else:
            st.success("🎉 ¡Excelente! Todos los artículos del catálogo tienen su información al 100%. No hay campos vacíos.")

        # --- SECCIÓN ADICIONAL: CARGA Y DETALLE DE IMAGEN DE ARTÍCULO ---
        st.write("---")
        st.subheader("🖼️ Detalle e Imagen del Artículo")
        
        filtro_estado_img = st.radio(
            "Filtrar artículos por estado de imagen:",
            ["Todos", "Sin imagen", "Con imagen"],
            horizontal=True,
            key="filtro_estado_imagen_select"
        )
        
        skus_con_img = obtener_skus_con_imagen()
        lista_skus_filtrados = df_art_filtrado['SKU'].dropna().tolist()
        
        if filtro_estado_img == "Sin imagen":
            lista_skus_filtrados = [s for s in lista_skus_filtrados if s not in skus_con_img]
        elif filtro_estado_img == "Con imagen":
            lista_skus_filtrados = [s for s in lista_skus_filtrados if s in skus_con_img]
            
        opc_skus_img = ["Seleccione un SKU..."] + lista_skus_filtrados
        sku_sel = st.selectbox("Seleccione un SKU para administrar su imagen:", opc_skus_img, key="sku_select_img")
        
        if sku_sel != "Seleccione un SKU...":
            art_row = df_art_filtrado[df_art_filtrado['SKU'] == sku_sel].iloc[0]
            
            import glob
            os.makedirs("imagenes_articulos", exist_ok=True)
            
            # Escanear localmente
            matching_files_local = glob.glob(f"imagenes_articulos/{sku_sel}(*.*")
            
            imagen_final_path = None
            if matching_files_local:
                imagen_final_path = matching_files_local[0]
            else:
                # Buscar en GitHub si el token está disponible
                if "github_token" in st.secrets and st.secrets["github_token"]:
                    try:
                        GITHUB_TOKEN = st.secrets["github_token"]
                        url_list = f"https://api.github.com/repos/{REPO_OWNER}/{REPO_NAME}/contents/imagenes_articulos?ref={BRANCH}"
                        headers = {"Authorization": f"token {GITHUB_TOKEN}", "Accept": "application/vnd.github.v3+json"}
                        res_list = requests.get(url_list, headers=headers)
                        if res_list.status_code == 200:
                            items = res_list.json()
                            for item in items:
                                if item["name"].startswith(f"{sku_sel}("):
                                    github_file_path = f"imagenes_articulos/{item['name']}"
                                    if descargar_imagen_desde_github(github_file_path):
                                        imagen_final_path = github_file_path
                                        break
                    except Exception:
                        pass
            
            col_ficha, col_cargar = st.columns(2)
            
            with col_ficha:
                st.write("##### Ficha Técnica del Artículo")
                
                # Widget HTML/JS para copiar SKU automáticamente al seleccionar, con botón físico de respaldo
                copiar_html = f"""
                    <div style="margin-bottom: 12px; font-family: sans-serif; display: flex; align-items: center; gap: 8px;">
                        <button id="btn-copiar-sku" style="
                            background-color: #EC2024;
                            color: white;
                            border: none;
                            padding: 8px 16px;
                            border-radius: 4px;
                            cursor: pointer;
                            font-weight: bold;
                            font-size: 13px;
                            display: inline-flex;
                            align-items: center;
                            gap: 8px;
                            box-shadow: 0 2px 4px rgba(0,0,0,0.1);
                        ">
                            📋 Copiar SKU
                        </button>
                        <span id="msg-copiado" style="color: #2E7D32; font-weight: bold; display: none; font-size: 13px; animation: fadeIn 0.3s;">¡Copiado al portapapeles!</span>
                    </div>
                    <script>
                        function copiarTexto() {{
                            navigator.clipboard.writeText("{sku_sel}").then(function() {{
                                var msg = document.getElementById("msg-copiado");
                                msg.style.display = "inline";
                                setTimeout(function() {{
                                    msg.style.display = "none";
                                }}, 2000);
                            }}).catch(function(err) {{
                                console.error("Error al copiar: ", err);
                            }});
                        }}
                        // Ejecución inmediata
                        try {{
                            copiarTexto();
                        }} catch(e) {{}}
                        // Listener del botón
                        document.getElementById("btn-copiar-sku").addEventListener("click", copiarTexto);
                    </script>
                """
                components.html(copiar_html, height=45)
                
                st.markdown(f"**SKU / Código:** `{sku_sel}`")
                st.markdown(f"**Nombre:** {art_row['Nombre']}")
                st.markdown(f"**Calibre / Espesor:** {art_row['Calibre_Espesor']}")
                st.markdown(f"**Dimensiones:** {art_row['Dimensiones_Pieza']}")
                st.markdown(f"**Acabado Superficial:** {art_row['Acabado_Superficial']}")
                
                st.write("")
                if imagen_final_path and os.path.exists(imagen_final_path):
                    st.image(imagen_final_path, caption=f"Imagen cargada para {sku_sel}", use_container_width=True)
                    
                    if st.button("🗑️ Eliminar Imagen de Artículo", use_container_width=True, key=f"btn_del_img_{sku_sel}"):
                        if eliminar_imagen_de_github(imagen_final_path):
                            obtener_skus_con_imagen.clear()
                            st.success("¡Imagen eliminada correctamente del servidor y GitHub!")
                            st.rerun()
                        else:
                            st.error("Error al eliminar la imagen en GitHub.")
                else:
                    st.info("Este artículo no cuenta con una imagen asociada actualmente.")
            
            with col_cargar:
                st.write("##### Cargar / Reemplazar Imagen")
                
                file_uploaded = st.file_uploader("Subir archivo de imagen (PNG, JPG, JPEG):", type=["png", "jpg", "jpeg"], key=f"file_uploader_{sku_sel}")
                
                from streamlit_paste_button import paste_image_button as pbutton
                paste_result = pbutton(
                    "📋 Pegar captura de pantalla desde el portapapeles",
                    key=f"paste_button_{sku_sel}"
                )
                
                nueva_imagen_data = None
                img_ext = ".png"
                
                if file_uploaded:
                    nueva_imagen_data = Image.open(file_uploaded)
                    _, ext = os.path.splitext(file_uploaded.name)
                    if ext.lower() in [".png", ".jpg", ".jpeg"]:
                        img_ext = ext.lower()
                elif paste_result.image_data is not None:
                    nueva_imagen_data = paste_result.image_data
                    img_ext = ".png"
                    
                if nueva_imagen_data is not None:
                    st.write("---")
                    st.warning("⚠️ Vista Previa de la Nueva Imagen (Aún no se ha guardado):")
                    st.image(nueva_imagen_data, caption="Vista Previa de la Carga", use_container_width=True)
                    
                    c_save, c_cancel = st.columns(2)
                    
                    with c_save:
                        if st.button("💾 Guardar y Sincronizar Imagen", use_container_width=True, key=f"btn_save_img_{sku_sel}"):
                            # 1. Crear nombre final en inglés: SKU(dd-Mon-yy)
                            meses_en = {
                                1: "Jan", 2: "Feb", 3: "Mar", 4: "Apr", 5: "May", 6: "Jun",
                                7: "Jul", 8: "Aug", 9: "Sep", 10: "Oct", 11: "Nov", 12: "Dec"
                            }
                            hoy = datetime.date.today()
                            fecha_ingles = f"{hoy.day:02d}-{meses_en[hoy.month]}-{str(hoy.year)[-2:]}"
                            nombre_archivo = f"{sku_sel}({fecha_ingles}){img_ext}"
                            nuevo_path = f"imagenes_articulos/{nombre_archivo}"
                            
                            if imagen_final_path and os.path.exists(imagen_final_path):
                                eliminar_imagen_de_github(imagen_final_path)
                            
                            try:
                                if img_ext in [".jpg", ".jpeg"] and nueva_imagen_data.mode in ("RGBA", "P"):
                                    nueva_imagen_data = nueva_imagen_data.convert("RGB")
                                nueva_imagen_data.save(nuevo_path)
                                
                                if subir_imagen_a_github(nuevo_path):
                                    obtener_skus_con_imagen.clear()
                                    st.success("¡Imagen guardada y sincronizada correctamente en GitHub!")
                                    st.rerun()
                                else:
                                    st.error("Error al sincronizar la imagen con el repositorio de GitHub.")
                            except Exception as ex_save:
                                st.error(f"Error al guardar el archivo localmente: {ex_save}")
                                
                    with c_cancel:
                        if st.button("❌ Descartar", use_container_width=True, key=f"btn_discard_img_{sku_sel}"):
                            st.rerun()

    else:
        st.info("ℹ️ No hay artículos registrados en el catálogo maestro actualmente o el archivo en GitHub está vacío.")

elif opcion_menu == "🏭 Industria 4.0":
    st.title("🏭 Manufactura Inteligente e Industria 4.0")
    st.markdown("##### Integración Tecnológica y Automatización de Procesos Logísticos")
    
    st.markdown("""
    <div style="background-color: #F8FAFC; border-left: 5px solid #EC2024; padding: 15px; border-radius: 4px; margin-bottom: 20px;">
        <p style="font-family: 'Questrial', sans-serif; font-size: 15px; color: #1E293B; margin: 0; line-height: 1.6;">
            La digitalización y la interconectividad son los pilares fundamentales del ecosistema industrial contemporáneo. 
            Esta sección describe cómo la presente plataforma de Remisiones de Materiales se alinea estratégicamente 
            con las tendencias de la <b>Manufactura Inteligente</b> y la <b>Industria 4.0</b>.
        </p>
    </div>
    """, unsafe_allow_html=True)
    
    tab_ind1, tab_ind2, tab_ind3 = st.tabs([
        "🔬 Justificación de Manufactura Inteligente", 
        "🎯 Beneficios Estratégicos", 
        "💻 Stack Tecnológico"
    ])
    
    with tab_ind1:
        st.write("")
        st.markdown("""
        <h4 style="color: #EC2024; font-family: 'Montserrat', sans-serif; font-weight: bold; margin-bottom: 10px;">
            El Rol del Dato en la Cadena Logística
        </h4>
        <p style="font-family: 'Questrial', sans-serif; font-size: 14px; line-height: 1.6; color: #111111;">
            En la manufactura de metales y ensambles de precisión, la logística interna suele ser un cuello de botella. 
            La transformación digital implica convertir flujos físicos en información estructurada utilizable para la toma de decisiones.
        </p>
        <p style="font-family: 'Questrial', sans-serif; font-size: 14px; line-height: 1.6; color: #111111;">
            <b>Habilitación del Gemelo Digital (Digital Twin):</b> Cada tarima empacada en la planta no es solo un bulto físico de metal; 
            a través de este sistema, se asocia dinámicamente con un identificador único <code>TPM-XXXX</code> que representa digitalmente su 
            contenido granular (SKUs, POs, proyectos y parcialidades). Esto permite rastrear virtualmente la ubicación, 
            estatus e historial de cada paquete sin necesidad de inspecciones físicas manuales, logrando una representación en tiempo real del inventario.
        </p>
        <p style="font-family: 'Questrial', sans-serif; font-size: 14px; line-height: 1.6; color: #111111;">
            <b>Filosofía de IoT Industrial (IIoT):</b> Aunque no utilicemos sensores físicos directos en esta etapa, el esquema de 
            registro descentralizado por medio de terminales móviles y la sincronización con una base de datos centralizada en la nube 
            sigue la arquitectura básica del IIoT: capturar el dato directamente en la fuente (línea de producción / empaque) 
            y comunicarlo de manera ágil hacia las capas de toma de decisiones.
        </p>
        """, unsafe_allow_html=True)
        
    with tab_ind2:
        st.write("")
        st.markdown('<h4 style="color: #EC2024; font-family: \'Montserrat\', sans-serif; font-weight: bold; margin-bottom: 15px;">Impacto y Retorno de Inversión Tecnológica</h4>', unsafe_allow_html=True)
        
        col_b1, col_b2 = st.columns(2)
        with col_b1:
            st.markdown("""
            <div class="report-card">
                <h5 style="color: #111111; font-family: 'Montserrat', sans-serif; font-weight: bold; margin-top: 0;"><span style="color: #EC2024;">📍</span> Trazabilidad de Extremo a Extremo</h5>
                <p style="font-family: 'Questrial', sans-serif; font-size: 13.5px; line-height: 1.5; color: #334155; margin-bottom: 0;">
                    Registro inalterable de cada movimiento: desde la creación de la tarima por el líder del área, pasando por el empaque de productos específicos, hasta su despacho final documentado en una remisión firmada.
                </p>
            </div>
            <div class="report-card">
                <h5 style="color: #111111; font-family: 'Montserrat', sans-serif; font-weight: bold; margin-top: 0;"><span style="color: #EC2024;">⚡</span> Eliminación del Error Humano</h5>
                <p style="font-family: 'Questrial', sans-serif; font-size: 13.5px; line-height: 1.5; color: #334155; margin-bottom: 0;">
                    El sistema valida automáticamente las cantidades y referencias contra el catálogo maestro de artículos y líderes, impidiendo errores de dedo, duplicados o envío de folios erróneos a los clientes.
                </p>
            </div>
            """, unsafe_allow_html=True)
            
        with col_b2:
            st.markdown("""
            <div class="report-card">
                <h5 style="color: #111111; font-family: 'Montserrat', sans-serif; font-weight: bold; margin-top: 0;"><span style="color: #EC2024;">⏱️</span> Agilidad en Operaciones</h5>
                <p style="font-family: 'Questrial', sans-serif; font-size: 13.5px; line-height: 1.5; color: #334155; margin-bottom: 0;">
                    Generación en 3 segundos de manifiestos y anexos en formato PDF certificado y listos para impresión. Lo que antes requería transcripciones manuales de varias horas, ahora se realiza de forma automatizada e inmediata.
                </p>
            </div>
            <div class="report-card">
                <h5 style="color: #111111; font-family: 'Montserrat', sans-serif; font-weight: bold; margin-top: 0;"><span style="color: #EC2024;">☁️</span> Persistencia Híbrida y Resiliencia</h5>
                <p style="font-family: 'Questrial', sans-serif; font-size: 13.5px; line-height: 1.5; color: #334155; margin-bottom: 0;">
                    El sistema combina almacenamiento local en caché con sincronización instantánea en la nube mediante APIs (GitHub API). Esto garantiza que la aplicación siga operando en planta incluso ante interrupciones de red.
                </p>
            </div>
            """, unsafe_allow_html=True)

    with tab_ind3:
        st.write("")
        st.markdown('<h4 style="color: #EC2024; font-family: \'Montserrat\', sans-serif; font-weight: bold; margin-bottom: 15px;">Arquitectura de Software y Datos</h4>', unsafe_allow_html=True)
        
        st.markdown("""
        <p style="font-family: 'Questrial', sans-serif; font-size: 14px; line-height: 1.6; color: #111111;">
            La aplicación ha sido diseñada bajo estándares de desarrollo ágil y tecnologías robustas de procesamiento de datos:
        </p>
        """, unsafe_allow_html=True)
        
        col_st1, col_st2, col_st3, col_st4 = st.columns(4)
        with col_st1:
            st.metric("🎨 Front-End UI", "Streamlit")
            st.caption("Entorno web reactivo de alta velocidad, estilizado con CSS de marca.")
        with col_st2:
            st.metric("⚙️ Data Engine", "Python / Pandas")
            st.caption("Procesamiento y modelado de datos tabulares en memoria.")
        with col_st3:
            st.metric("📁 Cloud Storage", "GitHub API")
            st.caption("Sincronización en la nube mediante API controlada por tokens.")
        with col_st4:
            st.metric("📄 Document Maker", "ReportLab")
            st.caption("Motor de precisión para la construcción de reportes PDF oficiales.")
            
        st.markdown("<br/>", unsafe_allow_html=True)
        col_st5, col_st6 = st.columns(2)
        with col_st5:
            st.metric("📦 Zip Archiver", "Python ZipFile")
            st.caption("Compresión en memoria de múltiples reportes de etiqueta para optimizar la descarga.")
        with col_st6:
            st.metric("📧 Email Standard", "RFC 822 / MIME")
            st.caption("Ensamblado dinámico de correos con formato de borrador de Outlook, tablas HTML e imágenes.")

elif opcion_menu == "📖 Manual de Operación":
    st.title("📖 Manual de Operación del Sistema")
    st.markdown("##### Guía de Usuario para el Control de Remisiones y Tarimas")
    
    st.markdown("""
    <div style="background-color: #F8FAFC; border-left: 5px solid #EC2024; padding: 15px; border-radius: 4px; margin-bottom: 20px;">
        <p style="font-family: 'Questrial', sans-serif; font-size: 15px; color: #1E293B; margin: 0; line-height: 1.6;">
            Bienvenido al manual oficial de operación de la aplicación. 
            Utilice las pestañas inferiores para comprender el funcionamiento detallado de cada módulo del sistema.
        </p>
    </div>
    """, unsafe_allow_html=True)
    
    tab_man1, tab_man2, tab_man3, tab_man4, tab_man5 = st.tabs([
        "🔑 Acceso y Seguridad", 
        "📦 Módulo Tarimas", 
        "🚚 Módulo Remisiones", 
        "🔍 Consultas e Impresión", 
        "⚙️ Mantenimiento (Admin)"
    ])
    
    with tab_man1:
        st.write("")
        st.markdown("""
        <h4 style="color: #EC2024; font-family: 'Montserrat', sans-serif; font-weight: bold;">Control de Credenciales y Perfiles</h4>
        <p style="font-family: 'Questrial', sans-serif; font-size: 14px; line-height: 1.6;">
            La aplicación restringe sus funciones mediante credenciales de acceso para garantizar la integridad de los datos de inventario:
        </p>
        <ul style="font-family: 'Questrial', sans-serif; font-size: 14px; line-height: 1.6;">
            <li><b>Perfil Operador / Líder:</b> Permite operar la mayor parte del sistema (visualizar tableros, cargar proyectos, registrar tarimas, generar remisiones y descargar reportes). 
            Para ingresar, use su nombre tal cual aparece en el catálogo de líderes (ej. <code>Jesús Morales</code>, <code>Miguel Alvarado</code>) y la contraseña general <code>Metales</code>.</li>
            <li><b>Perfil Administrador:</b> Cuenta con privilegios especiales para realizar modificaciones en caliente de bases de datos, corregir inventarios, modificar catálogos y purgar registros. 
            Para ingresar, use el usuario <code>admin</code> y la contraseña corporativa <code>SigramaMetales2026</code>.</li>
            <li><b>Cierre de Sesión:</b> Al concluir sus labores, es recomendable presionar el botón <b>🚪 Cerrar Sesión</b> en la barra lateral para evitar que otros usuarios realicen movimientos bajo su firma.</li>
        </ul>
        """, unsafe_allow_html=True)
        
    with tab_man2:
        st.write("")
        st.markdown("""
        <h4 style="color: #EC2024; font-family: 'Montserrat', sans-serif; font-weight: bold;">Empaque e Inventariado de Tarimas</h4>
        <p style="font-family: 'Questrial', sans-serif; font-size: 14px; line-height: 1.6;">
            El Módulo Tarimas sirve para registrar y desglosar el contenido de los paquetes que se preparan en planta:
        </p>
        <ol style="font-family: 'Questrial', sans-serif; font-size: 14px; line-height: 1.6;">
            <li><b>Carga de Plantilla Excel:</b> Suba el archivo en formato Excel con las columnas requeridas (Tarima, Producto/SKU, PO, Proyecto, Parcialidad, Descripcion, Cantidad).</li>
            <li><b>Asignación de Líder y Tipo:</b> Indique el líder de línea a cargo del empaque y especifique si la tarima física es Cuadrada o Rectangular.</li>
            <li><b>Procesamiento:</b> Presione <i>Procesar e Integrar Plantilla Avanzada</i>. El sistema validará los SKUs contra el catálogo oficial, generará automáticamente folios correlativos únicos de tarima (TPM-XXXX), y subirá la información estructurada a las bases de datos de GitHub de manera segura.</li>
        </ol>
        """, unsafe_allow_html=True)
        
    with tab_man3:
        st.write("")
        st.markdown("""
        <h4 style="color: #EC2024; font-family: 'Montserrat', sans-serif; font-weight: bold;">Salida de Materiales y Remisión</h4>
        <p style="font-family: 'Questrial', sans-serif; font-size: 14px; line-height: 1.6;">
            El Módulo Remisiones agrupa las tarimas que saldrán de la planta con destino a un cliente:
        </p>
        <ol style="font-family: 'Questrial', sans-serif; font-size: 14px; line-height: 1.6;">
            <li><b>Selección de Tarimas:</b> El dropdown muestra todas las tarimas con estatus <code>Disponible</code> (es decir, creadas pero que aún no han sido asignadas a ninguna remisión). Seleccione una o más tarimas para agruparlas.</li>
            <li><b>Datos de Despacho:</b> Especifique el líder que autoriza la salida, el almacén de origen, el nombre del receptor (cliente) y la dirección destino.</li>
            <li><b>Emisión:</b> Haga clic en <i>🚀 Confirmar Salida y Generar Nueva Remisión</i>. La aplicación generará un folio oficial correlativo (ej. <code>E0028</code>), marcará el estatus de las tarimas como <code>Remesada</code> para que no puedan volver a usarse y guardará la transacción en la nube.</li>
            <li><b>Notificación de Salida:</b> Tras emitir la remisión, use el panel inferior para descargar el borrador de correo (.eml) que contiene la información formateada y lista para ser despachada a través de Outlook a los destinatarios configurados.</li>
        </ol>
        """, unsafe_allow_html=True)
        
    with tab_man4:
        st.write("")
        st.markdown("""
        <h4 style="color: #EC2024; font-family: 'Montserrat', sans-serif; font-weight: bold;">Centro de Consultas y Descarga Documental</h4>
        <p style="font-family: 'Questrial', sans-serif; font-size: 14px; line-height: 1.6;">
            La consulta y obtención de documentos certificados se centraliza en tres áreas clave:
        </p>
        <ul style="font-family: 'Questrial', sans-serif; font-size: 14px; line-height: 1.6;">
            <li><b>Dashboard e Históricos:</b> Muestra métricas agregadas del inventario (tarimas disponibles, remesadas, etc.) y gráficos interactivos de barras y líneas para analizar la cantidad de productos por proyecto en tiempo real.</li>
            <li><b>Centro de Consultas (Impresión Masiva):</b> Permite ver la lista completa de tarimas, filtrarlas por folio, estatus o fecha, y seleccionar filas específicas para generar reportes anejos en PDF con el desglose exacto de su contenido.</li>
            <li><b>Descarga Documental de Remisiones y EML:</b> En la parte inferior del Módulo Remisiones, use la tabla interactiva de remisiones históricas. Puede seleccionar **una o múltiples remisiones** a la vez:
            - Si selecciona una sola remisión, podrá descargar su PDF y Excel oficial de manera independiente.
            - Si selecciona múltiples remisiones, podrá generar un correo consolidado (.eml) que adjuntará cada remisión por separado en PDF y agrupará todas las etiquetas de las tarimas involucradas en un archivo <code>Etiquetas_Tarimas.zip</code>. El cuerpo del correo se generará de acuerdo con los lineamientos oficiales de Industria Sigrama.</li>
        </ul>
        """, unsafe_allow_html=True)
        
    with tab_man5:
        st.write("")
        st.markdown("""
        <h4 style="color: #EC2024; font-family: 'Montserrat', sans-serif; font-weight: bold;">Mantenimiento Crítico (Acceso Restringido)</h4>
        <p style="font-family: 'Questrial', sans-serif; font-size: 14px; line-height: 1.6;">
            El módulo <b>⚙️ Mantenimiento y Catálogos</b> está reservado para el perfil <b>Administrador</b> (o personal de soporte de TI):
        </p>
        <ul style="font-family: 'Questrial', sans-serif; font-size: 14px; line-height: 1.6;">
            <li><b>Ajustar Cantidades:</b> Editor interactivo tipo Excel que permite modificar directamente las cantidades o descripciones de una tarima ya guardada para corregir capturas erróneas.</li>
            <li><b>Catálogo de Líderes y Artículos:</b> Permite dar de alta nuevos códigos SKU o empleados autorizados, ya sea uno a uno o cargando un archivo Excel masivo.</li>
            <li><b>Contador de Tarimas (Consecutivos):</b> Si necesita reiniciar la numeración o saltar folios (ej. continuar en 150), digite el consecutivo deseado y presione guardar.</li>
            <li><b>Purga de Datos (Eliminación):</b> Permite eliminar remisiones o tarimas seleccionadas para liberar espacio, o aplicar un <b>Reset de Fábrica Total</b>. 
            <i>Nota: Debido al riesgo operativo, esta pestaña se bloquea por completo a nivel de base de datos si el rol del usuario no es Administrador, incluso si se ingresa con la clave de soporte técnico.</i></li>
        </ul>
        """, unsafe_allow_html=True)

elif opcion_menu == "🏢 Reporte por Receptor":
    st.title("🏢 Reporte de Piezas Remisionadas por Receptor")
    st.markdown("Consulte el detalle de piezas y tarimas enviadas a cada cliente o receptor de remisiones.")
    
    if not st.session_state.BD_Datos_Generales_Remision.empty and not st.session_state.BD_Detalle_Tarimas.empty:
        df_rems = st.session_state.BD_Datos_Generales_Remision.copy()
        receptores_disp = ["Seleccione un Receptor...", "Todos"] + sorted(df_rems['Nombre_Receptor'].dropna().unique().tolist())
        
        col_rec1, col_rec2 = st.columns([3, 1])
        with col_rec1:
            receptor_sel = st.selectbox("Seleccione el Receptor (Cliente):", receptores_disp, key="sel_receptor_reporte")
            
        if receptor_sel != "Seleccione un Receptor...":
            if receptor_sel == "Todos":
                df_rems_filtradas = df_rems.copy()
            else:
                df_rems_filtradas = df_rems[df_rems['Nombre_Receptor'] == receptor_sel]
            
            tarimas_asociadas = []
            import ast
            for _, row in df_rems_filtradas.iterrows():
                t_val = row['Tarimas_Asociadas']
                t_list = []
                if isinstance(t_val, str):
                    try:
                        t_list = ast.literal_eval(t_val)
                    except Exception:
                        t_list = [t_val]
                elif isinstance(t_val, list):
                    t_list = t_val
                
                for t in t_list:
                    tarimas_asociadas.append({
                        "ID_Tarima": str(t).strip(),
                        "Folio_Remision": row['Folio_Remision'],
                        "Fecha_Salida": row['Fecha_Hora_Salida'],
                        "Receptor": row['Nombre_Receptor']
                    })
                    
            if tarimas_asociadas:
                df_tarimas_receptor = pd.DataFrame(tarimas_asociadas)
                df_detalle = st.session_state.BD_Detalle_Tarimas.copy()
                df_detalle['ID_Tarima'] = df_detalle['ID_Tarima'].astype(str).str.strip()
                
                df_cruce = pd.merge(df_tarimas_receptor, df_detalle, on="ID_Tarima", how="inner")
                
                if not df_cruce.empty:
                    if "BD_Articulos" in st.session_state and not st.session_state.BD_Articulos.empty:
                        df_cruce = pd.merge(df_cruce, st.session_state.BD_Articulos[['SKU', 'Nombre']], on="SKU", how="left")
                    else:
                        df_cruce['Nombre'] = "N/A"
                        
                    df_cruce = df_cruce.rename(columns={
                        "ID_Tarima": "Tarima",
                        "Folio_Remision": "Remisión",
                        "Fecha_Salida": "Fecha Envío",
                        "Receptor": "Cliente / Receptor",
                        "SKU": "Código (SKU)",
                        "Nombre": "Descripción Comercial",
                        "Cantidad": "Piezas"
                    })
                    
                    columnas_mostrar = ["Remisión", "Fecha Envío", "Cliente / Receptor", "Tarima", "Código (SKU)", "Descripción Comercial", "Piezas", "PO", "Proyecto"]
                    columnas_mostrar = [c for c in columnas_mostrar if c in df_cruce.columns]
                    
                    df_mostrar = df_cruce[columnas_mostrar].copy()
                    
                    for col in df_mostrar.columns:
                        if col != "Piezas":
                            df_mostrar[col] = df_mostrar[col].astype(str)
                    df_mostrar["Piezas"] = pd.to_numeric(df_mostrar["Piezas"], errors='coerce').fillna(0).astype(int)
                    
                    total_pzs = df_mostrar["Piezas"].sum()
                    
                    st.write("---")
                    col_met1, col_met2 = st.columns(2)
                    col_met1.metric(f"📦 Total Piezas Remisionadas a {receptor_sel}", f"{total_pzs:,} Pzs")
                    col_met2.metric("🚚 Total Remisiones Relacionadas", df_mostrar["Remisión"].nunique())
                    
                    st.write("#### Detalle Operativo")
                    st.dataframe(df_mostrar, use_container_width=True, hide_index=True)
                    
                    with col_rec2:
                        st.write("") # Alineación vertical
                        csv = df_mostrar.to_csv(index=False).encode('utf-8')
                        st.download_button(
                            label="📥 Descargar CSV",
                            data=csv,
                            file_name=f"Reporte_Receptor_{receptor_sel}.csv",
                            mime="text/csv",
                            use_container_width=True
                        )
                else:
                    st.warning("No hay registros de piezas para las tarimas enviadas a este receptor.")
            else:
                st.warning("No hay tarimas asociadas a las remisiones de este receptor.")
    else:
        st.info("No hay datos suficientes (Falta Base de Remisiones o Detalle de Tarimas).")

elif opcion_menu == "⚙️ Mantenimiento y Catálogos":
    st.title("⚙️ Panel de Mantenimiento Avanzado del Sistema")
    st.warning("⚠️ Acción Crítica: Las modificaciones realizadas impactan directamente en los archivos de GitHub.")
    
    # Inicialización del catálogo básico de personal operativo en memoria
    if "BD_Lideres" not in st.session_state:
        st.session_state.BD_Lideres = pd.DataFrame([
            {"ID_Lider": "LID-01", "Nombre_Lider": "Jesus Morales", "Area": "Metales", "Estatus": "Activo"}
        ])

    tab1, tab2, tab3, tab4, tab5, tab6, tab7, tab8 = st.tabs(["📝 Ajustar Cantidades", "👥 Catálogo de Líderes", "⚠️ Purga de Datos", "📦 Catálogo de Artículos", "🔢 Contador de Tarimas", "🖼️ Carpeta de Imágenes", "🏢 Catálogo de Receptores", "📧 Listas de Correo"])



    # --- SUB-MÓDULO 1: MODIFICACIÓN DIRECTA DE DETALLES DE INVENTARIO ---
    with tab1:
        st.subheader("✏️ Edición Rápida de Inventario (Detalle Tarimas)")
        if not st.session_state.BD_Detalle_Tarimas.empty:
            st.markdown("Haga doble clic sobre cualquier celda permitida (**SKU, PO, Parcialidad, Descripcion o Cantidad**) para corregir errores:")
            
            # Editor interactivo bloqueando únicamente llaves primarias y relacionales del sistema
            df_editable = st.data_editor(
                st.session_state.BD_Detalle_Tarimas, 
                use_container_width=True, 
                disabled=["ID_Detalle", "ID_Tarima", "Proyecto"], 
                hide_index=True, 
                key="editor_mantenimiento_cantidades_final"
            )
            
            if st.button("💾 Guardar Cambios de Inventario en GitHub"):
                st.session_state.BD_Detalle_Tarimas = df_editable
                subir_excel_a_github("BD_Detalle_Tarimas.xlsx", st.session_state.BD_Detalle_Tarimas)
                st.success("✅ ¡Inventario corregido y sincronizado con éxito!"); st.rerun()
        else: 
            st.info("No existen registros en el desglose de inventario para modificar.")
# =============================================================================
# SECCIÓN 17B: MANTENIMIENTO - CATALOGO DE LÍDERES MEDIANTE ARCHIVOS MASIVOS
# =============================================================================
    with tab2:
        st.subheader("👤 Administración del Personal de Líderes")
        
        # Generación de archivo muestra descargable para los operadores de IT
        df_l_plantilla = pd.DataFrame([
            {"Nombre_Lider": "Nombre Ejemplo 1", "Area": "Metales"},
            {"Nombre_Lider": "Nombre Ejemplo 2", "Area": "Embarques"}
        ])
        
        from openpyxl.styles import Font, PatternFill
        buf_l = io.BytesIO()
        with pd.ExcelWriter(buf_l, engine='openpyxl') as wr_l:
            df_l_plantilla.to_excel(wr_l, index=False, sheet_name='Plantilla_Lideres')
            ws_l = wr_l.sheets['Plantilla_Lideres']
            
            fill_header_lider = PatternFill(start_color="757575", end_color="757575", fill_type="solid")
            font_header_lider = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
            for col_idx in range(1, 3):
                cell = ws_l.cell(row=1, column=col_idx)
                cell.font, cell.fill = font_header_lider, fill_header_lider
                
        st.write("Descargue el formato base para rellenar la lista de personal autorizado:")
        st.download_button(label="📥 Descargar Plantilla de Personal (.xlsx)", data=buf_l.getvalue(), file_name="plantilla_lideres_sigrama.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key="btn_download_plantilla_lideres_u")
        st.write("---")
        
        # Procesador de integración masiva de empleados
        arch_lideres = st.file_uploader("Suba la Plantilla de Personal Rellenada:", type=["xlsx"], key="uploader_lideres_masivo_u")
        if arch_lideres and st.button("🚀 Procesar e Integrar Personal Masivo"):
            try:
                df_l_excel = pd.read_excel(arch_lideres)
                if not all(col in df_l_excel.columns for col in ["Nombre_Lider", "Area"]):
                    st.error("❌ Error: Columnas incompatibles. Use: Nombre_Lider, Area")
                else:
                    for _, row_l in df_l_excel.iterrows():
                        if pd.notna(row_l['Nombre_Lider']) and str(row_l['Nombre_Lider']).strip() != "":
                            n_id_l = f"LID-{(len(st.session_state.BD_Lideres) + 1):02d}"
                            n_l_row = {"ID_Lider": n_id_l, "Nombre_Lider": str(row_l['Nombre_Lider']).strip(), "Area": str(row_l['Area']).strip() if pd.notna(row_l['Area']) else "Metales", "Estatus": "Activo"}
                            st.session_state.BD_Lideres = pd.concat([st.session_state.BD_Lideres, pd.DataFrame([n_l_row])], ignore_index=True)
                    subir_excel_a_github("BD_Lideres.xlsx", st.session_state.BD_Lideres)
                    st.success("✅ ¡Lista de personal integrada y respaldada con éxito!"); st.rerun()
            except Exception as e: st.error(f"Error: {e}")
# =============================================================================
# SECCIÓN 17C: MANTENIMIENTO - REGISTRO TRADICIONAL Y RESPALDO COMPACTO
# =============================================================================
        st.write("---")
        with st.expander("➕ Dar de Alta Nuevo Líder Individual"):
            c_l1, c_l2 = st.columns(2)
            with c_l1: nuevo_nom = st.text_input("Nombre Completo del Líder:", key="txt_input_nuevo_lider_name_f")
            with c_l2: nueva_area = st.text_input("Área de Adscripción:", "Metales", key="txt_input_nuevo_lider_area_f")
            if st.button("➕ Registrar Líder Individual"):
                if nuevo_nom:
                    n_row = {"ID_Lider": f"LID-{(len(st.session_state.BD_Lideres) + 1):02d}", "Nombre_Lider": nuevo_nom, "Area": nueva_area, "Estatus": "Activo"}
                    st.session_state.BD_Lideres = pd.concat([st.session_state.BD_Lideres, pd.DataFrame([n_row])], ignore_index=True)
                    subir_excel_a_github("BD_Lideres.xlsx", st.session_state.BD_Lideres)
                    st.success("Líder registrado exitosamente."); st.rerun()
                    
        st.write("📋 Catálogo Maestro de Líderes Autorizados (Editable):")
        df_l_edit = st.data_editor(st.session_state.BD_Lideres, use_container_width=True, hide_index=True, key="editor_catalogo_lideres_master_final")
        if st.button("💾 Sincronizar Cambios Manuales de Líderes"):
            st.session_state.BD_Lideres = df_l_edit
            subir_excel_a_github("BD_Lideres.xlsx", st.session_state.BD_Lideres)
            st.success("Catálogo de personal operativo actualizado en GitHub.")
# =============================================================================
# SECCIÓN 17D: MANTENIMIENTO - PURGA SELECCIONADA Y CIERRE GENERAL DEL SISTEMA
# =============================================================================
    with tab3:
        st.subheader("🚨 Reset de Fábrica y Purga de Datos Controlada")
        if st.session_state.rol != "Administrador":
            st.error("🔒 Área Bloqueada: Solo el Administrador del sistema tiene permisos para purgar o eliminar registros de las bases de datos.")
        else:
            metodo_purga = st.radio("Método de Purga:", ["❌ Purga Total Automática (Reset Completo)", "🔍 Seleccionar Registros Específicos para Eliminar"], horizontal=True, key="radio_metodo_purga_master_final")
            st.write("---")
            
            if metodo_purga == "❌ Purga Total Automática (Reset Completo)":
                st.error("⚠️ Peligro: Esta acción vaciará por completo todos los históricos en el repositorio.")
                if st.checkbox("Confirmo que deseo aplicar un Reset de Fábrica Total.", key="chk_total_purga_final_final"):
                    if st.button("🗑️ EJECUTAR PURGA MAESTRA TOTAL"):
                        st.session_state.BD_Tarimas = pd.DataFrame(columns=st.session_state.BD_Tarimas.columns)
                        st.session_state.BD_Detalle_Tarimas = pd.DataFrame(columns=st.session_state.BD_Detalle_Tarimas.columns)
                        st.session_state.BD_Datos_Generales_Remision = pd.DataFrame(columns=st.session_state.BD_Datos_Generales_Remision.columns)
                        subir_excel_a_github("BD_Tarimas.xlsx", st.session_state.BD_Tarimas)
                        subir_excel_a_github("BD_Detalle_Tarimas.xlsx", st.session_state.BD_Detalle_Tarimas)
                        subir_excel_a_github("BD_Datos_Generales_Remision.xlsx", st.session_state.BD_Datos_Generales_Remision)
                        import os
                        if os.path.exists("consecutivo_override.txt"):
                            try:
                                os.remove("consecutivo_override.txt")
                            except Exception:
                                pass
                        st.session_state["siguiente_numero_tpm"] = 1
                        st.success("💥 Sistema purgado por completo a ceros de forma masiva."); st.rerun()
            else:
                st.markdown("### 📦 1. Eliminar Tarimas del Inventario")
                if not st.session_state.BD_Tarimas.empty:
                    df_tar_vista = st.session_state.BD_Tarimas.copy().drop(columns=["Es_Nueva"], errors="ignore")
                    sel_tarimas = st.dataframe(df_tar_vista, use_container_width=True, on_select="rerun", selection_mode="multi-row", key="tabla_purga_tarimas_final_f")
                    filas_tar = sel_tarimas.get("selection", {}).get("rows", [])
                    if filas_tar:
                        # Filtrar índices válidos para evitar errores de desfase de Streamlit en la recarga
                        filas_tar_validas = [i for i in filas_tar if i < len(df_tar_vista)]
                        ids_tar_eliminar = df_tar_vista.iloc[filas_tar_validas]['ID_Tarima'].tolist() if filas_tar_validas else []
                        
                        if st.button("🗑️ Eliminar Tarimas Seleccionadas") and ids_tar_eliminar:
                            st.session_state.BD_Tarimas = st.session_state.BD_Tarimas[~st.session_state.BD_Tarimas['ID_Tarima'].isin(ids_tar_eliminar)]
                            st.session_state.BD_Detalle_Tarimas = st.session_state.BD_Detalle_Tarimas[~st.session_state.BD_Detalle_Tarimas['ID_Tarima'].isin(ids_tar_eliminar)]
                            subir_excel_a_github("BD_Tarimas.xlsx", st.session_state.BD_Tarimas)
                            subir_excel_a_github("BD_Detalle_Tarimas.xlsx", st.session_state.BD_Detalle_Tarimas)
                            st.session_state["siguiente_numero_tpm"] = obtener_siguiente_consecutivo_tpm()
                            st.success("✅ Tarimas removidas."); st.rerun()
                else: st.write("No hay tarimas registradas.")
                
                st.write("---")
                st.markdown("### 🚚 2. Eliminar Remisiones de Salida")
                if not st.session_state.BD_Datos_Generales_Remision.empty:
                    df_rem_purga = st.session_state.BD_Datos_Generales_Remision.copy()
                    if 'Tarimas_Asociadas' in df_rem_purga.columns:
                        df_rem_purga['Tarimas_Asociadas'] = df_rem_purga['Tarimas_Asociadas'].astype(str)
                    sel_remisiones = st.dataframe(df_rem_purga, use_container_width=True, on_select="rerun", selection_mode="multi-row", key="tabla_purga_remisiones_final_f")
                    filas_rem = sel_remisiones.get("selection", {}).get("rows", [])
                    if filas_rem:
                        # Filtrar índices válidos para evitar errores de desfase de Streamlit en la recarga
                        filas_rem_validas = [i for i in filas_rem if i < len(st.session_state.BD_Datos_Generales_Remision)]
                        if filas_rem_validas:
                            ids_rem_eliminar = st.session_state.BD_Datos_Generales_Remision.iloc[filas_rem_validas]['Folio_Remision'].tolist()
                            tarimas_afectadas = []
                            import ast
                            for idx in filas_rem_validas:
                                raw_tarimas = st.session_state.BD_Datos_Generales_Remision.iloc[idx]['Tarimas_Asociadas']
                                if isinstance(raw_tarimas, str):
                                    try:
                                        t_list = ast.literal_eval(raw_tarimas)
                                    except Exception:
                                        t_list = [raw_tarimas]
                                elif isinstance(raw_tarimas, list):
                                    t_list = raw_tarimas
                                else:
                                    t_list = []
                                for t in t_list:
                                                 st.write("---")
                                                 
                st.write("---")
                st.markdown("### 📋 3. Eliminar Órdenes de Compra (POs) y Requerimientos")
                if not st.session_state.BD_POs_Cabecera.empty:
                    df_po_purga = st.session_state.BD_POs_Cabecera.copy()
                    sel_pos = st.dataframe(df_po_purga, use_container_width=True, on_select="rerun", selection_mode="multi-row", key="tabla_purga_pos_final_f")
                    filas_po = sel_pos.get("selection", {}).get("rows", [])
                    if filas_po:
                        filas_po_validas = [i for i in filas_po if i < len(df_po_purga)]
                        if filas_po_validas:
                            pos_a_eliminar = df_po_purga.iloc[filas_po_validas]['PO'].astype(str).str.strip().tolist()
                            
                            if st.button("🗑️ Eliminar POs Seleccionadas") and pos_a_eliminar:
                                # Filtrar en cabecera
                                st.session_state.BD_POs_Cabecera = st.session_state.BD_POs_Cabecera[~st.session_state.BD_POs_Cabecera['PO'].astype(str).str.strip().isin(pos_a_eliminar)]
                                # Filtrar en requerimientos
                                if "BD_Requerimientos_POs" in st.session_state and not st.session_state.BD_Requerimientos_POs.empty:
                                    st.session_state.BD_Requerimientos_POs = st.session_state.BD_Requerimientos_POs[~st.session_state.BD_Requerimientos_POs['PO'].astype(str).str.strip().isin(pos_a_eliminar)]
                                
                                # Guardar a GitHub y disco
                                subir_excel_a_github("BD_POs_Cabecera.xlsx", st.session_state.BD_POs_Cabecera)
                                subir_excel_a_github("BD_Requerimientos_POs.xlsx", st.session_state.BD_Requerimientos_POs)
                                
                                st.success("✅ Órdenes de Compra y requerimientos eliminados correctamente."); st.rerun()
                else:
                    st.write("No hay Órdenes de Compra registradas.")
                    
                st.markdown("### 🔄 4. Sincronización y Reparación de Estatus de Tarimas")
                st.info("Utilice esta herramienta para buscar tarimas marcadas como 'Remesadas' pero que no pertenecen a ninguna remisión activa (por ejemplo, remisiones borradas manualmente), y regresarlas al estatus de 'Disponible'.")
                if st.button("🔄 Sincronizar y Reparar Estatus de Tarimas"):
                    corregidas = sincronizar_estatus_tarimas(auto_save=True)
                    if corregidas > 0:
                        st.success(f"✅ ¡Operación completada! Se corrigieron y liberaron {corregidas} tarimas huérfanas.")
                        st.rerun()
                    else:
                        st.info("ℹ️ No se detectaron tarimas con estatus incorrecto. Todo está sincronizado correctamente.")
   
    with tab4:
        st.subheader("📦 Administración y Sincronización del Catálogo de Artículos")
        st.markdown("Utilice este panel para actualizar de forma masiva mediante archivos o editar directamente registros específicos en caliente:")
        
        # Sub-pestañas para separar bases de datos
        sub_tab1, sub_tab2 = st.tabs(["📦 Catálogo Maestro Detallado", "📋 Lista de SKUs Autorizados"])
        
        with sub_tab1:
            st.markdown("#### 📦 Catálogo Maestro de Artículos (Detallado)")
            st.info("💡 **Guía de uso:** Este es el catálogo detallado que contiene Nombre, Calibre, Dimensiones y Acabados de los artículos. Requiere la plantilla estricta de 5 columnas.")
            
            with st.expander("📥 Carga Masiva de Catálogo Maestro mediante Excel", expanded=False):
                c_art1, c_art2 = st.columns(2)
                
                with c_art1:
                    st.write("##### Obtener Plantilla Estructurada")
                    columnas_oficiales = ["SKU", "Nombre", "Calibre_Espesor", "Dimensiones_Pieza", "Acabado_Superficial"]
                    df_plantilla_art = pd.DataFrame(columns=columnas_oficiales)
        
                    buf_p_art = io.BytesIO()
                    with pd.ExcelWriter(buf_p_art, engine='openpyxl') as p_writer:
                        df_plantilla_art.to_excel(p_writer, index=False, sheet_name='Datos_Sistema')
                        worksheet = p_writer.sheets['Datos_Sistema']
                        worksheet.column_dimensions['A'].width = 20
                        worksheet.column_dimensions['B'].width = 25
                        worksheet.column_dimensions['C'].width = 20
                        worksheet.column_dimensions['D'].width = 25
                        worksheet.column_dimensions['E'].width = 20
        
                        opciones_calibres = '"10GA,12GA,14GA,16GA,10GACR,12GACR,14GACR,16GACR","125AL","250AL","188AL"'
                        dv_calibre = DataValidation(type="list", formula1=opciones_calibres, allow_blank=True)
                        worksheet.add_data_validation(dv_calibre)
                        dv_calibre.add("C2:C2000")
        
                        opciones_acabados = '"Decapado,Ansi 61,Galvanizado,Otro"'
                        dv_acabado = DataValidation(type="list", formula1=opciones_acabados, allow_blank=True)
                        worksheet.add_data_validation(dv_acabado)
                        dv_acabado.add("E2:E2000")
        
                    buf_p_art.seek(0)
                    st.download_button(
                        label="📊 Descargar Plantilla Base Artículos (.xlsx)",
                        data=buf_p_art.getvalue(),
                        file_name="Plantilla_Maestra_Articulos.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key="btn_download_plantilla_maestra_articulos_sgc_v2"
                    )
                    
                    st.write("---")
                    st.write("##### Artículos en Tarimas Sin Registro en Catálogo")
                    
                    skus_en_tarimas = set()
                    if "BD_Detalle_Tarimas" in st.session_state and not st.session_state.BD_Detalle_Tarimas.empty:
                        skus_en_tarimas = set(st.session_state.BD_Detalle_Tarimas['SKU'].astype(str).str.strip().dropna().unique())
                        
                    # Los SKUs válidos son la unión de Artículos y SKUs Autorizados
                    set_skus_art = set(st.session_state.BD_Articulos['SKU'].astype(str).str.strip().unique()) if "BD_Articulos" in st.session_state and not st.session_state.BD_Articulos.empty else set()
                    set_skus_aut = set(st.session_state.BD_SKUs_Autorizados['SKU'].astype(str).str.strip().unique()) if "BD_SKUs_Autorizados" in st.session_state and not st.session_state.BD_SKUs_Autorizados.empty else set()
                    skus_validos_sistema = set_skus_art | set_skus_aut
                    
                    skus_sin_registro = sorted(list(skus_en_tarimas - skus_validos_sistema))
                    skus_sin_registro = [s for s in skus_sin_registro if s and s.upper() not in ["TODOS", "SELECCIONE UN SKU..."]]
                    
                    if len(skus_sin_registro) > 0:
                        st.info(f"🔍 Se encontraron **{len(skus_sin_registro)}** artículos en tarimas que no están registrados en ninguna base de datos.")
                    else:
                        st.success("🎉 ¡Todos los artículos en tarimas están registrados en el sistema!")
                        
                    df_sin_registro = pd.DataFrame(columns=columnas_oficiales)
                    df_sin_registro["SKU"] = skus_sin_registro
                    
                    buf_sin_reg = io.BytesIO()
                    with pd.ExcelWriter(buf_sin_reg, engine='openpyxl') as p_writer:
                        df_sin_registro.to_excel(p_writer, index=False, sheet_name='Datos_Sistema')
                        worksheet = p_writer.sheets['Datos_Sistema']
                        worksheet.column_dimensions['A'].width = 20
                        worksheet.column_dimensions['B'].width = 25
                        worksheet.column_dimensions['C'].width = 20
                        worksheet.column_dimensions['D'].width = 25
                        worksheet.column_dimensions['E'].width = 20
                        
                        dv_calibre_sin = DataValidation(type="list", formula1=opciones_calibres, allow_blank=True)
                        worksheet.add_data_validation(dv_calibre_sin)
                        dv_calibre_sin.add("C2:C2000")
                        
                        dv_acabado_sin = DataValidation(type="list", formula1=opciones_acabados, allow_blank=True)
                        worksheet.add_data_validation(dv_acabado_sin)
                        dv_acabado_sin.add("E2:E2000")
                        
                    buf_sin_reg.seek(0)
                    st.download_button(
                        label="📥 Artículos Sin Registro",
                        data=buf_sin_reg.getvalue(),
                        file_name="Plantilla_Articulos_Sin_Registro.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key="btn_download_articulos_sin_registro",
                        disabled=len(skus_sin_registro) == 0,
                        use_container_width=True
                    )
                    
                    st.write("---")
                    st.write("##### Artículos en Órdenes de Compra (POs) Sin Registro en Catálogo")
                    
                    skus_en_pos = set()
                    if "BD_Requerimientos_POs" in st.session_state and not st.session_state.BD_Requerimientos_POs.empty:
                        skus_en_pos = set(st.session_state.BD_Requerimientos_POs['SKU'].astype(str).str.strip().dropna().unique())
                        
                    skus_sin_registro_pos = sorted(list(skus_en_pos - skus_validos_sistema))
                    skus_sin_registro_pos = [s for s in skus_sin_registro_pos if s and s.upper() not in ["TODOS", "SELECCIONE UN SKU..."]]
                    
                    if len(skus_sin_registro_pos) > 0:
                        st.info(f"🔍 Se encontraron **{len(skus_sin_registro_pos)}** artículos en Órdenes de Compra (POs) que no están registrados en el catálogo.")
                    else:
                        st.success("🎉 ¡Todos los artículos en Órdenes de Compra (POs) están registrados en el sistema!")
                        
                    df_sin_registro_pos = pd.DataFrame(columns=columnas_oficiales)
                    df_sin_registro_pos["SKU"] = skus_sin_registro_pos
                    
                    buf_sin_reg_pos = io.BytesIO()
                    with pd.ExcelWriter(buf_sin_reg_pos, engine='openpyxl') as p_writer:
                        df_sin_registro_pos.to_excel(p_writer, index=False, sheet_name='Datos_Sistema')
                        worksheet = p_writer.sheets['Datos_Sistema']
                        worksheet.column_dimensions['A'].width = 20
                        worksheet.column_dimensions['B'].width = 25
                        worksheet.column_dimensions['C'].width = 20
                        worksheet.column_dimensions['D'].width = 25
                        worksheet.column_dimensions['E'].width = 20
                        
                        dv_calibre_sin_pos = DataValidation(type="list", formula1=opciones_calibres, allow_blank=True)
                        worksheet.add_data_validation(dv_calibre_sin_pos)
                        dv_calibre_sin_pos.add("C2:C2000")
                        
                        dv_acabado_sin_pos = DataValidation(type="list", formula1=opciones_acabados, allow_blank=True)
                        worksheet.add_data_validation(dv_acabado_sin_pos)
                        dv_acabado_sin_pos.add("E2:E2000")
                        
                    buf_sin_reg_pos.seek(0)
                    st.download_button(
                        label="📥 Artículos de POs Sin Registro",
                        data=buf_sin_reg_pos.getvalue(),
                        file_name="Plantilla_Articulos_POs_Sin_Registro.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key="btn_download_articulos_pos_sin_registro",
                        disabled=len(skus_sin_registro_pos) == 0,
                        use_container_width=True
                    )
                    
                with c_art2:
                    st.write("##### Cargar / Sobrescribir Catálogo Maestro")
                    arch_articulos = st.file_uploader("Suba la Plantilla de Artículos Rellenada (Estricta 5 columnas):", type=["xlsx"], key="uploader_articulos_masivo_f")
                    
                    if arch_articulos:
                        if st.button("🔄 Procesar e Integrar Catálogo Maestro", use_container_width=True):
                            try:
                                df_art_excel = pd.read_excel(arch_articulos)
                                columnas_requeridas = ["SKU", "Nombre", "Calibre_Espesor", "Dimensiones_Pieza", "Acabado_Superficial"]
                                
                                columnas_correctas = True
                                for col in columnas_requeridas:
                                    if col not in df_art_excel.columns:
                                        columnas_correctas = False
                                        
                                if not columnas_correctas:
                                    st.error("❌ Error: Columnas incompatibles. Use la estructura oficial de 5 columnas.")
                                else:
                                    df_art_excel = df_art_excel.dropna(subset=["SKU"])
                                    df_art_excel["SKU"] = df_art_excel["SKU"].astype(str).str.strip().str.upper()
                                    
                                    # Conserva anteriores y actualiza los nuevos
                                    if "BD_Articulos" in st.session_state and not st.session_state.BD_Articulos.empty:
                                        df_anterior = st.session_state.BD_Articulos.copy()
                                        df_anterior["SKU"] = df_anterior["SKU"].astype(str).str.strip().str.upper()
                                        df_anterior = df_anterior[~df_anterior["SKU"].isin(df_art_excel["SKU"])]
                                        st.session_state.BD_Articulos = pd.concat([df_anterior, df_art_excel], ignore_index=True)
                                    else:
                                        st.session_state.BD_Articulos = df_art_excel
                                        
                                    if subir_excel_a_github("BD_Articulos.xlsx", st.session_state.BD_Articulos):
                                        st.success("✅ ¡Catálogo maestro actualizado en GitHub!")
                                        st.rerun()
                                    else:
                                        st.error("❌ Error al guardar en GitHub.")
                            except Exception as e:
                                st.error(f"Error: {e}")
                                
            st.write("##### ✏️ Edición Directa, Alta y Baja del Catálogo Maestro Detallado")
            if "BD_Articulos" in st.session_state and not st.session_state.BD_Articulos.empty:
                df_art_editable = st.data_editor(
                    st.session_state.BD_Articulos,
                    use_container_width=True,
                    disabled=["SKU"],
                    hide_index=True,
                    num_rows="dynamic",
                    key="editor_mantenimiento_articulos_directo_v3",
                    column_config={
                        "SKU": st.column_config.TextColumn("SKU (Bloqueado/Llave)"),
                        "Nombre": st.column_config.TextColumn("Descripción Comercial"),
                        "Calibre_Espesor": st.column_config.SelectboxColumn("Calibre / Espesor", options=["10GA", "12GA", "14GA", "16GA", "10GACR", "12GACR", "14GACR", "16GACR", "125AL", "250AL", "188AL"]),
                        "Dimensiones_Pieza": st.column_config.TextColumn("Dimensiones"),
                        "Acabado_Superficial": st.column_config.SelectboxColumn("Acabado Superficial", options=["Decapado", "Ansi 61", "Galvanizado", "Otro"])
                    }
                )
                if st.button("💾 Guardar Cambios del Catálogo Maestro en GitHub", use_container_width=True):
                    df_art_final = df_art_editable.dropna(subset=["SKU"])
                    df_art_final["SKU"] = df_art_final["SKU"].astype(str).str.strip().str.upper()
                    st.session_state.BD_Articulos = df_art_final
                    if subir_excel_a_github("BD_Articulos.xlsx", st.session_state.BD_Articulos):
                        st.success("💥 Modificaciones de catálogo maestro sincronizadas.")
                        st.rerun()
                    else:
                        st.error("❌ Error de comunicación con GitHub.")
            else:
                st.warning("⚠️ No existen registros activos en el catálogo detallado.")
                
        with sub_tab2:
            st.markdown("#### 📋 Lista de SKUs Autorizados (Nombres de SKU)")
            st.info("💡 **Guía de uso:** Esta es una base de datos simple de una sola columna. Úsela para subir listas masivas de SKUs autorizados (como su lista mensual) sin necesidad de llenar descripciones o calibre. No se mezclará con el catálogo detallado.")
            
            with st.expander("📥 Carga Masiva de SKUs Autorizados mediante Excel", expanded=False):
                c_aut1, c_aut2 = st.columns(2)
                
                with c_aut1:
                    st.write("##### Obtener Plantilla Simple de SKUs")
                    df_plantilla_aut = pd.DataFrame(columns=["SKU"])
                    buf_p_aut = io.BytesIO()
                    df_plantilla_aut.to_excel(buf_p_aut, index=False, sheet_name="SKUs")
                    buf_p_aut.seek(0)
                    st.download_button(
                        label="📥 Descargar Plantilla Simple SKUs (.xlsx)",
                        data=buf_p_aut.getvalue(),
                        file_name="Plantilla_Simple_SKUs.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key="btn_download_plantilla_simple_skus"
                    )
                    
                with c_aut2:
                    st.write("##### Subir / Actualizar Lista de SKUs")
                    arch_skus_aut = st.file_uploader("Suba el Excel con los SKUs autorizados (Solo columna SKU):", type=["xlsx"], key="uploader_skus_autorizados")
                    
                    if arch_skus_aut:
                        if st.button("🔄 Procesar e Integrar Lista de SKUs", use_container_width=True):
                            try:
                                df_excel_aut = pd.read_excel(arch_skus_aut)
                                df_excel_aut.columns = [str(c).strip() for c in df_excel_aut.columns]
                                
                                sku_col = None
                                for col in df_excel_aut.columns:
                                    if col.upper() == "SKU":
                                        sku_col = col
                                        break
                                        
                                if sku_col is None:
                                    st.error("❌ Error: El archivo debe contener una columna llamada 'SKU'.")
                                else:
                                    df_excel_aut = df_excel_aut[[sku_col]].rename(columns={sku_col: "SKU"}).dropna(subset=["SKU"])
                                    df_excel_aut["SKU"] = df_excel_aut["SKU"].astype(str).str.strip().str.upper()
                                    
                                    # Integración inteligente: conserva anteriores y suma nuevos
                                    if "BD_SKUs_Autorizados" in st.session_state and not st.session_state.BD_SKUs_Autorizados.empty:
                                        df_prev = st.session_state.BD_SKUs_Autorizados.copy()
                                        df_prev["SKU"] = df_prev["SKU"].astype(str).str.strip().str.upper()
                                        df_prev = df_prev[~df_prev["SKU"].isin(df_excel_aut["SKU"])]
                                        st.session_state.BD_SKUs_Autorizados = pd.concat([df_prev, df_excel_aut], ignore_index=True)
                                    else:
                                        st.session_state.BD_SKUs_Autorizados = df_excel_aut
                                        
                                    if subir_excel_a_github("BD_SKUs_Autorizados.xlsx", st.session_state.BD_SKUs_Autorizados):
                                        st.success("✅ ¡Lista de SKUs Autorizados integrada con éxito en GitHub!")
                                        st.rerun()
                                    else:
                                        st.error("❌ Error al guardar en GitHub.")
                            except Exception as e:
                                st.error(f"Error al procesar: {e}")
                                
            st.write("##### ✏️ Edición y Bajas de la Lista de SKUs Autorizados")
            if "BD_SKUs_Autorizados" in st.session_state and not st.session_state.BD_SKUs_Autorizados.empty:
                df_aut_editable = st.data_editor(
                    st.session_state.BD_SKUs_Autorizados,
                    use_container_width=True,
                    hide_index=True,
                    num_rows="dynamic",
                    key="editor_skus_autorizados_directo",
                    column_config={"SKU": st.column_config.TextColumn("SKU Autorizado")}
                )
                if st.button("💾 Guardar Cambios de SKUs Autorizados en GitHub", use_container_width=True):
                    df_aut_final = df_aut_editable.dropna(subset=["SKU"])
                    df_aut_final["SKU"] = df_aut_final["SKU"].astype(str).str.strip().str.upper()
                    st.session_state.BD_SKUs_Autorizados = df_aut_final
                    if subir_excel_a_github("BD_SKUs_Autorizados.xlsx", st.session_state.BD_SKUs_Autorizados):
                        st.success("✅ Cambios en SKUs Autorizados guardados exitosamente.")
                        st.rerun()
                    else:
                        st.error("❌ Error al guardar modificaciones en GitHub.")
            else:
                st.warning("⚠️ No existen registros activos en la lista de SKUs autorizados.")
    # =============================================================================
    # PESTAÑA 5: CONFIGURADOR DEL FOLIO CONSECUTIVO TPM
    # =============================================================================
    try:
        # Intentamos usar la variable de la quinta pestaña si existe
        with tab5:
            st.write("##### 🔢 Inicializar o Cambiar Consecutivo de Tarimas (TPM)")
            st.info("💡 **Guía de uso:** Defina el número numérico en el que desea que continúe la siguiente tarima que genere el sistema (Ejemplo: Si pone 56, la siguiente será TPM-0056).")
    
            if "siguiente_numero_tpm" not in st.session_state or st.session_state["siguiente_numero_tpm"] is None:
                st.session_state["siguiente_numero_tpm"] = obtener_siguiente_consecutivo_tpm()
    
            nuevo_consecutivo = st.number_input(
                "Indique el siguiente número de folio a generar:",
                min_value=1, max_value=9999,
                value=int(st.session_state["siguiente_numero_tpm"]),
                step=1, key="input_consecutivo_manual_tpm"
            )
    
            col_btn1, col_btn2 = st.columns(2)
            with col_btn1:
                if st.button("💾 Guardar Nuevo Consecutivo de Folio", use_container_width=True):
                    st.session_state["siguiente_numero_tpm"] = nuevo_consecutivo
                    try:
                        with open("consecutivo_override.txt", "w", encoding="utf-8") as f:
                            f.write(str(nuevo_consecutivo))
                    except Exception as e:
                        st.warning(f"⚠️ No se pudo guardar el consecutivo en disco: {e}")
                    st.success(f"💥 ¡Consecutivo actualizado! La próxima tarima nueva se creará con el folio: **TPM-{nuevo_consecutivo:04d}**")
                    st.rerun()
            with col_btn2:
                if st.button("🔄 Restablecer a Consecutivo Automático", use_container_width=True):
                    import os
                    if os.path.exists("consecutivo_override.txt"):
                        try:
                            os.remove("consecutivo_override.txt")
                        except Exception:
                            pass
                    st.session_state["siguiente_numero_tpm"] = obtener_siguiente_consecutivo_tpm()
                    st.success("🔄 Restablecido al siguiente consecutivo automático de la base de datos.")
                    st.rerun()
        with tab6:
            renderizar_explorador_imagenes()
    except NameError:
        # Si tab5 no existe, desplegamos un contenedor nativo expandible para que no rompa la aplicación
        with st.expander("🔢 Contador y Consecutivo de Tarimas (TPM)", expanded=True):
            st.write("##### 🔢 Inicializar o Cambiar Consecutivo de Tarimas (TPM)")
            st.info("💡 **Guía de uso:** Defina el número numérico en el que desea que continúe la siguiente tarima que genere el sistema (Ejemplo: Si pone 56, la siguiente será TPM-0056).")
    
            if "siguiente_numero_tpm" not in st.session_state or st.session_state["siguiente_numero_tpm"] is None:
                st.session_state["siguiente_numero_tpm"] = obtener_siguiente_consecutivo_tpm()
    
            nuevo_consecutivo = st.number_input(
                "Indique el siguiente número de folio a generar:",
                min_value=1, max_value=9999,
                value=int(st.session_state["siguiente_numero_tpm"]),
                step=1, key="input_consecutivo_manual_tpm_fallback"
            )
    
            col_btn1, col_btn2 = st.columns(2)
            with col_btn1:
                if st.button("💾 Guardar Nuevo Consecutivo de Folio ", use_container_width=True):
                    st.session_state["siguiente_numero_tpm"] = nuevo_consecutivo
                    try:
                        with open("consecutivo_override.txt", "w", encoding="utf-8") as f:
                            f.write(str(nuevo_consecutivo))
                    except Exception as e:
                        st.warning(f"⚠️ No se pudo guardar el consecutivo en disco: {e}")
                    st.success(f"💥 ¡Consecutivo actualizado! La próxima tarima nueva se creará con el folio: **TPM-{nuevo_consecutivo:04d}**")
                    st.rerun()
            with col_btn2:
                if st.button("🔄 Restablecer a Consecutivo Automático ", use_container_width=True):
                    import os
                    if os.path.exists("consecutivo_override.txt"):
                        try:
                            os.remove("consecutivo_override.txt")
                        except Exception:
                            pass
                    st.session_state["siguiente_numero_tpm"] = obtener_siguiente_consecutivo_tpm()
                    st.success("🔄 Restablecido al siguiente consecutivo automático de la base de datos.")
                    st.rerun()

        with st.expander("🖼️ Carpeta de Imágenes de Artículos (Mantenimiento)", expanded=False):
            renderizar_explorador_imagenes()

    # --- SUB-MÓDULO 7: CATÁLOGO DE RECEPTORES / DESTINOS ---
    with tab7:
        st.subheader("🏢 Catálogo de Receptores / Destinos de Remisiones")
        
        # 1. Plantilla descargable
        df_r_plantilla = pd.DataFrame([
            {"Nombre_Receptor": "Cliente Ejemplo 1", "Direccion": "Calle Principal 123, Ciudad, Estado"},
            {"Nombre_Receptor": "Cliente Ejemplo 2", "Direccion": "Av. Tecnologico 456, Ciudad, Estado"}
        ])
        
        from openpyxl.styles import Font, PatternFill
        buf_r = io.BytesIO()
        with pd.ExcelWriter(buf_r, engine='openpyxl') as wr_r:
            df_r_plantilla.to_excel(wr_r, index=False, sheet_name='Plantilla_Receptores')
            ws_r = wr_r.sheets['Plantilla_Receptores']
            
            fill_header = PatternFill(start_color="757575", end_color="757575", fill_type="solid")
            font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
            for col_idx in range(1, 3):
                cell = ws_r.cell(row=1, column=col_idx)
                cell.font, cell.fill = font_header, fill_header
                
        st.write("Descargue el formato base para rellenar la lista de receptores autorizados:")
        st.download_button(label="📥 Descargar Plantilla de Receptores (.xlsx)", data=buf_r.getvalue(), file_name="plantilla_receptores_sigrama.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key="btn_download_plantilla_receptores")
        st.write("---")
        
        # 2. Carga masiva
        arch_receptores = st.file_uploader("Suba la Plantilla de Receptores Rellenada:", type=["xlsx"], key="uploader_receptores_masivo")
        if arch_receptores and st.button("🚀 Procesar e Integrar Receptores Masivos"):
            try:
                df_r_excel = pd.read_excel(arch_receptores)
                if not all(col in df_r_excel.columns for col in ["Nombre_Receptor", "Direccion"]):
                    st.error("❌ Error: Columnas incompatibles. Use: Nombre_Receptor, Direccion")
                else:
                    for _, row_r in df_r_excel.iterrows():
                        if pd.notna(row_r['Nombre_Receptor']) and str(row_r['Nombre_Receptor']).strip() != "":
                            n_id_r = f"REC-{(len(st.session_state.BD_Receptores) + 1):02d}"
                            n_r_row = {"ID_Receptor": n_id_r, "Nombre_Receptor": str(row_r['Nombre_Receptor']).strip(), "Direccion": str(row_r['Direccion']).strip() if pd.notna(row_r['Direccion']) else "N/A", "Estatus": "Activo"}
                            st.session_state.BD_Receptores = pd.concat([st.session_state.BD_Receptores, pd.DataFrame([n_r_row])], ignore_index=True)
                    subir_excel_a_github("BD_Receptores.xlsx", st.session_state.BD_Receptores)
                    st.success("✅ ¡Lista de receptores integrada y respaldada con éxito!"); st.rerun()
            except Exception as e: st.error(f"Error: {e}")
            
        st.write("---")
        
        # 3. Alta individual
        with st.expander("➕ Dar de Alta Nuevo Receptor Individual"):
            c_rec1, c_rec2 = st.columns(2)
            with c_rec1: nuevo_rec_nom = st.text_input("Nombre del Receptor / Cliente:", key="txt_input_nuevo_receptor_name")
            with c_rec2: nueva_rec_dir = st.text_input("Dirección del Destino:", key="txt_input_nuevo_receptor_dir")
            if st.button("➕ Registrar Receptor Individual"):
                if nuevo_rec_nom:
                    n_id_r = f"REC-{(len(st.session_state.BD_Receptores) + 1):02d}"
                    n_row = {"ID_Receptor": n_id_r, "Nombre_Receptor": nuevo_rec_nom.strip(), "Direccion": nueva_rec_dir.strip(), "Estatus": "Activo"}
                    st.session_state.BD_Receptores = pd.concat([st.session_state.BD_Receptores, pd.DataFrame([n_row])], ignore_index=True)
                    subir_excel_a_github("BD_Receptores.xlsx", st.session_state.BD_Receptores)
                    st.success("Receptor registrado exitosamente."); st.rerun()
                    
        # 4. Tabla de edición y visualización
        st.write("📋 Catálogo Maestro de Receptores Autorizados (Editable):")
        df_r_edit = st.data_editor(st.session_state.BD_Receptores, use_container_width=True, hide_index=True, key="editor_catalogo_receptores_master")
        if st.button("💾 Sincronizar Cambios Manuales de Receptores"):
            st.session_state.BD_Receptores = df_r_edit
            subir_excel_a_github("BD_Receptores.xlsx", st.session_state.BD_Receptores)
            st.success("Catálogo de receptores actualizado en GitHub.")

    with tab8:
        st.subheader("📧 Configuración de Listas de Distribución de Correo")
        st.markdown("Defina los correos electrónicos por defecto para el envío de notificaciones de salida de material. Puede incluir múltiples direcciones separándolas por punto y coma (`;`).")
        
        cfg_actual = obtener_emails_config()
        
        cfg_to = st.text_area("Destinatarios Principales (Para):", value=cfg_actual.get("dest_to", ""), help="Ejemplo: logistica@sigrama.com.mx; almacen@sigrama.com.mx")
        cfg_cc = st.text_area("Con Copia (CC):", value=cfg_actual.get("dest_cc", ""), help="Ejemplo: calidad@sigrama.com.mx")
        
        if st.button("💾 Guardar Listas de Distribución", use_container_width=True):
            if guardar_emails_config(cfg_to, cfg_cc):
                st.success("Configuración de correos guardada exitosamente.")
                st.rerun()

# =============================================================================================
# 15. CARGA MASIVA HISTÓRICA
# =============================================================================
elif opcion_menu == "🕰️ Carga Histórica":
    st.title("🕰️ Carga Masiva de Remisiones y Tarimas Históricas")
    st.markdown("Esta sección permite cargar un archivo de Excel para dar de alta tarimas antiguas y registrarlas inmediatamente como remisiones enviadas, consolidando tu inventario histórico en un solo paso.")
    
    st.warning("⚠️ **ATENCIÓN:** El archivo de Excel debe contener estrictamente las siguientes columnas: **Tarima, Producto/SKU, PO, Proyecto, Parcialidad, Descripcion, Cantidad, Fecha de Remisión, Folio de Remisión**.")
    
    with st.form("form_carga_historica"):
        col1, col2 = st.columns(2)
        with col1:
            st.info("💡 **NUEVO:** El Folio de Remisión (Ej. E0001) ahora se leerá automáticamente de tu archivo Excel para cada tarima, permitiéndote subir múltiples remisiones en un solo documento.")
            receptor_input = st.text_input("Receptor (Destino para todas):", value="Galvatec Industrias")
        with col2:
            direccion_input = st.text_input("Dirección del Receptor:", value="Prol. Valle Guadiana 919, Parque Industrial II, 35078 Gómez Palacio, Dgo.")
            
        archivo_cargado = st.file_uploader("📥 Selecciona el archivo Excel con el historial", type=['xlsx'])
        btn_procesar = st.form_submit_button("🚀 Procesar Carga Histórica")
        
    if btn_procesar:
        if not archivo_cargado:
            st.error("❌ Debes subir el archivo Excel.")
        else:
            try:
                df_hist = pd.read_excel(archivo_cargado)
                # Soportar ambas formas del encabezado: "Folio de Remisión" o "Folio Remisión"
                if "Folio Remisión" in df_hist.columns and "Folio de Remisión" not in df_hist.columns:
                    df_hist = df_hist.rename(columns={"Folio Remisión": "Folio de Remisión"})
                    
                cols_requeridas = ["Tarima", "Producto/SKU", "PO", "Proyecto", "Parcialidad", "Descripcion", "Cantidad", "Fecha de Remisión", "Folio de Remisión"]
                
                # Validar columnas
                faltantes = [c for c in cols_requeridas if c not in df_hist.columns]
                if faltantes:
                    st.error(f"❌ El archivo no tiene el formato correcto. Faltan las columnas: {', '.join(faltantes)}")
                else:
                    # Auditoría de Datos
                    st.session_state['df_historico_auditado'] = df_hist
                    st.session_state['receptor_auditado'] = receptor_input
                    st.session_state['direccion_auditada'] = direccion_input
            except Exception as e:
                st.error(f"❌ Ocurrió un error al leer el archivo: {e}")

    # Sección de Auditoría y Confirmación
    if 'df_historico_auditado' in st.session_state:
        df_hist = st.session_state['df_historico_auditado']
        receptor_input = st.session_state['receptor_auditado']
        direccion_input = st.session_state['direccion_auditada']
        
        st.markdown("---")
        st.subheader("📋 Auditoría de Información a Cargar")
        
        tarimas_excel = df_hist['Tarima'].dropna().unique().tolist()
        folios_excel = df_hist['Folio de Remisión'].dropna().unique().tolist()
        skus_excel = df_hist['Producto/SKU'].dropna().unique().tolist()
        
        # Validar si los folios ya existen
        folios_existentes = []
        if not st.session_state.BD_Datos_Generales_Remision.empty:
            folios_bd = st.session_state.BD_Datos_Generales_Remision['Folio_Remision'].astype(str).str.strip().tolist()
            for f in folios_excel:
                if str(f).strip() in folios_bd:
                    folios_existentes.append(str(f).strip())
                    
        # Validar si los SKUs existen
        skus_no_encontrados = []
        skus_art = set(st.session_state.BD_Articulos['SKU'].astype(str).str.strip().tolist()) if "BD_Articulos" in st.session_state and not st.session_state.BD_Articulos.empty else set()
        skus_aut = set(st.session_state.BD_SKUs_Autorizados['SKU'].astype(str).str.strip().tolist()) if "BD_SKUs_Autorizados" in st.session_state and not st.session_state.BD_SKUs_Autorizados.empty else set()
        skus_bd = skus_art | skus_aut
        for s in skus_excel:
            if str(s).strip().upper() not in skus_bd:
                skus_no_encontrados.append(str(s).strip())
                    
        col_res1, col_res2 = st.columns(2)
        with col_res1:
            st.metric("Total Tarimas a Crear", len(tarimas_excel))
            st.metric("Total Remisiones a Generar", len(folios_excel))
        with col_res2:
            st.metric("Total de Partidas (Filas)", len(df_hist))
            st.metric("SKUs únicos procesados", len(skus_excel))
            
        hay_conflictos = False
        if folios_existentes:
            hay_conflictos = True
            st.error(f"⚠️ **ALERTA DE CONFLICTO:** Los siguientes folios ya existen en el sistema: {', '.join(folios_existentes)}. Por seguridad, no puedes cargar duplicados.")
            
        if skus_no_encontrados:
            st.warning(f"⚠️ **ADVERTENCIA:** Se encontraron {len(skus_no_encontrados)} SKUs que no están en el Catálogo de Artículos. (Se pueden cargar, pero la descripción no se enriquecerá automáticamente).")
            
        if hay_conflictos:
            st.error("❌ Debes corregir el archivo Excel y volver a subirlo para evitar duplicar folios.")
            if st.button("Cancelar Carga"):
                del st.session_state['df_historico_auditado']
                st.rerun()
        else:
            st.success("✅ **Auditoría superada sin conflictos graves.** La información está lista para consolidarse.")
            
            if st.button("✅ Confirmar y Ejecutar Carga Masiva", use_container_width=True):
                with st.spinner("Procesando la carga en las bases de datos..."):
                    try:
                        mapa_tpm = {}
                        nuevo_id = obtener_siguiente_consecutivo_tpm()
                        for t in tarimas_excel:
                            mapa_tpm[t] = f"TPM-{nuevo_id:04d}"
                            nuevo_id += 1
                            
                        st.session_state["siguiente_numero_tpm"] = nuevo_id
                        
                        # 3. Preparar BD_Tarimas
                        nuevas_tarimas = []
                        for t in tarimas_excel:
                            sub_df = df_hist[df_hist['Tarima'] == t]
                            fecha_str = str(sub_df.iloc[0]['Fecha de Remisión']).split(' ')[0]
                            nuevas_tarimas.append({
                                "ID_Tarima": mapa_tpm[t],
                                "Tarima_Origen_Excel": t,
                                "Fecha_Creacion": fecha_str,
                                "Ubicacion_Actual": "Planta Rio",
                                "Creado_Por": "Carga Historica",
                                "Tipo_Tarima": "Mixta",
                                "Estatus": "Remesada",
                                "Es_Nueva": "No"
                            })
                        df_nuevas_tarimas = pd.DataFrame(nuevas_tarimas)
                        st.session_state.BD_Tarimas = pd.concat([st.session_state.BD_Tarimas, df_nuevas_tarimas], ignore_index=True)
                        
                        # 4. Preparar BD_Detalle_Tarimas
                        nuevos_detalles = []
                        id_det = 1
                        if not st.session_state.BD_Detalle_Tarimas.empty:
                            try:
                                id_det = int(st.session_state.BD_Detalle_Tarimas['ID_Detalle'].max()) + 1
                            except:
                                id_det = len(st.session_state.BD_Detalle_Tarimas) + 1
                                
                        for _, row in df_hist.iterrows():
                            t_excel = row['Tarima']
                            nuevos_detalles.append({
                                "ID_Detalle": id_det,
                                "ID_Tarima": mapa_tpm[t_excel],
                                "SKU": row['Producto/SKU'],
                                "PO": clean_po_val(row['PO']),
                                "Proyecto": clean_project_val(row['Proyecto']),
                                "Parcialidad": row['Parcialidad'],
                                "Descripcion": row['Descripcion'],
                                "Cantidad": row['Cantidad']
                            })
                            id_det += 1
                        df_nuevos_detalles = pd.DataFrame(nuevos_detalles)
                        st.session_state.BD_Detalle_Tarimas = pd.concat([st.session_state.BD_Detalle_Tarimas, df_nuevos_detalles], ignore_index=True)
                        
                        # 5. Preparar BD_Datos_Generales_Remision (Agrupando por Folio)
                        import json
                        nuevas_remisiones = []
                        
                        id_rem = 1
                        if not st.session_state.BD_Datos_Generales_Remision.empty:
                            try:
                                id_rem = int(st.session_state.BD_Datos_Generales_Remision['ID_Remision'].max()) + 1
                            except:
                                id_rem = len(st.session_state.BD_Datos_Generales_Remision) + 1
                                
                        for folio in folios_excel:
                            sub_df_folio = df_hist[df_hist['Folio de Remisión'] == folio]
                            fecha_rem = str(sub_df_folio.iloc[0]['Fecha de Remisión']).split(' ')[0]
                            tarimas_de_este_folio_excel = sub_df_folio['Tarima'].dropna().unique().tolist()
                            tarimas_asignadas_list = [mapa_tpm[t] for t in tarimas_de_este_folio_excel]
                            
                            nuevas_remisiones.append({
                                "ID_Remision": id_rem,
                                "Folio_Remision": str(folio).strip(),
                                "Fecha_Hora_Salida": fecha_rem,
                                "Nombre_Emisor": "Carga Historica",
                                "Direccion_Emisor": "N/A",
                                "Nombre_Receptor": receptor_input.strip(),
                                "Direccion_Receptor": direccion_input.strip(),
                                "Tarimas_Asociadas": json.dumps(tarimas_asignadas_list)
                            })
                            id_rem += 1
                            
                        df_nuevas_remisiones = pd.DataFrame(nuevas_remisiones)
                        st.session_state.BD_Datos_Generales_Remision = pd.concat([st.session_state.BD_Datos_Generales_Remision, df_nuevas_remisiones], ignore_index=True)
                        
                        # 6. Guardar todo
                        subir_excel_a_github("BD_Tarimas.xlsx", st.session_state.BD_Tarimas)
                        subir_excel_a_github("BD_Detalle_Tarimas.xlsx", st.session_state.BD_Detalle_Tarimas)
                        subir_excel_a_github("BD_Datos_Generales_Remision.xlsx", st.session_state.BD_Datos_Generales_Remision)
                        
                        try:
                            with open("consecutivo_override.txt", "w", encoding="utf-8") as f:
                                f.write(str(nuevo_id))
                        except:
                            pass
                            
                        del st.session_state['df_historico_auditado']
                        st.success(f"✅ ¡Carga Exitosa! Se crearon {len(tarimas_excel)} tarimas en {len(folios_excel)} remisiones.")
                        st.balloons()
                    except Exception as e:
                        st.error(f"❌ Ocurrió un error crítico durante la carga: {e}")
                        
# =============================================================================
# 16. ANÁLISIS DE FALTANTES POR FABRICAR
# =============================================================================
elif opcion_menu == "📉 Análisis de Faltantes":
    st.title("📉 Análisis de Faltantes de Producción")
    st.markdown("Gestione los requerimientos de sus Órdenes de Compra (POs) y visualice la dispersión de entregas, stock y faltantes de producción en tiempo real.")
    
    # Asegurar inicialización en sesión
    if "BD_POs_Cabecera" not in st.session_state:
        st.session_state.BD_POs_Cabecera = pd.DataFrame(columns=["PO", "Fecha_Pedido", "Proyecto", "Solicitante", "Requisicion", "Destino"])
    if "BD_Requerimientos_POs" not in st.session_state:
        st.session_state.BD_Requerimientos_POs = pd.DataFrame(columns=["PO", "SKU", "Fecha_Entrega", "Cantidad_Requerida", "Parcialidad"])

    tab_ins, tab_mat, tab_dl = st.tabs([
        "📥 Ingreso de Órdenes de Compra (POs)", 
        "📊 Matriz de Avance por PO", 
        "📉 Descarga de Reportes"
    ])
    
    with tab_ins:
        st.write("")
        st.subheader("📥 Carga de Requerimientos Oficiales")
        
        # Botón para descargar plantilla de 2 hojas
        df_template_gen = pd.DataFrame([{
            "PO": "26032107",
            "Fecha_Pedido": "12/06/2026",
            "Proyecto": "SWBD R5",
            "Solicitante": "MAURICIO DOMINGUEZ",
            "Requisicion": "22035",
            "Destino": "ALMACEN SIGRAMA RIO XIX"
        }])
        
        df_template_det = pd.DataFrame([{
            "ID": "12-A-6359-02",
            "23-feb-2026": 6,
            "01-mar-2026": "-",
            "13-mar-2026": 13,
            "01-abr-2026": "-",
            "13-abr-2026": 12,
            "13-may-2026": 5,
            "13-jun-2026": "-",
            "14-jul-2026": 5,
            "13-ago-2026": 5,
            "13-sep-2026": 12
        }, {
            "ID": "P17690-17",
            "23-feb-2026": 15,
            "01-mar-2026": "-",
            "13-mar-2026": 33,
            "01-abr-2026": "-",
            "13-abr-2026": 30,
            "13-may-2026": 12,
            "13-jun-2026": "-",
            "14-jul-2026": 12,
            "13-ago-2026": 12,
            "13-sep-2026": 30
        }])
        
        buf_t = io.BytesIO()
        with pd.ExcelWriter(buf_t, engine='openpyxl') as writer_t:
            df_template_gen.to_excel(writer_t, sheet_name="Datos_Generales", index=False)
            df_template_det.to_excel(writer_t, sheet_name="Detalle_Entregas", index=False)
            
            from openpyxl.styles import Font, PatternFill, Alignment
            workbook_t = writer_t.book
            
            header_fill = PatternFill(start_color="EC2024", end_color="EC2024", fill_type="solid")
            header_font = Font(name="Arial", color="FFFFFF", bold=True, size=11)
            center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            # Formatear Datos_Generales
            sheet_gen = workbook_t["Datos_Generales"]
            for cell in sheet_gen[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_alignment
            
            # Ancho de columnas para Datos_Generales
            for col in sheet_gen.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                col_letter = col[0].column_letter
                sheet_gen.column_dimensions[col_letter].width = max(max_len + 5, 22)
                
            # Formatear Detalle_Entregas
            sheet_det = workbook_t["Detalle_Entregas"]
            for cell in sheet_det[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_alignment
                
            # Ancho de columnas para Detalle_Entregas
            for col in sheet_det.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                col_letter = col[0].column_letter
                sheet_det.column_dimensions[col_letter].width = max(max_len + 5, 18)
        
        st.download_button(
            label="📥 Descargar Plantilla Oficial de Carga (2 Hojas)",
            data=buf_t.getvalue(),
            file_name="Plantilla_Requerimientos_PO.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        
        st.write("---")
        excel_po = st.file_uploader("Subir Archivo de Requerimientos de PO (Excel):", type=['xlsx'], key="po_file_uploader")
        modo_carga = st.radio("Método de Carga:", ["Acumular registros en la base de datos", "Sobrescribir base de datos completa"], horizontal=True, key="po_load_mode")
        
        if excel_po:
            if st.button("🚀 Integrar Requerimientos", use_container_width=True):
                try:
                    xl = pd.ExcelFile(excel_po)
                    if "Datos_Generales" in xl.sheet_names and "Detalle_Entregas" in xl.sheet_names:
                        df_gen = xl.parse("Datos_Generales")
                        df_det_matrix = xl.parse("Detalle_Entregas")
                        
                        if df_gen.empty or df_det_matrix.empty:
                            st.error("❌ Una o ambas pestañas están vacías.")
                        else:
                            # 1. Procesar Cabecera
                            row_gen = df_gen.iloc[0].to_dict()
                            po_num = clean_po_val(row_gen.get("PO", ""))
                            row_gen["PO"] = po_num
                            if "Proyecto" in row_gen:
                                row_gen["Proyecto"] = clean_project_val(row_gen["Proyecto"])
                            
                            if not po_num or po_num == "NAN":
                                st.error("❌ El campo 'PO' en la hoja Datos_Generales no es válido.")
                            else:
                                # 2. Procesar Detalle de Entregas (Despivotar Matriz)
                                col_sku_name = df_det_matrix.columns[0]
                                df_det_matrix = df_det_matrix.rename(columns={col_sku_name: 'SKU'})
                                df_det_matrix['SKU'] = df_det_matrix['SKU'].astype(str).str.strip().str.upper()
                                
                                date_cols = df_det_matrix.columns[1:].tolist()
                                df_flat = df_det_matrix.melt(id_vars=['SKU'], value_vars=date_cols, var_name='Fecha_Entrega', value_name='Cantidad_Requerida')
                                
                                # Limpiar cantidades y filtrar vacíos
                                df_flat['Cantidad_Requerida'] = pd.to_numeric(df_flat['Cantidad_Requerida'], errors='coerce').fillna(0)
                                df_flat = df_flat[df_flat['Cantidad_Requerida'] > 0]
                                
                                # Normalizar Fechas
                                def normalize_date_col(col_name):
                                    c_clean = str(col_name).lower().strip()
                                    meses = {
                                        'ene': 'jan', 'feb': 'feb', 'mar': 'mar', 'abr': 'apr', 'may': 'may', 'jun': 'jun',
                                        'jul': 'jul', 'ago': 'aug', 'sep': 'sep', 'oct': 'oct', 'nov': 'nov', 'dic': 'dec'
                                    }
                                    for esp, eng in meses.items():
                                        if esp in c_clean:
                                            c_clean = c_clean.replace(esp, eng)
                                    try:
                                        return pd.to_datetime(c_clean).strftime('%Y-%m-%d')
                                    except:
                                        return str(col_name).strip()
                                        
                                df_flat['Fecha_Entrega'] = df_flat['Fecha_Entrega'].apply(normalize_date_col)
                                
                                # Asignar Parcialidades secuenciales cronológicamente
                                unique_dates = sorted(df_flat['Fecha_Entrega'].unique())
                                date_to_parcialidad = {date: f"P{idx+1}" for idx, date in enumerate(unique_dates)}
                                df_flat['Parcialidad'] = df_flat['Fecha_Entrega'].map(date_to_parcialidad)
                                df_flat['PO'] = po_num
                                
                                # Reordenar columnas
                                df_flat = df_flat[["PO", "SKU", "Fecha_Entrega", "Cantidad_Requerida", "Parcialidad"]]
                                
                                # Sincronizar bases locales de sesión
                                if modo_carga == "Sobrescribir base de datos completa":
                                    st.session_state.BD_POs_Cabecera = pd.DataFrame([row_gen])
                                    st.session_state.BD_Requerimientos_POs = df_flat
                                else:
                                    st.session_state.BD_POs_Cabecera = pd.concat([
                                        st.session_state.BD_POs_Cabecera[st.session_state.BD_POs_Cabecera['PO'] != po_num],
                                        pd.DataFrame([row_gen])
                                    ], ignore_index=True)
                                    
                                    st.session_state.BD_Requerimientos_POs = pd.concat([
                                        st.session_state.BD_Requerimientos_POs[st.session_state.BD_Requerimientos_POs['PO'] != po_num],
                                        df_flat
                                    ], ignore_index=True)
                                    
                                # Guardar cambios en GitHub
                                res_cab = subir_excel_a_github("BD_POs_Cabecera.xlsx", st.session_state.BD_POs_Cabecera)
                                res_req = subir_excel_a_github("BD_Requerimientos_POs.xlsx", st.session_state.BD_Requerimientos_POs)
                                
                                if res_cab and res_req:
                                    st.success(f"✅ Requerimientos de la PO {po_num} integrados y guardados en GitHub con éxito.")
                                else:
                                    st.error("❌ Ocurrió un error al intentar guardar los archivos en GitHub.")
                    else:
                        st.error("❌ El Excel cargado debe contener las hojas 'Datos_Generales' y 'Detalle_Entregas'.")
                except Exception as e:
                    st.error(f"❌ Error al procesar el archivo Excel: {e}")
                    
        st.write("---")
        st.subheader("📋 Órdenes de Compra Registradas")
        if not st.session_state.BD_POs_Cabecera.empty:
            st.dataframe(st.session_state.BD_POs_Cabecera, use_container_width=True, hide_index=True)
            st.info("💡 **Consejo de Navegación:** Para consultar el avance, dispersión de piezas y realizar el **Mapeo de Coincidencia** de estas POs con la producción real, diríjase a la pestaña **'📊 Matriz de Avance por PO'** en la parte superior.")
        else:
            st.info("💡 No hay POs registradas actualmente.")

    with tab_mat:
        st.write("")
        st.subheader("📊 Matriz de Avance y Dispersión por PO")
        
        if st.session_state.BD_POs_Cabecera.empty:
            st.info("ℹ️ No hay Órdenes de Compra registradas. Cargue una PO en la pestaña 'Ingreso de POs' para comenzar.")
        else:
            pos_disponibles = sorted(st.session_state.BD_POs_Cabecera['PO'].astype(str).str.strip().unique().tolist())
            pos_seleccionadas = st.multiselect(
                "Seleccione las POs a consultar (puede elegir una o varias):",
                options=pos_disponibles,
                default=[pos_disponibles[0]] if pos_disponibles else [],
                key="po_selector_matriz"
            )
            
            if not pos_seleccionadas:
                st.info("💡 Seleccione al menos una Orden de Compra (PO) en el selector superior para mostrar los datos de avance.")
            else:
                df_cab_selected = st.session_state.BD_POs_Cabecera[st.session_state.BD_POs_Cabecera['PO'].astype(str).str.strip().isin(pos_seleccionadas)].copy()
                
                # Obtener datos generales consolidados
                if len(pos_seleccionadas) == 1:
                    cab_info = df_cab_selected.iloc[0].to_dict()
                    st.markdown("##### 📄 Datos Generales")
                    mc1, mc2, mc3, mc4 = st.columns(4)
                    with mc1:
                        st.metric("Proyecto / Uso", cab_info.get("Proyecto", "N/A"))
                    with mc2:
                        fecha_val = cab_info.get("Fecha_Pedido", "N/A")
                        if pd.notnull(fecha_val) and hasattr(fecha_val, 'strftime'):
                            fecha_str = fecha_val.strftime("%Y-%m-%d")
                        else:
                            fecha_str = str(fecha_val)
                        st.metric("Fecha Pedido", fecha_str)
                    with mc3:
                        st.metric("Solicitante", cab_info.get("Solicitante", "N/A"))
                    with mc4:
                        st.metric("Requisición", cab_info.get("Requisicion", "N/A"))
                    st.info(f"📍 **Destino (L.A.B.):** {cab_info.get('Destino', 'N/A')}")
                else:
                    st.markdown("##### 📄 Datos Generales (Consolidado de POs)")
                    df_cab_display = df_cab_selected.copy()
                    if 'Fecha_Pedido' in df_cab_display.columns:
                        df_cab_display['Fecha_Pedido'] = df_cab_display['Fecha_Pedido'].apply(
                            lambda x: x.strftime("%Y-%m-%d") if pd.notnull(x) and hasattr(x, 'strftime') else str(x)
                        )
                    st.dataframe(df_cab_display, use_container_width=True, hide_index=True)
                    
                    # Generar diccionario para EML/fácil acceso
                    cab_info = {
                        "Proyecto": ", ".join(df_cab_selected["Proyecto"].astype(str).str.strip().unique()),
                        "Solicitante": ", ".join(df_cab_selected["Solicitante"].astype(str).str.strip().unique()),
                        "Requisicion": ", ".join(df_cab_selected["Requisicion"].astype(str).str.strip().unique()),
                        "Destino": ", ".join(df_cab_selected["Destino"].astype(str).str.strip().unique()),
                    }
                    
                # Obtener requerimientos de las POs
                df_req_po_raw = st.session_state.BD_Requerimientos_POs[st.session_state.BD_Requerimientos_POs['PO'].astype(str).str.strip().isin(pos_seleccionadas)].copy()
                
                if df_req_po_raw.empty:
                    st.warning("⚠️ No se encontraron partidas registradas para estas POs.")
                else:
                    # Obtener parcialidades únicas ordenadas cronológicamente por su fecha
                    df_dates_map = df_req_po_raw.groupby(['PO', 'Parcialidad'], as_index=False)['Fecha_Entrega'].min().sort_values(by='Fecha_Entrega')
                    lista_parcialidades = []
                    parcialidades_dict = {}
                    
                    for _, p_row in df_dates_map.iterrows():
                        po_name = str(p_row['PO']).strip()
                        p_val = str(p_row['Parcialidad']).strip()
                        f_val = p_row['Fecha_Entrega']
                        f_str = f_val.strftime("%Y-%m-%d") if pd.notnull(f_val) and hasattr(f_val, 'strftime') else str(f_val)
                        
                        op_key = f"{po_name}::{p_val}"
                        lista_parcialidades.append(op_key)
                        parcialidades_dict[op_key] = f"PO {po_name} - {p_val} ({f_str})"
                        
                    st.write("")
                    st.markdown("📅 **Filtro de Parcialidades a Revisar**")
                    parcialidades_selec = st.multiselect(
                        "Seleccione las parcialidades que desea incluir en la matriz y el cálculo de faltantes:",
                        options=lista_parcialidades,
                        default=lista_parcialidades,
                        format_func=lambda x: parcialidades_dict.get(x, x),
                        key=f"filter_parcialidades_multiselect_{'_'.join(pos_seleccionadas)}"
                    )
                    
                    df_req_po_raw['po_p_key'] = df_req_po_raw['PO'].astype(str).str.strip() + "::" + df_req_po_raw['Parcialidad'].astype(str).str.strip()
                    df_req_po = df_req_po_raw[df_req_po_raw['po_p_key'].isin(parcialidades_selec)].copy()
                    
                    if df_req_po.empty:
                        st.warning("⚠️ Seleccione al menos una parcialidad en el filtro superior para mostrar los datos de avance.")
                    else:
                        # Agrupar por SKU y Fecha de Entrega para evitar duplicados en la matriz
                        df_req_grouped = df_req_po.groupby(['SKU', 'Fecha_Entrega'], as_index=False)['Cantidad_Requerida'].sum()
                        
                        # Fechas de entrega únicas ordenadas cronológicamente
                        fechas_columnas = sorted(df_req_grouped['Fecha_Entrega'].unique().tolist())
                        unique_skus = df_req_grouped['SKU'].unique().tolist()
                        
                        # Obtener estatus de tarimas
                        mapa_tarimas_status = {}
                        if not st.session_state.BD_Tarimas.empty:
                            for _, t_row in st.session_state.BD_Tarimas.iterrows():
                                mapa_tarimas_status[str(t_row['ID_Tarima']).strip().upper()] = str(t_row['Estatus']).strip().upper()
                        
                        # Obtener detalles de producción para estas POs
                        df_prod_po = pd.DataFrame()
                        if not st.session_state.BD_Detalle_Tarimas.empty:
                            df_prod = st.session_state.BD_Detalle_Tarimas.copy()
                            
                            # Normalización automática inteligente para sugerencias
                            import re
                            def clean_po(val):
                                if not val or pd.isna(val): return ""
                                val_str = str(val).strip().upper()
                                # Quitar prefijo PO: o PO- y caracteres no alfanuméricos
                                val_str = re.sub(r'^PO\s*[-:]*\s*', '', val_str)
                                val_str = re.sub(r'[^A-Z0-9]', '', val_str)
                                return val_str
                            
                            po_targets = [clean_po(p) for p in pos_seleccionadas]
                            
                            # Obtener listado de PO reales registradas en las tarimas de producción
                            pos_reales_produccion = sorted(df_prod['PO'].astype(str).str.strip().unique().tolist())
                            
                            # Buscar coincidencias sugeridas (por ejemplo, "PO-2602-0711" coincide con "26020711")
                            sugeridos_coincidencia = [
                                p for p in pos_reales_produccion
                                if clean_po(p) in po_targets or any(target in clean_po(p) for target in po_targets)
                            ]
                            
                            # Mostrar la interfaz de mapeo interactivo para el usuario
                            st.write("---")
                            st.markdown("🔗 **Mapeo de Coincidencia de POs en Producción (Tarimas)**")
                            st.info("💡 **Sugerencia Automática:** El sistema ha preseleccionado los valores que parecen corresponder a las POs consultadas. Puede agregar o quitar etiquetas manualmente según sea necesario.")
                            
                            pos_mapeadas = st.multiselect(
                                "Etiquetas de PO físicas detectadas en producción:",
                                options=pos_reales_produccion,
                                default=sugeridos_coincidencia,
                                key=f"mapeo_po_multiselect_{'_'.join(pos_seleccionadas)}"
                            )
                        
                        # Filtrar los detalles de producción con base en la selección del usuario
                        df_prod_po = df_prod[df_prod['PO'].astype(str).str.strip().isin(pos_mapeadas)].copy()
                    
                    if not df_prod_po.empty:
                        df_prod_po['ID_Tarima_Clean'] = df_prod_po['ID_Tarima'].astype(str).str.strip().str.upper()
                        df_prod_po['Estatus'] = df_prod_po['ID_Tarima_Clean'].map(mapa_tarimas_status).fillna("DISPONIBLE")
                        df_prod_po['Cantidad'] = pd.to_numeric(df_prod_po['Cantidad'], errors='coerce').fillna(0)
                        df_prod_po['Remesada'] = df_prod_po.apply(lambda r: r['Cantidad'] if r['Estatus'] == 'REMESADA' else 0, axis=1)
                        df_prod_po['Stock'] = df_prod_po.apply(lambda r: r['Cantidad'] if r['Estatus'] != 'REMESADA' else 0, axis=1)
                    
                    # Construir matriz de dispersión FIFO
                    matrix_rows = []
                    import ast
                    
                    # Inicializar totalizadores por columna de fecha
                    col_req_totals = {d: 0 for d in fechas_columnas}
                    col_ent_totals = {d: 0 for d in fechas_columnas}
                    col_stk_totals = {d: 0 for d in fechas_columnas}
                    
                    for sku in unique_skus:
                        sku_desc = "Sin Registro en Catálogo"
                        if "BD_Articulos" in st.session_state and not st.session_state.BD_Articulos.empty:
                            match_art = st.session_state.BD_Articulos[st.session_state.BD_Articulos['SKU'] == sku]
                            if not match_art.empty:
                                sku_desc = match_art.iloc[0]['Nombre']
                                
                        # Requerimientos para este SKU específico
                        sku_reqs = df_req_grouped[df_req_grouped['SKU'] == sku].sort_values(by='Fecha_Entrega')
                        total_req = int(sku_reqs['Cantidad_Requerida'].sum())
                        
                        # Consolidado producido real
                        total_rem = 0
                        total_stk = 0
                        if not df_prod_po.empty:
                            match_prod = df_prod_po[df_prod_po['SKU'] == sku]
                            total_rem = int(match_prod['Remesada'].sum())
                            total_stk = int(match_prod['Stock'].sum())
                        
                        # --- ALGORITMO FIFO ---
                        # Pasada 1: Asignar Remesado (Entregado)
                        rem_restante = total_rem
                        allocated_rem = {}
                        for _, req_row in sku_reqs.iterrows():
                            d = req_row['Fecha_Entrega']
                            req_qty = req_row['Cantidad_Requerida']
                            alloc = min(req_qty, rem_restante)
                            allocated_rem[d] = alloc
                            rem_restante -= alloc
                            
                        # Pasada 2: Asignar Stock (Disponible en Almacén)
                        stk_restante = total_stk
                        allocated_stk = {}
                        for _, req_row in sku_reqs.iterrows():
                            d = req_row['Fecha_Entrega']
                            req_qty = req_row['Cantidad_Requerida']
                            already_alloc = allocated_rem.get(d, 0)
                            needed = req_qty - already_alloc
                            alloc = min(needed, stk_restante)
                            allocated_stk[d] = alloc
                            stk_restante -= alloc
                            
                        # Construir renglón
                        total_cubierto = total_rem + total_stk
                        total_faltante = max(0, total_req - total_cubierto)
                        
                        # Obtener detalle de tarimas para este SKU
                        tarimas_text_app = "-"
                        if not df_prod_po.empty:
                            sku_tarimas = df_prod_po[df_prod_po['SKU'] == sku]
                            tarimas_info_list = []
                            for _, t_row in sku_tarimas.iterrows():
                                t_id = str(t_row['ID_Tarima']).strip()
                                t_qty = int(t_row['Cantidad'])
                                t_status = str(t_row['Estatus']).strip().upper()
                                status_emoji = "✅" if t_status == "REMESADA" else "📦"
                                tarimas_info_list.append(f"{status_emoji} {t_id} ({t_qty} pzs)")
                            
                            if tarimas_info_list:
                                tarimas_text_app = "\n".join(tarimas_info_list)
                                
                        row_data = {
                            "SKU": sku,
                            "Detalle Tarimas": tarimas_text_app,
                            "Total Requerido": total_req,
                            "Total Entregado": total_rem,
                            "Total Almacén": total_stk,
                            "Total Faltante": total_faltante
                        }
                        
                        # Columnas dinámicas de Fechas
                        for d in fechas_columnas:
                            match_date = sku_reqs[sku_reqs['Fecha_Entrega'] == d]
                            if match_date.empty:
                                row_data[d] = "-"
                            else:
                                req_val = int(match_date.iloc[0]['Cantidad_Requerida'])
                                ent_val = int(allocated_rem.get(d, 0))
                                stk_val = int(allocated_stk.get(d, 0))
                                falt_val = max(0, req_val - (ent_val + stk_val))
                                
                                status_sym = "✅" if falt_val == 0 else "⚠️"
                                row_data[d] = f"R:{req_val} | E:{ent_val} | S:{stk_val} | F:{falt_val} {status_sym}"
                                
                                # Acumular para totales globales por columna
                                col_req_totals[d] += req_val
                                col_ent_totals[d] += ent_val
                                col_stk_totals[d] += stk_val
                                
                        matrix_rows.append(row_data)
                        
                    # Totales globales a nivel de métricas
                    tot_req = sum(r["Total Requerido"] for r in matrix_rows)
                    tot_ent = sum(r["Total Entregado"] for r in matrix_rows)
                    tot_stk = sum(r["Total Almacén"] for r in matrix_rows)
                    tot_fal = sum(r["Total Faltante"] for r in matrix_rows)
                    
                    # Construir fila resumen "📈 % AVANCE"
                    summary_row = {
                        "SKU": "📈 % AVANCE",
                        "Detalle Tarimas": "",
                        "Total Requerido": f"{((tot_ent + tot_stk) / tot_req * 100):.1f}%" if tot_req > 0 else "0.0%",
                        "Total Entregado": tot_ent,
                        "Total Almacén": tot_stk,
                        "Total Faltante": tot_fal
                    }
                    
                    # Para cada fecha, calculamos el porcentaje de avance global de esa fecha
                    for d in fechas_columnas:
                        r_tot = col_req_totals[d]
                        e_tot = col_ent_totals[d]
                        s_tot = col_stk_totals[d]
                        
                        if r_tot > 0:
                            pct = (e_tot + s_tot) / r_tot
                            pct_val = min(100.0, pct * 100)
                            
                            # Dibujar barra de progreso de texto (10 bloques)
                            num_blocks = int(round(pct_val / 10))
                            bar = "█" * num_blocks + "░" * (10 - num_blocks)
                            sym = "✅" if pct_val >= 100 else "⚠️"
                            summary_row[d] = f"{bar} {pct_val:.1f}% {sym}"
                        else:
                            summary_row[d] = "-"
                            
                    # Añadir fila resumen a los datos de la matriz
                    matrix_rows.append(summary_row)
                    
                    df_matrix = pd.DataFrame(matrix_rows)
                    
                    st.write("")
                    col_met1, col_met2, col_met3, col_met4 = st.columns(4)
                    with col_met1:
                        st.metric("Total Requerido", f"{tot_req:,} PZS")
                    with col_met2:
                        st.metric("Total Entregado (Remesado)", f"{tot_ent:,} PZS")
                    with col_met3:
                        st.metric("Stock en Almacén", f"{tot_stk:,} PZS")
                    with col_met4:
                        st.metric("Pendiente (Faltante)", f"{tot_fal:,} PZS")
                        
                    # Aplicar estilos CSS solicitados a columnas específicas y centrar alineación
                    def style_matrix(df):
                        styler = df.style
                        # Centrar alineación de todo el texto por defecto
                        styler = styler.set_properties(**{"text-align": "center"})
                        
                        # "SKU" Negrita y más grande (doble/grande)
                        if "SKU" in df.columns:
                            styler = styler.set_properties(
                                subset=["SKU"],
                                **{"font-weight": "bold", "text-align": "center", "font-size": "18px"}
                            )
                            
                        # "Detalle Tarimas" alineado a la izquierda y tamaño cómodo
                        if "Detalle Tarimas" in df.columns:
                            styler = styler.set_properties(
                                subset=["Detalle Tarimas"],
                                **{"text-align": "left", "font-size": "13px"}
                            )
                            
                        # "Total Requerido" Fondo Negro, Letra Blanca
                        styler = styler.set_properties(
                            subset=["Total Requerido"],
                            **{"background-color": "#000000", "color": "#FFFFFF", "font-weight": "bold", "text-align": "center", "font-size": "22px"}
                        )
                        # "Total Entregado" Fondo Verde, Letra Blanca
                        styler = styler.set_properties(
                            subset=["Total Entregado"],
                            **{"background-color": "#2E7D32", "color": "#FFFFFF", "font-weight": "bold", "text-align": "center", "font-size": "22px"}
                        )
                        # "Total Almacén" Fondo Amarillo, Letra Negra
                        styler = styler.set_properties(
                            subset=["Total Almacén"],
                            **{"background-color": "#FBC02D", "color": "#000000", "font-weight": "bold", "text-align": "center", "font-size": "22px"}
                        )
                        # "Total Faltante" Fondo Rojo, Letra Blanca
                        styler = styler.set_properties(
                            subset=["Total Faltante"],
                            **{"background-color": "#C62828", "color": "#FFFFFF", "font-weight": "bold", "text-align": "center", "font-size": "22px"}
                        )
                        return styler
                    
                    # Copiar para visualización y forzar tipo texto en números para permitir alineación centrada
                    df_display = df_matrix.copy()
                    cols_to_str = ["Total Requerido", "Total Entregado", "Total Almacén", "Total Faltante"]
                    for col_name in cols_to_str:
                        if col_name in df_display.columns:
                            df_display[col_name] = df_display[col_name].astype(str)
                            
                    styled_matrix_df = style_matrix(df_display)
                    st.dataframe(
                        styled_matrix_df,
                        use_container_width=True,
                        hide_index=True,
                        column_config={
                            "SKU": st.column_config.TextColumn(
                                "SKU",
                                width=120
                            ),
                            "Detalle Tarimas": st.column_config.TextColumn(
                                "Detalle Tarimas",
                                help="Detalle de tarimas asociadas y sus cantidades/estatus",
                                width=200
                            ),
                            "Total Requerido": st.column_config.TextColumn(
                                "Total Requerido",
                                width=110
                            ),
                            "Total Entregado": st.column_config.TextColumn(
                                "Total Entregado",
                                width=110
                            ),
                            "Total Almacén": st.column_config.TextColumn(
                                "Total Almacén",
                                width=110
                            ),
                            "Total Faltante": st.column_config.TextColumn(
                                "Total Faltante",
                                width=110
                            ),
                        }
                    )
                    
                    st.session_state.last_matrix_df = df_matrix
                    st.session_state.last_matrix_po = pos_seleccionadas
                    
                    po_names_str = "_".join(pos_seleccionadas)
                    po_names_display = ", ".join(pos_seleccionadas)
                    
                    # Generar el Excel del avance para descarga directa
                    buf_eml_xlsx = io.BytesIO()
                    with pd.ExcelWriter(buf_eml_xlsx, engine='openpyxl') as writer_dl:
                        sheet_title = f"Avance_POs_{po_names_str}"[:30] # Límite de 30 caracteres
                        df_matrix.to_excel(writer_dl, index=False, sheet_name=sheet_title)
                        
                        from openpyxl.styles import Font, PatternFill, Alignment
                        workbook = writer_dl.book
                        sheet = workbook[sheet_title]
                        
                        header_fill = PatternFill(start_color="EC2024", end_color="EC2024", fill_type="solid")
                        header_font = Font(color="FFFFFF", bold=True)
                        center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                        
                        # Cabeceras
                        for cell in sheet[1]:
                            cell.fill = header_fill
                            cell.font = header_font
                            cell.alignment = center_alignment
                            
                        # Celdas
                        for row in sheet.iter_rows(min_row=2):
                            for cell in row:
                                cell.alignment = center_alignment
                                
                        # Ajustar ancho
                        for col in sheet.columns:
                            max_length = 0
                            column = col[0].column_letter
                            for cell in col:
                                try:
                                    if len(str(cell.value)) > max_length:
                                        max_length = len(str(cell.value))
                                except:
                                    pass
                            sheet.column_dimensions[column].width = min((max_length + 2), 60)
                            
                        sheet.auto_filter.ref = sheet.dimensions
                        sheet.freeze_panes = "A2"
                    
                    buf_eml_xlsx.seek(0)
                    
                    # Botón de descarga directa de Excel
                    st.write("")
                    st.download_button(
                        label=f"📥 Descargar Reporte de Avance Excel (POs: {po_names_display})",
                        data=buf_eml_xlsx.getvalue(),
                        file_name=f"Reporte_Avance_POs_{po_names_str}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key=f"btn_dl_matrix_direct_xlsx_{po_names_str}",
                        use_container_width=True
                    )
                    
                    st.write("---")
                    st.subheader("📩 Generar Notificación por Correo (Borrador Outlook)")
                    st.caption("Ensamblado dinámico de correos con formato de borrador de Outlook, tablas HTML y el reporte Excel adjunto de esta consulta.")
                    
                    # Cargar destinatarios configurados por defecto
                    cfg_emails_po = obtener_emails_config()
                    
                    col_em_po1, col_em_po2 = st.columns(2)
                    with col_em_po1:
                        eml_po_to = st.text_input("Para:", value=cfg_emails_po.get("dest_to", ""), key=f"eml_po_to_{po_names_str}")
                    with col_em_po2:
                        eml_po_cc = st.text_input("CC:", value=cfg_emails_po.get("dest_cc", ""), key=f"eml_po_cc_{po_names_str}")
                    
                    # Generar cuerpo HTML
                    cuerpo_html = generar_cuerpo_correo_po_html(po_names_display, cab_info, df_matrix, fechas_columnas)
                    
                    # Generar adjuntos
                    adjuntos_dict = {
                        f"Reporte_Avance_POs_{po_names_str}.xlsx": buf_eml_xlsx.getvalue()
                    }
                    
                    # Generar borrador EML
                    eml_subject = f"Reporte de Avance y Dispersión - POs {po_names_display}"
                    eml_bytes = generar_archivo_eml(eml_po_to, eml_po_cc, eml_subject, cuerpo_html, adjuntos_dict)
                    
                    st.download_button(
                        label="📩 Descargar Borrador de Correo de Avance (.eml)",
                        data=eml_bytes,
                        file_name=f"Correo_Avance_POs_{po_names_str}.eml",
                        mime="message/rfc822",
                        key=f"btn_dl_eml_po_{po_names_str}",
                        use_container_width=True
                    )

    with tab_dl:
        st.write("")
        st.subheader("📉 Exportar Reporte de Dispersión")
        if "last_matrix_df" in st.session_state and "last_matrix_po" in st.session_state:
            po_name = st.session_state.last_matrix_po
            df_dl = st.session_state.last_matrix_df.copy()
            
            if isinstance(po_name, list):
                po_name_str = "_".join(po_name)
                po_name_display = ", ".join(po_name)
            else:
                po_name_str = str(po_name)
                po_name_display = str(po_name)
                
            buf_dl = io.BytesIO()
            with pd.ExcelWriter(buf_dl, engine='openpyxl') as writer_dl:
                sheet_title = f"Avance_{po_name_str}"[:30] # Límite de 30 caracteres
                df_dl.to_excel(writer_dl, index=False, sheet_name=sheet_title)
                
                from openpyxl.styles import Font, PatternFill, Alignment
                workbook = writer_dl.book
                sheet = workbook[sheet_title]
                
                header_fill = PatternFill(start_color="EC2024", end_color="EC2024", fill_type="solid")
                header_font = Font(color="FFFFFF", bold=True)
                center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                
                # Cabeceras
                for cell in sheet[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = center_alignment
                    
                # Celdas
                for row in sheet.iter_rows(min_row=2):
                    for cell in row:
                        cell.alignment = center_alignment
                        
                # Ajustar ancho
                for col in sheet.columns:
                    max_length = 0
                    column = col[0].column_letter
                    for cell in col:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    sheet.column_dimensions[column].width = min((max_length + 2), 60)
                    
                sheet.auto_filter.ref = sheet.dimensions
                sheet.freeze_panes = "A2"
                
            st.download_button(
                label=f"📥 Descargar Reporte de Avance Excel (POs: {po_name_display})",
                data=buf_dl.getvalue(),
                file_name=f"Reporte_Avance_POs_{po_name_str}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="btn_download_po_matrix_xlsx",
                use_container_width=True
            )
        else:
            st.info("💡 Por favor consulte primero una PO en la pestaña 'Matriz de Avance por PO' para poder descargar su reporte de dispersión.")
            
# =============================================================================
# 17. CONSULTA POR LOTE SKU
# =============================================================================
elif opcion_menu == "📋 Consulta por Lote SKU":
    st.title("📋 Consulta de Inventario por Lote SKU")
    st.markdown("Carga un archivo de Excel con un listado de SKUs para obtener un reporte dinámico con el desglose de piezas disponibles, remesadas y los totales enviados a cada receptor.")
    
    archivo_skus = st.file_uploader("📥 Selecciona el archivo Excel con los SKUs a consultar", type=['xlsx'])
    
    if archivo_skus:
        try:
            df_skus_input = pd.read_excel(archivo_skus)
            st.success("✅ Archivo cargado correctamente. Selecciona la columna del SKU:")
            
            columnas_excel = df_skus_input.columns.tolist()
            col_sku_sel = st.selectbox("Columna que contiene el SKU/Material:", columnas_excel, key="lote_sku_col_select")
            
            # 1. Limpiar y obtener SKUs únicos del Excel
            skus_consultados = df_skus_input[col_sku_sel].astype(str).str.strip().dropna().unique().tolist()
            skus_consultados = [s for s in skus_consultados if s and s.upper() not in ["TODOS", "SELECCIONE UN SKU..."]]
            
            if not skus_consultados:
                st.warning("⚠️ No se encontraron SKUs válidos en la columna seleccionada.")
            else:
                # 2. Obtener la producción detallada de estos SKUs (Base)
                if not st.session_state.BD_Detalle_Tarimas.empty:
                    df_det = st.session_state.BD_Detalle_Tarimas.copy()
                    df_det['SKU'] = df_det['SKU'].astype(str).str.strip()
                    df_det_filtrado = df_det[df_det['SKU'].isin(skus_consultados)].copy()
                else:
                    df_det_filtrado = pd.DataFrame(columns=["ID_Detalle", "ID_Tarima", "SKU", "PO", "Proyecto", "Parcialidad", "Descripcion", "Cantidad"])
                
                # --- FILTROS GLOBALES INTERACTIVOS ---
                st.markdown("##### 🔍 Filtrar resultados de la consulta:")
                col_f1, col_f2, col_f3 = st.columns(3)
                
                with col_f1:
                    opciones_proy = ["Todos"] + sorted(df_det_filtrado['Proyecto'].dropna().astype(str).unique().tolist()) if not df_det_filtrado.empty else ["Todos"]
                    filtro_proy = st.selectbox("Filtrar por Proyecto Interno:", opciones_proy, key="lote_filtro_proy_unique")
                
                with col_f2:
                    df_temp_po = df_det_filtrado.copy()
                    if filtro_proy != "Todos":
                        df_temp_po = df_temp_po[df_temp_po['Proyecto'] == filtro_proy]
                    opciones_po = ["Todos"] + sorted(df_temp_po['PO'].dropna().astype(str).unique().tolist()) if not df_temp_po.empty else ["Todos"]
                    filtro_po = st.selectbox("Filtrar por Orden de Compra (PO):", opciones_po, key="lote_filtro_po_unique")
                    
                with col_f3:
                    df_temp_desc = df_temp_po.copy()
                    if filtro_po != "Todos":
                        df_temp_desc = df_temp_desc[df_temp_desc['PO'] == filtro_po]
                    opciones_desc = ["Todas"] + sorted(df_temp_desc['Descripcion'].dropna().astype(str).unique().tolist()) if not df_temp_desc.empty else ["Todas"]
                    filtro_desc = st.selectbox("Filtrar por Descripción de Proyecto:", opciones_desc, key="lote_filtro_desc_unique")
                
                with st.spinner("Procesando consulta..."):
                    # Aplicar filtros a la producción base
                    df_corte = df_det_filtrado.copy()
                    if filtro_proy != "Todos":
                        df_corte = df_corte[df_corte['Proyecto'] == filtro_proy]
                    if filtro_po != "Todos":
                        df_corte = df_corte[df_corte['PO'] == filtro_po]
                    if filtro_desc != "Todas":
                        df_corte = df_corte[df_corte['Descripcion'] == filtro_desc]
                        
                    # 3. Cruzar con BD_Tarimas para obtener el Estatus actual de la tarima
                    if not df_corte.empty and not st.session_state.BD_Tarimas.empty:
                        df_t = st.session_state.BD_Tarimas[['ID_Tarima', 'Estatus']].copy()
                        df_t['ID_Tarima'] = df_t['ID_Tarima'].astype(str).str.strip()
                        df_corte['ID_Tarima'] = df_corte['ID_Tarima'].astype(str).str.strip()
                        df_det_cruce = pd.merge(df_corte, df_t, on='ID_Tarima', how='left')
                    else:
                        df_det_cruce = df_corte.copy()
                        df_det_cruce['Estatus'] = 'Disponible'
                            
                        # 4. Relacionar con Receptores desde BD_Datos_Generales_Remision
                        mapa_tarima_receptor = {}
                        if "BD_Datos_Generales_Remision" in st.session_state and not st.session_state.BD_Datos_Generales_Remision.empty:
                            import ast
                            for _, r_row in st.session_state.BD_Datos_Generales_Remision.iterrows():
                                rec = r_row['Nombre_Receptor']
                                tarimas_str = r_row['Tarimas_Asociadas']
                                try:
                                    t_list = ast.literal_eval(tarimas_str)
                                except:
                                    t_list = [tarimas_str]
                                for t in t_list:
                                    mapa_tarima_receptor[str(t).strip()] = str(rec).strip()
                                    
                        df_det_cruce['Receptor'] = df_det_cruce['ID_Tarima'].map(mapa_tarima_receptor).fillna("N/A")
                        
                        # 5. Segmentar Cantidades (Disponible vs Remesado)
                        df_det_cruce['Cantidad'] = pd.to_numeric(df_det_cruce['Cantidad'], errors='coerce').fillna(0)
                        df_det_cruce['Cant_Disponible'] = df_det_cruce.apply(lambda r: r['Cantidad'] if str(r['Estatus']).strip() != 'Remesada' else 0, axis=1)
                        df_det_cruce['Cant_Remesada'] = df_det_cruce.apply(lambda r: r['Cantidad'] if str(r['Estatus']).strip() == 'Remesada' else 0, axis=1)
                        
                        # Columnas dinámicas por Receptor (solo de las tarimas que han sido remesadas)
                        receptores_remesados = df_det_cruce[df_det_cruce['Estatus'] == 'Remesada']['Receptor'].dropna().unique().tolist()
                        receptores_remesados = [r for r in receptores_remesados if r != "N/A"]
                        
                        for rec in receptores_remesados:
                            df_det_cruce[f"Remesado a: {rec}"] = df_det_cruce.apply(lambda r: r['Cantidad'] if str(r['Estatus']).strip() == 'Remesada' and r['Receptor'] == rec else 0, axis=1)
                            
                        # 6. Agrupar por SKU
                        agg_dict = {
                            'Cant_Disponible': 'sum',
                            'Cant_Remesada': 'sum',
                            'Cantidad': 'sum'
                        }
                        for rec in receptores_remesados:
                            agg_dict[f"Remesado a: {rec}"] = 'sum'
                            
                        if not df_det_cruce.empty:
                            df_grouped = df_det_cruce.groupby('SKU').agg(agg_dict).reset_index()
                            df_grouped = df_grouped.rename(columns={
                                'Cant_Disponible': 'Disponible (No Remesado)',
                                'Cant_Remesada': 'Total Remesado',
                                'Cantidad': 'Total Fabricado'
                            })
                        else:
                            df_grouped = pd.DataFrame(columns=['SKU', 'Disponible (No Remesado)', 'Total Remesado', 'Total Fabricado'])
                            
                        # 7. Asegurar que todos los SKUs de la lista cargada se muestren (incluso si tienen 0 en inventario)
                        df_base_skus = pd.DataFrame({'SKU': skus_consultados})
                        df_final = pd.merge(df_base_skus, df_grouped, on='SKU', how='left')
                        
                        # Llenar nulos con 0
                        columnas_numericas = ['Disponible (No Remesado)', 'Total Remesado', 'Total Fabricado'] + [f"Remesado a: {r}" for r in receptores_remesados]
                        for col in columnas_numericas:
                            if col in df_final.columns:
                                df_final[col] = df_final[col].fillna(0).astype(int)
                                
                        # Enriquecer con Nombre / Descripción
                        if "BD_Articulos" in st.session_state and not st.session_state.BD_Articulos.empty:
                            df_art = st.session_state.BD_Articulos[['SKU', 'Nombre']].copy()
                            df_art['SKU'] = df_art['SKU'].astype(str).str.strip()
                            df_final = pd.merge(df_final, df_art, on='SKU', how='left')
                            df_final['Nombre'] = df_final['Nombre'].fillna("Sin Registro en Catálogo")
                            
                            # Reordenar columnas para poner Nombre después de SKU
                            cols_orden = ['SKU', 'Nombre', 'Disponible (No Remesado)', 'Total Remesado'] + [f"Remesado a: {r}" for r in receptores_remesados] + ['Total Fabricado']
                            # Filtrar solo las que existen por seguridad
                            cols_orden = [c for c in cols_orden if c in df_final.columns]
                            df_final = df_final[cols_orden]
                            
                        # 8. Agregar Fila de Totales Generales al final
                        total_row = {}
                        for col in df_final.columns:
                            if col == 'SKU':
                                total_row['SKU'] = 'TOTALES'
                            elif col == 'Nombre':
                                total_row['Nombre'] = ''
                            else:
                                total_row[col] = df_final[col].sum()
                                
                        df_final = pd.concat([df_final, pd.DataFrame([total_row])], ignore_index=True)
                        
                        # --- PANEL DE VISUALIZACIÓN ---
                        st.markdown("---")
                        st.subheader("📊 Consolidado de Inventario por Lote")
                        
                        col_m1, col_m2, col_m3 = st.columns(3)
                        with col_m1:
                            st.metric("Total Disponible (En Patio)", f"{total_row.get('Disponible (No Remesado)', 0):,} PZS")
                        with col_m2:
                            st.metric("Total Remesado (Enviado)", f"{total_row.get('Total Remesado', 0):,} PZS")
                        with col_m3:
                            st.metric("Total Fabricado Consolidado", f"{total_row.get('Total Fabricado', 0):,} PZS")
                            
                        st.dataframe(df_final, use_container_width=True, hide_index=True)
                        
                        # --- MOTOR DE EXPORTACIÓN A EXCEL FORMATEADO ---
                        buf_c = io.BytesIO()
                        writer_c = pd.ExcelWriter(buf_c, engine='openpyxl')
                        df_final.to_excel(writer_c, index=False, sheet_name='Reporte por Lote')
                        
                        from openpyxl.styles import Font, PatternFill, Alignment
                        workbook = writer_c.book
                        sheet = workbook['Reporte por Lote']
                        
                        header_fill = PatternFill(start_color="D32F2F", end_color="D32F2F", fill_type="solid")
                        header_font = Font(color="FFFFFF", bold=True)
                        center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                        
                        # Dar formato al encabezado
                        for cell in sheet[1]:
                            cell.fill = header_fill
                            cell.font = header_font
                            cell.alignment = center_alignment
                            
                        # Formatear celdas de datos
                        for row in sheet.iter_rows(min_row=2):
                            for cell in row:
                                cell.alignment = center_alignment
                                
                        # Auto-ajuste de columnas
                        for col in sheet.columns:
                            max_length = 0
                            column = col[0].column_letter
                            for cell in col:
                                try:
                                    if len(str(cell.value)) > max_length:
                                        max_length = len(str(cell.value))
                                except:
                                    pass
                            sheet.column_dimensions[column].width = min((max_length + 2), 60)
                            
                        # Filtros e inmovilización
                        sheet.auto_filter.ref = sheet.dimensions
                        sheet.freeze_panes = "A2"
                        
                        writer_c.close()
                        buf_c.seek(0)
                        
                        st.download_button(
                            label="📥 Descargar Reporte de Lote en Excel (.xlsx)",
                            data=buf_c.getvalue(),
                            file_name="Reporte_Consolidado_Lote.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            key="btn_download_reporte_lote_excel_v3"
                        )
                        
        except Exception as e:
            st.error(f"❌ Ocurrió un error al procesar la consulta por lote: {e}")

    
