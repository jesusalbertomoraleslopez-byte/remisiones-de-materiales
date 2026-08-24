import streamlit as st
import streamlit.components.v1 as components
import os
import pandas as pd
import datetime
import io
import requests
import urllib.parse
from PIL import Image
import glob

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders

from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image as RLImage
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch

# 1. CONFIGURACIÓN E INTERFAZ BASE RESPONSIVA (Base de Datos SGP - SIGRAMA Metales)
st.set_page_config(
    page_title="Base de Datos SGP — Consulta de Piezas SIGRAMA",
    layout="wide",
    page_icon="🏭"
)

REPO_OWNER = "jesusalbertomoraleslopez-byte"
REPO_NAME = "remisiones-de-materiales"
BRANCH = "main"

def obtener_secret(key, default=None):
    try:
        if hasattr(st, 'secrets') and st.secrets is not None:
            if key in st.secrets:
                val = st.secrets[key]
                return val if val is not None else default
    except Exception:
        pass
    return default

# Estilos CSS Inspirados en la Interfaz Oficial SGP SIGRAMA
st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=Montserrat:wght@400;600;700;800&family=Questrial&display=swap');

    html, body, [class*="css"], .stApp {
        font-family: 'Questrial', sans-serif !important;
        background-color: #F1F5F9 !important;
    }

    /* Reducir espacio superior vacio del contenedor principal de Streamlit */
    .block-container, [data-testid="stBlockContainer"] {
        padding-top: 0.8rem !important;
        padding-bottom: 1rem !important;
        margin-top: 0 !important;
    }
    header[data-testid="stHeader"] {
        height: 0 !important;
        background: transparent !important;
        display: none !important;
    }

    /* Barra Superior SGP Compacta en Rojo Corporativo SIGRAMA #EC2024 y Negro #111111 */
    .sgp-header {
        background: linear-gradient(90deg, #A81B1E 0%, #EC2024 45%, #111111 100%) !important;
        padding: 8px 16px !important;
        border-radius: 6px !important;
        color: #FFFFFF !important;
        margin-bottom: 8px !important;
        box-shadow: 0 2px 6px rgba(0,0,0,0.12);
    }
    .sgp-header h2 {
        color: #FFFFFF !important;
        font-family: 'Montserrat', sans-serif !important;
        font-weight: 800 !important;
        letter-spacing: 0.5px !important;
        margin: 0 !important;
        font-size: 17px !important;
    }
    .sgp-header p {
        color: #F8FAFC !important;
        margin: 1px 0 0 0 !important;
        font-size: 11px !important;
    }

    /* Tarjeta de Ficha Técnica SGP Estilo Tablero Industrial */
    .sgp-card {
        background-color: #111111 !important;
        border: 2px solid #334155 !important;
        border-radius: 8px !important;
        padding: 20px !important;
        color: #FFFFFF !important;
    }
    .sgp-sku-title {
        background-color: #FFFFFF !important;
        color: #111111 !important;
        font-family: 'Montserrat', sans-serif !important;
        font-size: 32px !important;
        font-weight: 800 !important;
        text-align: center !important;
        padding: 10px !important;
        border-radius: 6px !important;
        border: 2px solid #EC2024 !important;
        margin-bottom: 15px !important;
        letter-spacing: 2px !important;
    }
    .sgp-label {
        font-family: 'Montserrat', sans-serif !important;
        font-weight: 700 !important;
        color: #94A3B8 !important;
        font-size: 12px !important;
        text-transform: uppercase !important;
    }
    .sgp-value {
        font-family: 'Questrial', sans-serif !important;
        font-size: 16px !important;
        color: #FFFFFF !important;
        font-weight: bold !important;
    }

    /* Restricción y Escalado a 0.8 veces para la Imagen o Plano de Pieza */
    .stImage > img, [data-testid="stImage"] img {
        max-height: 256px !important;
        max-width: 80% !important;
        width: auto !important;
        object-fit: contain !important;
        margin: 0 auto !important;
        display: block !important;
        border-radius: 8px !important;
        border: 1px solid #CBD5E1 !important;
        background-color: #FFFFFF !important;
        padding: 5px !important;
        box-shadow: 0 2px 8px rgba(0,0,0,0.08) !important;
    }
</style>
""", unsafe_allow_html=True)

# 2. CARGA DE DATOS DESDE GITHUB CON CACHÉ
@st.cache_data(ttl=120, max_entries=15)
def cargar_excel_desde_github(file_name):
    try:
        url = f"https://api.github.com/repos/{REPO_OWNER}/{REPO_NAME}/contents/{file_name}?ref={BRANCH}&ts={int(__import__('time').time())}"
        headers = {"Cache-Control": "no-cache", "Accept": "application/vnd.github.v3.raw"}
        token = obtener_secret("github_token")
        if token:
            headers["Authorization"] = f"token {token}"
            
        res = requests.get(url, headers=headers)
        if res.status_code == 200:
            archivo_bytes = res.content
            try:
                return pd.read_excel(io.BytesIO(archivo_bytes), sheet_name='Datos_Sistema')
            except Exception:
                return pd.read_excel(io.BytesIO(archivo_bytes), sheet_name=0)
    except Exception:
        pass

    if os.path.exists(file_name):
        try:
            try:
                return pd.read_excel(file_name, sheet_name='Datos_Sistema')
            except Exception:
                return pd.read_excel(file_name, sheet_name=0)
        except Exception:
            pass
    return pd.DataFrame()

# Helper de estandarización unificada de fechas a formato DD/MM/YYYY
def normalizar_fecha_display(val):
    """Convierte cualquier formato de fecha (ISO YYYY-MM-DD, DD/MM/YYYY, Timestamp) al formato unificado DD/MM/YYYY."""
    if pd.isna(val) or val is None:
        return "N/A"
    val_str = str(val).strip()
    if not val_str or val_str.upper() in ['N/A', 'NONE', 'NAN', 'S/N']:
        return "N/A"
    if ' ' in val_str:
        val_str = val_str.split()[0]
    try:
        # Formato ISO YYYY-MM-DD -> DD/MM/YYYY
        if len(val_str) == 10 and val_str[4] == '-' and val_str[7] == '-':
            parts = val_str.split('-')
            return f"{parts[2]}/{parts[1]}/{parts[0]}"
        dt = pd.to_datetime(val_str, errors='coerce', dayfirst=True)
        if pd.notna(dt):
            return dt.strftime('%d/%m/%Y')
    except Exception:
        pass
    return val_str

# Helper para la generación de archivos EML con múltiples adjuntos e imágenes inline (CID)
def generar_archivo_eml(dest_to, dest_cc, subject, body_html, adjuntos_dict, inline_images_dict=None):
    """Genera un archivo EML en memoria como borrador de Outlook con adjuntos e imágenes inline (CID)."""
    msg = MIMEMultipart('related')
    msg['Subject'] = subject
    if dest_to and str(dest_to).strip():
        msg['To'] = str(dest_to).strip()
    if dest_cc and str(dest_cc).strip():
        msg['Cc'] = str(dest_cc).strip()
    msg['X-Unsent'] = '1'

    msg_alt = MIMEMultipart('alternative')
    msg.attach(msg_alt)
    
    body_part = MIMEText(body_html, 'html', 'utf-8')
    msg_alt.attach(body_part)
    
    # Imágenes Inline (CID)
    if inline_images_dict:
        from email.mime.image import MIMEImage
        for cid, img_bytes in inline_images_dict.items():
            if img_bytes:
                img_part = MIMEImage(img_bytes)
                img_part.add_header('Content-ID', f'<{cid}>')
                img_part.add_header('Content-Disposition', 'inline', filename=f"{cid}.png")
                msg.attach(img_part)

    # Adjuntos de Archivos (.xlsx, .pdf)
    for filename, file_bytes in adjuntos_dict.items():
        if not file_bytes: continue
        if hasattr(file_bytes, 'getvalue'): file_bytes = file_bytes.getvalue()
        elif hasattr(file_bytes, 'read'): file_bytes = file_bytes.read()
            
        part = MIMEBase('application', 'octet-stream')
        part.set_payload(file_bytes)
        encoders.encode_base64(part)
        part.add_header('Content-Disposition', f'attachment; filename="{filename}"')
        msg.attach(part)
        
    import gc
    val = msg.as_bytes()
    gc.collect()
    return val

def generar_excel_consulta_inventario(sku_actual, spec_dict, pzs_disp, pzs_rem, pzs_tot, tar_tot, df_tabla_export):
    """Genera un archivo Excel profesional estilizado con openpyxl (Ficha Técnica, Métricas y Tabla Auto-ajustada)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Inventario_{sku_actual}"[:30]
    ws.views.sheetView[0].showGridLines = True
    
    fill_header = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    fill_accent = PatternFill(start_color="EC2024", end_color="EC2024", fill_type="solid")
    fill_kpi = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
    fill_zebra = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    fill_white = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    
    font_title = Font(name="Arial", size=13, bold=True, color="FFFFFF")
    font_sub = Font(name="Arial", size=9, italic=True, color="E2E8F0")
    font_hdr = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    font_bold = Font(name="Arial", size=9.5, bold=True, color="0F172A")
    font_regular = Font(name="Arial", size=9.5, color="1E293B")
    font_kpi_num = Font(name="Arial", size=13, bold=True, color="EC2024")
    font_kpi_lbl = Font(name="Arial", size=8.5, bold=True, color="475569")

    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )

    # 1. Cabecera Principal (Filas 1 a 2)
    max_cols = max(len(df_tabla_export.columns), 9)
    col_max_letter = get_column_letter(max_cols)
    
    ws.merge_cells(f"A1:{col_max_letter}1")
    ws["A1"] = f"REPORTE DE INVENTARIO Y UBICACIÓN FÍSICA DE PIEZA — {sku_actual}"
    ws["A1"].font = font_title
    ws["A1"].fill = fill_header
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    ws.merge_cells(f"A2:{col_max_letter}2")
    f_hoy = datetime.datetime.now().strftime("%d/%m/%Y %H:%M")
    ws["A2"] = f"INDUSTRIA SIGRAMA S.A. DE C.V. | Planta Metales Diagonal | Generado: {f_hoy}"
    ws["A2"].font = font_sub
    ws["A2"].fill = fill_header
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 16

    # Línea Roja de Acento (Fila 3)
    ws.merge_cells(f"A3:{col_max_letter}3")
    ws["A3"].fill = fill_accent
    ws.row_dimensions[3].height = 3.5

    # 2. Ficha Técnica del Artículo (Filas 5 y 6)
    ws["A5"] = "DESCRIPCIÓN COMERCIAL:"
    ws["A5"].font = font_bold
    ws["B5"] = spec_dict.get('nombre', 'N/A')
    ws["B5"].font = font_regular

    ws["E5"] = "CALIBRE / ESPESOR:"
    ws["E5"].font = font_bold
    ws["F5"] = spec_dict.get('calibre', 'N/A')
    ws["F5"].font = font_regular

    ws["A6"] = "DIMENSIONES PIEZA:"
    ws["A6"].font = font_bold
    ws["B6"] = spec_dict.get('dims', 'N/A')
    ws["B6"].font = font_regular

    ws["E6"] = "MATERIAL / ACABADO:"
    ws["E6"].font = font_bold
    ws["F6"] = spec_dict.get('acabado', 'N/A')
    ws["F6"].font = font_regular

    ws.row_dimensions[5].height = 18
    ws.row_dimensions[6].height = 18

    # 3. Bloque de Métricas KPI (Filas 8 y 9)
    kpis = [
        ("DISPONIBLES PLANTA", pzs_disp, "A", "B"),
        ("REMESADAS (ENVIADAS)", pzs_rem, "C", "D"),
        ("TOTAL PIEZAS", pzs_tot, "E", "F"),
        ("TARIMAS FÍSICAS", tar_tot, "G", col_max_letter)
    ]
    for lbl, val, c1, c2 in kpis:
        cell_lbl = f"{c1}8"
        cell_val = f"{c1}9"
        if c1 != c2:
            ws.merge_cells(f"{c1}8:{c2}8")
            ws.merge_cells(f"{c1}9:{c2}9")
        ws[cell_lbl] = lbl
        ws[cell_lbl].font = font_kpi_lbl
        ws[cell_lbl].fill = fill_kpi
        ws[cell_lbl].alignment = Alignment(horizontal="center", vertical="center")

        ws[cell_val] = val
        ws[cell_val].font = font_kpi_num
        ws[cell_val].fill = fill_kpi
        ws[cell_val].alignment = Alignment(horizontal="center", vertical="center")
        if isinstance(val, (int, float)):
            ws[cell_val].number_format = '#,##0'

    ws.row_dimensions[8].height = 15
    ws.row_dimensions[9].height = 22

    # 4. Tabla de Datos (Fila 11)
    start_row = 11
    headers = list(df_tabla_export.columns)
    for col_num, h_text in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col_num, value=h_text)
        cell.font = font_hdr
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
    ws.row_dimensions[start_row].height = 24

    # Filas de datos
    current_row = start_row + 1
    for r_idx, row_data in df_tabla_export.iterrows():
        fill_row = fill_zebra if r_idx % 2 == 1 else fill_white
        for c_idx, val in enumerate(row_data, 1):
            col_name = headers[c_idx - 1]
            cell = ws.cell(row=current_row, column=c_idx, value=val)
            cell.font = font_regular
            cell.fill = fill_row
            cell.border = thin_border

            if "ID Tarima" in col_name or "Fecha" in col_name or "Estatus Tarima" in col_name:
                cell.alignment = Alignment(horizontal="center", vertical="center")
                if "ID Tarima" in col_name or "Estatus Tarima" in col_name:
                    cell.font = font_bold
            elif "Piezas" in col_name or isinstance(val, (int, float)):
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.number_format = '#,##0'
                cell.font = font_bold
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

        ws.row_dimensions[current_row].height = 19
        current_row += 1

    # Auto-fit dinámico de ancho de columnas
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.row in [1, 2, 3]: continue
            val_str = str(cell.value or '')
            if len(val_str) > max_len: max_len = len(val_str)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 13)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()

def generar_html_correo_sku(sku, spec_dict, pzs_disp, pzs_rem, pzs_tot, tar_tot, df_tabla, has_logo_cid=False, has_img_cid=False):
    """Genera el cuerpo HTML corporativo réplica exacta del reporte PDF oficial."""
    f_hoy = datetime.datetime.now().strftime("%d/%m/%Y")
    
    logo_cell = '<div style="font-size: 22px; font-weight: 800; color: #EC2024; font-family: sans-serif;">SIGRAMA</div>'
    if has_logo_cid:
        logo_cell = '<img src="cid:logo_sigrama_cid" style="width: 140px; height: auto;" alt="SIGRAMA">'

    img_html_cell = f'<div style="border: 1px dashed #CBD5E1; border-radius: 6px; padding: 25px 10px; color: #94A3B8; font-size: 11px; background-color: #FFFFFF;">📷 Sin fotografía o plano registrado</div>'
    if has_img_cid:
        img_html_cell = f'<img src="cid:foto_sku_cid" style="max-height: 175px; max-width: 100%; width: auto; height: auto; object-fit: contain; border-radius: 4px; border: 1px solid #CBD5E1; padding: 4px; background-color: #FFFFFF;">'

    filas_html = ""
    if not df_tabla.empty:
        for idx, r in df_tabla.iterrows():
            bg = "#ffffff" if idx % 2 == 0 else "#f8fafc"
            filas_html += f"""
            <tr style="background-color: {bg};">
                <td style="padding: 5px 6px; border: 1px solid #cbd5e1; font-size: 10px; font-weight: bold; text-align: center;">{r.get('ID Tarima (TPM)', 'N/A')}</td>
                <td style="padding: 5px 6px; border: 1px solid #cbd5e1; font-size: 10px; text-align: center;">{r.get('Fecha Empaque', 'N/A')}</td>
                <td style="padding: 5px 6px; border: 1px solid #cbd5e1; font-size: 10px;">{r.get('Líder Empaque', 'N/A')}</td>
                <td style="padding: 5px 6px; border: 1px solid #cbd5e1; font-size: 10px;">{r.get('Ubicación Actual', 'N/A')}</td>
                <td style="padding: 5px 6px; border: 1px solid #cbd5e1; font-size: 10px; text-align: center; font-weight: bold;">{r.get('Estatus Tarima', 'N/A')}</td>
                <td style="padding: 5px 6px; border: 1px solid #cbd5e1; font-size: 10px; text-align: center; font-weight: bold;">{r.get('Piezas', 0)}</td>
                <td style="padding: 5px 6px; border: 1px solid #cbd5e1; font-size: 10px;">{r.get('Proyecto', 'N/A')}</td>
                <td style="padding: 5px 6px; border: 1px solid #cbd5e1; font-size: 10px;">{r.get('PO', 'N/A')}</td>
                <td style="padding: 5px 6px; border: 1px solid #cbd5e1; font-size: 10px;">{r.get('Parcialidad', 'N/A')}</td>
                <td style="padding: 5px 6px; border: 1px solid #cbd5e1; font-size: 10px;">{r.get('Descripción Proyecto', 'N/A')}</td>
                <td style="padding: 5px 6px; border: 1px solid #cbd5e1; font-size: 10px;">{r.get('Estatus de Remisión / Destino', 'En Planta')}</td>
            </tr>
            """

    html = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="utf-8">
        <style>
            body {{ font-family: 'Segoe UI', Helvetica, Arial, sans-serif; color: #1E293B; background-color: #FFFFFF; margin: 0; padding: 15px; }}
            .pdf-container {{ max-width: 920px; margin: 0 auto; }}
            .header-table {{ width: 100%; border-collapse: collapse; margin-bottom: 6px; }}
            .header-title {{ font-size: 16px; font-weight: 800; color: #0F172A; margin: 0; }}
            .header-sub {{ font-size: 11.5px; font-weight: 800; color: #EC2024; margin: 2px 0 0 0; }}
            .header-meta {{ font-size: 10px; color: #64748B; margin-top: 3px; text-transform: uppercase; font-weight: 600; }}
            .red-line {{ height: 2px; background-color: #EC2024; margin-bottom: 12px; border: none; }}
            
            .card-spec {{ width: 100%; border: 1px solid #CBD5E1; border-radius: 6px; background-color: #F8FAFC; border-collapse: collapse; margin-bottom: 12px; }}
            .spec-lbl {{ font-size: 10px; font-weight: bold; color: #64748B; text-transform: uppercase; padding: 3px 6px; width: 170px; }}
            .spec-val {{ font-size: 11px; font-weight: bold; color: #0F172A; padding: 3px 6px; }}
            
            .metric-table {{ width: 100%; border-collapse: separate; border-spacing: 6px 0; margin-bottom: 12px; }}
            .metric-cell {{ background-color: #F8FAFC; border: 1px solid #CBD5E1; border-radius: 6px; padding: 6px 4px; text-align: center; width: 25%; }}
            .metric-num {{ font-size: 15px; font-weight: bold; color: #EC2024; }}
            .metric-lbl {{ font-size: 9.5px; font-weight: bold; color: #64748B; margin-top: 1px; }}
            
            .data-table {{ width: 100%; border-collapse: collapse; margin-bottom: 15px; border: 1px solid #CBD5E1; }}
            .data-table th {{ background-color: #1E293B; color: #FFFFFF; font-size: 10px; font-weight: bold; padding: 6px 4px; border: 1px solid #CBD5E1; text-align: center; }}
            .data-table td {{ font-size: 10px; padding: 5px 6px; border: 1px solid #CBD5E1; color: #1E293B; }}
            
            .portal-box {{ background-color: #F8FAFC; border: 1px solid #CBD5E1; border-left: 4px solid #EC2024; padding: 10px 14px; border-radius: 4px; margin: 15px 0; }}
        </style>
    </head>
    <body>
    <div class="pdf-container">

        <!-- 1. ENCABEZADO ESTILO REPORTE PDF OFICIAL -->
        <table class="header-table">
            <tr>
                <td style="width: 150px; vertical-align: middle;">
                    {logo_cell}
                </td>
                <td style="vertical-align: middle; padding-left: 15px;">
                    <div class="header-title">INDUSTRIA SIGRAMA S.A. DE C.V.</div>
                    <div class="header-sub">BASE DE DATOS SGP — PLANTA METALES DIAGONAL</div>
                    <div class="header-meta">REPORTE OFICIAL DE PIEZA | BÚSQUEDA: {sku} | FECHA DE EMISIÓN: {f_hoy}</div>
                </td>
            </tr>
        </table>
        <div class="red-line"></div>

        <!-- 2. FICHA TÉCNICA Y DIBUJO/PLANO EN TARJETA ELEGANTE (REPLICA EXACTA DEL PDF) -->
        <table class="card-spec">
            <tr>
                <td style="vertical-align: top; padding: 8px; width: 68%;">
                    <table style="width: 100%; border-collapse: collapse;">
                        <tr>
                            <td class="spec-lbl">NÚMERO DE PARTE / PLANO:</td>
                            <td class="spec-val" style="color: #EC2024; font-size: 12px;">{sku}</td>
                        </tr>
                        <tr>
                            <td class="spec-lbl">DESCRIPCIÓN COMERCIAL:</td>
                            <td class="spec-val">{spec_dict.get('nombre', 'N/A')}</td>
                        </tr>
                        <tr>
                            <td class="spec-lbl">CALIBRE / ESPESOR:</td>
                            <td class="spec-val">{spec_dict.get('calibre', 'N/A')}</td>
                        </tr>
                        <tr>
                            <td class="spec-lbl">DIMENSIONES PIEZA:</td>
                            <td class="spec-val">{spec_dict.get('dims', 'N/A')}</td>
                        </tr>
                        <tr>
                            <td class="spec-lbl">MATERIAL / ACABADO:</td>
                            <td class="spec-val">{spec_dict.get('acabado', 'N/A')}</td>
                        </tr>
                    </table>
                </td>
                <td style="vertical-align: middle; text-align: center; padding: 8px; width: 32%;">
                    {img_html_cell}
                </td>
            </tr>
        </table>

        <!-- 3. BLOQUE DE TARJETAS DE MÉTRICAS -->
        <table class="metric-table">
            <tr>
                <td class="metric-cell">
                    <div class="metric-num">{pzs_disp:,} PZS</div>
                    <div class="metric-lbl">Disponibles Planta</div>
                </td>
                <td class="metric-cell">
                    <div class="metric-num">{pzs_rem:,} PZS</div>
                    <div class="metric-lbl">Remesadas (Enviadas)</div>
                </td>
                <td class="metric-cell">
                    <div class="metric-num">{pzs_tot:,} PZS</div>
                    <div class="metric-lbl">Total Piezas</div>
                </td>
                <td class="metric-cell">
                    <div class="metric-num">{tar_tot}</div>
                    <div class="metric-lbl">Tarimas Físicas</div>
                </td>
            </tr>
        </table>

        <!-- 4. TABLA DE DESGLOSE HISTÓRICO DE TARIMAS -->
        <table class="data-table">
            <thead>
                <tr>
                    <th>ID Tarima (TPM)</th>
                    <th>Fecha Empaque</th>
                    <th>Líder Empaque</th>
                    <th>Ubicación Actual</th>
                    <th>Estatus Tarima</th>
                    <th>Piezas</th>
                    <th>Proyecto</th>
                    <th>PO</th>
                    <th>Parcialidad</th>
                    <th>Descripción Proyecto</th>
                    <th>Estatus de Remisión / Destino</th>
                </tr>
            </thead>
            <tbody>
                {filas_html}
            </tbody>
        </table>

        <!-- 5. ENLACE AL PORTAL -->
        <div class="portal-box">
            <p style="margin: 0; font-size: 11px; color: #1E293B; font-weight: bold;">
                🔍 <b>Portal de Consulta de Inventario en Tiempo Real:</b>
            </p>
            <p style="margin: 3px 0 0 0; font-size: 10.5px; color: #334155;">
                Para consultar piezas, planos y disponibilidad en línea, ingrese a: 
                <a href="https://remisiones.streamlit.app/" target="_blank" style="color: #EC2024; font-weight: bold; text-decoration: underline;">https://remisiones.streamlit.app/</a>
            </p>
        </div>

        <p style="font-size: 9.5px; color: #64748B; margin-top: 8px;">* Se adjuntan a este correo el reporte interactivo en Excel (.xlsx) y el archivo de impresión PDF (.pdf).</p>

    </div>
    </body>
    </html>
    """
    return html

# Helper de resolución de imágenes
@st.cache_data(ttl=300, max_entries=50)
def obtener_imagen_sku(sku):
    if not sku or str(sku).strip().upper() in ['', 'NAN', 'NONE', 'N/A']:
        return None

    sku_clean = str(sku).strip().upper()
    os.makedirs("imagenes_articulos", exist_ok=True)

    if os.path.exists("imagenes_articulos"):
        for f in os.listdir("imagenes_articulos"):
            f_upper = f.upper()
            if f_upper.startswith(f"{sku_clean}(") or f_upper.startswith(f"{sku_clean}."):
                local_path = os.path.join("imagenes_articulos", f)
                try:
                    return Image.open(local_path)
                except Exception:
                    pass

    token = obtener_secret("github_token")
    if token:
        try:
            url_list = f"https://api.github.com/repos/{REPO_OWNER}/{REPO_NAME}/contents/imagenes_articulos?ref={BRANCH}"
            headers = {"Authorization": f"token {token}", "Accept": "application/vnd.github.v3+json"}
            res = requests.get(url_list, headers=headers)
            if res.status_code == 200:
                for item in res.json():
                    name_upper = item["name"].upper()
                    if name_upper.startswith(f"{sku_clean}(") or name_upper.startswith(f"{sku_clean}."):
                        download_url = item.get("download_url")
                        if not download_url:
                            download_url = f"https://raw.githubusercontent.com/{REPO_OWNER}/{REPO_NAME}/{BRANCH}/imagenes_articulos/{urllib.parse.quote(item['name'])}"
                        
                        img_res = requests.get(download_url)
                        if img_res.status_code == 200:
                            local_save_path = os.path.join("imagenes_articulos", item["name"])
                            with open(local_save_path, "wb") as f_out:
                                f_out.write(img_res.content)
                            try:
                                return Image.open(local_save_path)
                            except Exception:
                                pass
        except Exception:
            pass
    return None

# Generador de PDF oficial para impresión en Orientación Horizontal (Landscape)
def generar_pdf_consulta_reportlab(tipo_busqueda, valor_busqueda, df_tabla, spec_info=None, img_local_path=None):
    """Genera un reporte PDF profesional en orientacion Horizontal (Landscape) con proporciones de logo exactas y diseño ejecutivo."""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter), leftMargin=28, rightMargin=28, topMargin=24, bottomMargin=24)
    story = []
    styles = getSampleStyleSheet()
    
    style_company = ParagraphStyle('C_PDF', fontName="Helvetica-Bold", fontSize=14, textColor=colors.HexColor("#0F172A"), leading=16)
    style_subtitle = ParagraphStyle('S_PDF', fontName="Helvetica-Bold", fontSize=9.5, textColor=colors.HexColor("#EC2024"), leading=12)
    style_meta = ParagraphStyle('M_PDF', fontName="Helvetica", fontSize=8, textColor=colors.HexColor("#64748B"), leading=10)
    
    style_spec_lbl = ParagraphStyle('SL_PDF', fontName="Helvetica-Bold", fontSize=8, textColor=colors.HexColor("#64748B"))
    style_spec_val = ParagraphStyle('SV_PDF', fontName="Helvetica-Bold", fontSize=8.5, textColor=colors.HexColor("#0F172A"))
    
    style_hdr = ParagraphStyle('H_PDF', fontName="Helvetica-Bold", fontSize=7.5, textColor=colors.white, alignment=1)
    style_cell = ParagraphStyle('C_PDF', fontName="Helvetica", fontSize=7, leading=9, textColor=colors.HexColor("#1E293B"))
    style_cell_center = ParagraphStyle('CC_PDF', fontName="Helvetica", fontSize=7, leading=9, textColor=colors.HexColor("#1E293B"), alignment=1)
    style_cell_bold = ParagraphStyle('CB_PDF', fontName="Helvetica-Bold", fontSize=7.5, leading=9.5, textColor=colors.HexColor("#0F172A"), alignment=1)
    
    # 1. LOGOTIPO CON PROPORCIÓN EXACTA (SIN DEFORMACIÓN)
    logo_cell = Paragraph("", style_cell)
    if os.path.exists("logo_sigrama.png"):
        try:
            im_pil = Image.open("logo_sigrama.png")
            aspect = im_pil.height / im_pil.width
            target_w = 1.7 * inch
            target_h = target_w * aspect
            logo_cell = RLImage("logo_sigrama.png", width=target_w, height=target_h)
        except Exception:
            pass
            
    p_comp = Paragraph("INDUSTRIA SIGRAMA S.A. DE C.V.", style_company)
    p_sub = Paragraph("BASE DE DATOS SGP — PLANTA METALES DIAGONAL", style_subtitle)
    p_meta = Paragraph(f"REPORTE OFICIAL DE PIEZA | BÚSQUEDA: <b>{valor_busqueda}</b> | FECHA DE EMISIÓN: {datetime.date.today().strftime('%d/%m/%Y')}", style_meta)
    
    header_table = Table([[logo_cell, [p_comp, p_sub, p_meta]]], colWidths=[1.8*inch, 8.4*inch])
    header_table.setStyle(TableStyle([
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('PADDING', (0,0), (-1,-1), 0),
        ('BOTTOMPADDING', (0,0), (-1,-1), 6),
        ('LINEBELOW', (0,0), (-1,-1), 2, colors.HexColor("#EC2024")),
    ]))
    story.append(header_table)
    story.append(Spacer(1, 0.08 * inch))
    
    # 2. FICHA TÉCNICA Y DIBUJO/PLANO EN TARJETA ELEGANTE
    if spec_info:
        info_data = [
            [Paragraph("NÚMERO DE PARTE / PLANO:", style_spec_lbl), Paragraph(str(valor_busqueda), style_spec_val)],
            [Paragraph("DESCRIPCIÓN COMERCIAL:", style_spec_lbl), Paragraph(str(spec_info.get('nombre', 'N/A')), style_spec_val)],
            [Paragraph("CALIBRE / ESPESOR:", style_spec_lbl), Paragraph(str(spec_info.get('calibre', 'N/A')), style_spec_val)],
            [Paragraph("DIMENSIONES PIEZA:", style_spec_lbl), Paragraph(str(spec_info.get('dims', 'N/A')), style_spec_val)],
            [Paragraph("MATERIAL / ACABADO:", style_spec_lbl), Paragraph(str(spec_info.get('acabado', 'N/A')), style_spec_val)],
        ]
        info_table = Table(info_data, colWidths=[1.8*inch, 5.0*inch])
        info_table.setStyle(TableStyle([
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('PADDING', (0,0), (-1,-1), 2),
        ]))
        
        cell_img = Paragraph("<i>Sin imagen asignada</i>", style_cell_center)
        if img_local_path and os.path.exists(img_local_path):
            try:
                im_part = Image.open(img_local_path)
                p_aspect = im_part.height / im_part.width
                max_w = 2.4 * inch
                max_h = 1.15 * inch
                
                calc_w = max_w
                calc_h = max_w * p_aspect
                if calc_h > max_h:
                    calc_h = max_h
                    calc_w = max_h / p_aspect
                    
                cell_img = RLImage(img_local_path, width=calc_w, height=calc_h)
            except Exception:
                pass
                
        card_table = Table([[info_table, cell_img]], colWidths=[7.2*inch, 3.0*inch])
        card_table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,-1), colors.HexColor("#F8FAFC")),
            ('BOX', (0,0), (-1,-1), 0.75, colors.HexColor("#CBD5E1")),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('PADDING', (0,0), (-1,-1), 5)
        ]))
        story.append(card_table)
        story.append(Spacer(1, 0.1 * inch))
        
    # 3. TABLA DE DATOS CON ENCABEZADO CORPORATIVO ROJO/GRIS
    if not df_tabla.empty:
        df_clean = df_tabla.copy()
        for col in df_clean.columns:
            df_clean[col] = df_clean[col].astype(str).str.replace("➡️", " -> ").str.replace("➡", " -> ")
            
        headers = [Paragraph(str(col), style_hdr) for col in df_clean.columns]
        rows = [headers]
        
        for _, r in df_clean.iterrows():
            row_cells = []
            for col in df_clean.columns:
                val_str = str(r[col])
                if col in ['ID Tarima (TPM)', 'ID Tarima', 'Estatus Tarima', 'Estatus']:
                    style_use = style_cell_bold
                elif col in ['Fecha Empaque', 'Piezas', 'Cantidad', 'PO', 'Proyecto', 'Parcialidad']:
                    style_use = style_cell_center
                else:
                    style_use = style_cell
                row_cells.append(Paragraph(val_str, style_use))
            rows.append(row_cells)

        num_cols = len(df_clean.columns)
        if num_cols == 11:
            col_widths = [62, 58, 65, 60, 52, 36, 52, 65, 52, 82, 150]
        else:
            col_widths = [734 / num_cols] * num_cols

        t_rep = Table(rows, colWidths=col_widths, repeatRows=1)
        t_rep.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#1E293B")),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor("#CBD5E1")),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('TOPPADDING', (0,0), (-1,-1), 4),
            ('BOTTOMPADDING', (0,0), (-1,-1), 4),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor("#F8FAFC")]),
        ]))
        story.append(t_rep)
        
    doc.build(story)
    buffer.seek(0)
    return buffer.getvalue()

# Cargar Bases de Datos
df_articulos = cargar_excel_desde_github("BD_Articulos.xlsx")
df_skus_aut = cargar_excel_desde_github("BD_SKUs_Autorizados.xlsx")
df_tarimas = cargar_excel_desde_github("BD_Tarimas.xlsx")
df_detalle = cargar_excel_desde_github("BD_Detalle_Tarimas.xlsx")
df_remisiones = cargar_excel_desde_github("BD_Datos_Generales_Remision.xlsx")
df_actividad = cargar_excel_desde_github("BD_Actividad_Log.xlsx")

# --- RENDERIZADO DEL BANNER OFICIAL ESCALADO A 0.8X (80%) ---
c_banner1, c_banner2, c_banner3 = st.columns([0.1, 0.8, 0.1])
with c_banner2:
    if os.path.exists("REMISIONES APP.png"):
        st.image("REMISIONES APP.png", use_container_width=True)
    else:
        try:
            url_b = f"https://raw.githubusercontent.com/{REPO_OWNER}/{REPO_NAME}/{BRANCH}/REMISIONES%20APP.png"
            res_b = requests.get(url_b)
            if res_b.status_code == 200:
                with open("REMISIONES APP.png", "wb") as f_b:
                    f_b.write(res_b.content)
                st.image("REMISIONES APP.png", use_container_width=True)
            else:
                st.markdown("""
                <div class="sgp-header">
                    <h2>APLICACIÓN DE REMISIONES — SIGRAMA METALES</h2>
                    <p>Planta Metales Diagonal | Consulta de Piezas, Planos, Ubicación en Tarimas e Historial de Envíos</p>
                </div>
                """, unsafe_allow_html=True)
        except Exception:
            st.markdown("""
            <div class="sgp-header">
                <h2>APLICACIÓN DE REMISIONES — SIGRAMA METALES</h2>
                <p>Planta Metales Diagonal | Consulta de Piezas, Planos, Ubicación en Tarimas e Historial de Envíos</p>
            </div>
            """, unsafe_allow_html=True)

# Listas de SKUs
skus_con_foto = sorted([s for s in set(df_articulos['SKU'].dropna().astype(str).str.strip().str.upper()) if s not in ['', 'NAN', 'NONE', 'N/A']]) if not df_articulos.empty and 'SKU' in df_articulos.columns else []
skus_con_tarima = sorted([s for s in set(df_detalle['SKU'].dropna().astype(str).str.strip().str.upper()) if s not in ['', 'NAN', 'NONE', 'N/A']]) if not df_detalle.empty and 'SKU' in df_detalle.columns else []

skus_todos_set = set(skus_con_foto) | set(skus_con_tarima)
if not df_skus_aut.empty and 'SKU' in df_skus_aut.columns:
    skus_todos_set.update(df_skus_aut['SKU'].dropna().astype(str).str.strip().str.upper())
skus_todos = sorted([s for s in skus_todos_set if s not in ['', 'NAN', 'NONE', 'N/A']])

# Pestañas Principales
tab_sgp_piezas, tab_proyectos_po, tab_bitacora_global = st.tabs([
    "🧩 Consulta de Piezas y Fotografías",
    "📌 Consulta por Proyecto / PO",
    "📜 Bitácora e Historial Global"
])

# =============================================================================
# PESTAÑA 1: CONSULTA DE PIEZAS ESTILO SGP (SENCILLO Y DIRECTO)
# =============================================================================
with tab_sgp_piezas:
    # --- BARRA DE CONTROL Y NAVEGACIÓN SUPERIOR (SGP) ---
    if "sgp_sku_idx" not in st.session_state:
        st.session_state["sgp_sku_idx"] = 0

    lista_skus_activas = skus_con_foto if skus_con_foto else skus_todos

    def sgp_first(): st.session_state["sgp_sku_idx"] = 0
    def sgp_prev(): 
        if st.session_state["sgp_sku_idx"] > 0: st.session_state["sgp_sku_idx"] -= 1
    def sgp_next(): 
        if st.session_state["sgp_sku_idx"] < len(lista_skus_activas) - 1: st.session_state["sgp_sku_idx"] += 1
    def sgp_last(): st.session_state["sgp_sku_idx"] = len(lista_skus_activas) - 1

    def on_sgp_select():
        sel = st.session_state.get("sgp_select_sku_key")
        if sel in lista_skus_activas:
            st.session_state["sgp_sku_idx"] = lista_skus_activas.index(sel)

    idx_curr = max(0, min(st.session_state["sgp_sku_idx"], len(lista_skus_activas) - 1)) if lista_skus_activas else 0

    # BARRA SUPERIOR DE BOTONES NAVEGADORES (SGP STYLE)
    st.markdown("#### ⚡ Barra de Navegación de Piezas")
    c_btn1, c_btn2, c_sel, c_btn3, c_btn4, c_clear, c_print = st.columns([1, 1, 3.5, 1, 1, 1.2, 1.2])

    with c_btn1: st.button("⏮️ Primero", use_container_width=True, on_click=sgp_first, key="sgp_btn_first")
    with c_btn2: st.button("◀️ Anterior", use_container_width=True, on_click=sgp_prev, key="sgp_btn_prev")

    with c_sel:
        sku_nav_input = st.selectbox(
            f"Pieza {idx_curr+1} de {len(lista_skus_activas)}:",
            options=lista_skus_activas,
            index=idx_curr,
            key="sgp_select_sku_key",
            on_change=on_sgp_select,
            label_visibility="collapsed"
        )

    with c_btn3: st.button("Siguiente ▶️", use_container_width=True, on_click=sgp_next, key="sgp_btn_next")
    with c_btn4: st.button("Último ⏭️", use_container_width=True, on_click=sgp_last, key="sgp_btn_last")

    with c_clear:
        if st.button("🧹 Limpiar", use_container_width=True, key="sgp_btn_clear"):
            st.session_state["sgp_sku_idx"] = 0
            if "sgp_select_sku_key" in st.session_state: del st.session_state["sgp_select_sku_key"]
            st.rerun()

    with c_print:
        components.html("""
        <button onclick="window.parent.print()" style="
            background-color: #EC2024; color: white; border: none; padding: 7px 12px;
            font-size: 13px; font-weight: bold; border-radius: 6px; cursor: pointer; width: 100%;
            font-family: 'Montserrat', sans-serif; box-shadow: 0 2px 4px rgba(0,0,0,0.1);
        ">🖨️ Imprimir</button>
        """, height=38)

    sku_actual = lista_skus_activas[idx_curr] if lista_skus_activas else None

    # --- PANEL PRINCIPAL DE LA PIEZA (ESTILO SGP DE LA FOTO) ---
    if sku_actual:
        st.write("")
        col_sgp_left, col_sgp_right = st.columns([1.1, 1.2])

        spec_dict = {'nombre': 'No especificado', 'calibre': 'N/A', 'dims': 'N/A', 'acabado': 'N/A'}
        if not df_articulos.empty and 'SKU' in df_articulos.columns:
            df_m = df_articulos[df_articulos['SKU'].astype(str).str.strip().str.upper() == sku_actual]
            if not df_m.empty:
                r_m = df_m.iloc[0]
                spec_dict['nombre'] = str(r_m.get('Nombre', 'No especificado'))
                spec_dict['calibre'] = str(r_m.get('Calibre_Espesor', 'N/A'))
                spec_dict['dims'] = str(r_m.get('Dimensiones_Pieza', 'N/A'))
                spec_dict['acabado'] = str(r_m.get('Acabado_Superficial', 'N/A'))

        with col_sgp_left:
            st.markdown(f"""
            <div class="sgp-card">
                <div class="sgp-label" style="text-align:center;">NÚMERO DE PARTE / PLANO</div>
                <div class="sgp-sku-title">{sku_actual}</div>
                <div style="margin-bottom: 12px;">
                    <div class="sgp-label">DESCRIPCIÓN COMERCIAL:</div>
                    <div class="sgp-value" style="color: #F8FAFC;">{spec_dict['nombre']}</div>
                </div>
                <div style="display: flex; gap: 15px; margin-bottom: 12px;">
                    <div style="flex:1;">
                        <div class="sgp-label">CALIBRE / ESPESOR:</div>
                        <div class="sgp-value">{spec_dict['calibre']}</div>
                    </div>
                    <div style="flex:1;">
                        <div class="sgp-label">ACABADO / MATERIAL:</div>
                        <div class="sgp-value" style="color: #EC2024;">{spec_dict['acabado']}</div>
                    </div>
                </div>
                <div>
                    <div class="sgp-label">DIMENSIONES PIEZA:</div>
                    <div class="sgp-value">{spec_dict['dims']}</div>
                </div>
            </div>
            """, unsafe_allow_html=True)

        with col_sgp_right:
            img_sku = obtener_imagen_sku(sku_actual)
            img_path_local = None
            if img_sku:
                st.image(img_sku, caption=f"Plano / Fotografía de Pieza {sku_actual}", use_container_width=True)
                matching_local = glob.glob(f"imagenes_articulos/{sku_actual}(*.*")
                if matching_local: img_path_local = matching_local[0]
            else:
                st.info(f"📷 *Sin imagen fotográfica asignada para la pieza {sku_actual} en catálogo.*")

        st.write("")
        st.write("---")

        # --- BOTÓN CLARO Y PROMINENTE PARA VER HISTORIAL Y UBICACIÓN FÍSICA ---
        with st.expander(f"📜 VER HISTORIAL Y UBICACIÓN FÍSICA DE LA PIEZA `{sku_actual}`", expanded=True):
            if not df_detalle.empty:
                df_sub_det = df_detalle[df_detalle['SKU'].astype(str).str.strip().str.upper() == sku_actual].copy()
                
                if df_sub_det.empty:
                    st.info(f"ℹ️ El SKU **{sku_actual}** está registrado en el catálogo master, pero **aún no cuenta con tarimas físicamente registradas en planta**.")
                else:
                    if not df_tarimas.empty:
                        df_sub_det = pd.merge(
                            df_sub_det, 
                            df_tarimas[['ID_Tarima', 'Fecha_Creacion', 'Creado_Por', 'Ubicacion_Actual', 'Estatus']], 
                            on="ID_Tarima", 
                            how="left"
                        )
                    else:
                        df_sub_det['Fecha_Creacion'] = "N/A"
                        df_sub_det['Creado_Por'] = "N/A"
                        df_sub_det['Ubicacion_Actual'] = "Metales"
                        df_sub_det['Estatus'] = "Disponible"

                    df_sub_det['Ubicacion_Actual'] = df_sub_det['Ubicacion_Actual'].fillna("Metales")
                    df_sub_det['Estatus'] = df_sub_det['Estatus'].fillna("Disponible")
                    df_sub_det['Cantidad'] = pd.to_numeric(df_sub_det['Cantidad'], errors='coerce').fillna(0).astype(int)

                    rem_map = {}
                    rem_date_map = {}
                    if not df_remisiones.empty:
                        import ast
                        for _, r_row in df_remisiones.iterrows():
                            fol = str(r_row.get('Folio_Remision', ''))
                            fec_raw = r_row.get('Fecha_Hora_Salida', '')
                            fec = normalizar_fecha_display(fec_raw)
                            rec = str(r_row.get('Nombre_Receptor', ''))
                            dir_rec = str(r_row.get('Direccion_Receptor', ''))
                            asoc = r_row.get('Tarimas_Asociadas', '')
                            if isinstance(asoc, str):
                                try: asoc = ast.literal_eval(asoc)
                                except Exception: asoc = [asoc]
                            if isinstance(asoc, list):
                                for t_id in asoc:
                                    t_str = str(t_id).strip()
                                    rem_map[t_str] = f"Remisión {fol} ({fec}) ➡️ {rec} [{dir_rec}]"
                                    rem_date_map[t_str] = fec_raw

                    df_sub_det['Detalle_Remision'] = df_sub_det['ID_Tarima'].astype(str).str.strip().map(lambda x: rem_map.get(x, "En Planta / Almacén"))

                    # Helper para obtener timestamp comparable seguro del último movimiento
                    def get_latest_dt(row):
                        t_id = str(row.get('ID_Tarima', '')).strip()
                        d_rem = rem_date_map.get(t_id)
                        d_emp = row.get('Fecha_Creacion')
                        
                        dt_rem = pd.to_datetime(d_rem, format='mixed', dayfirst=True, errors='coerce') if d_rem else pd.NaT
                        dt_emp = pd.to_datetime(d_emp, format='mixed', dayfirst=True, errors='coerce') if d_emp else pd.NaT
                        
                        if pd.notna(dt_rem) and pd.notna(dt_emp): return max(dt_rem, dt_emp)
                        if pd.notna(dt_rem): return dt_rem
                        if pd.notna(dt_emp): return dt_emp
                        return pd.Timestamp.min

                    # ORDENAR ESTRICTAMENTE DE MÁS RECIENTE A MÁS ANTIGUO (ÚLTIMO MOVIMIENTO AL PRINCIPADO)
                    df_sub_det['_dt_ult_mov'] = df_sub_det.apply(get_latest_dt, axis=1)
                    df_sub_det = df_sub_det.sort_values(by='_dt_ult_mov', ascending=False)
                    
                    if 'Fecha_Creacion' in df_sub_det.columns:
                        df_sub_det['Fecha_Creacion'] = df_sub_det['Fecha_Creacion'].apply(normalizar_fecha_display)

                    pzs_disp = df_sub_det[df_sub_det['Estatus'] == 'Disponible']['Cantidad'].sum()
                    pzs_rem = df_sub_det[df_sub_det['Estatus'] == 'Remesada']['Cantidad'].sum()
                    pzs_tot = df_sub_det['Cantidad'].sum()
                    tar_tot = df_sub_det['ID_Tarima'].nunique()

                    m1, m2, m3, m4 = st.columns(4)
                    m1.metric("📦 Piezas Disponibles en Planta", f"{pzs_disp:,} PZS")
                    m2.metric("🚚 Piezas Remesadas (Enviadas)", f"{pzs_rem:,} PZS")
                    m3.metric("🏗️ Total Piezas Registradas", f"{pzs_tot:,} PZS")
                    m4.metric("📊 Total Tarimas Físicas", f"{tar_tot} Tarimas")

                    st.write("")
                    cols_mostrar = ['ID_Tarima', 'Fecha_Creacion', 'Creado_Por', 'Ubicacion_Actual', 'Estatus', 'Cantidad', 'Proyecto', 'PO', 'Parcialidad', 'Descripcion', 'Detalle_Remision']
                    df_tabla_export = df_sub_det[[c for c in cols_mostrar if c in df_sub_det.columns]].copy()
                    df_tabla_export = df_tabla_export.rename(columns={
                        'ID_Tarima': 'ID Tarima (TPM)',
                        'Fecha_Creacion': 'Fecha Empaque',
                        'Creado_Por': 'Líder Empaque',
                        'Ubicacion_Actual': 'Ubicación Actual',
                        'Estatus': 'Estatus Tarima',
                        'Cantidad': 'Piezas',
                        'Descripcion': 'Descripción Proyecto',
                        'Detalle_Remision': 'Estatus de Remisión / Destino'
                    })

                    st.dataframe(df_tabla_export, use_container_width=True, hide_index=True)

                    c_dl1, c_dl2, c_dl3 = st.columns(3)
                    
                    # 1. Excel con Formato Ejecutivo (openpyxl)
                    xl_bytes = generar_excel_consulta_inventario(sku_actual, spec_dict, pzs_disp, pzs_rem, pzs_tot, tar_tot, df_tabla_export)
                    
                    with c_dl1:
                        st.download_button(
                            label=f"📥 Descargar Registro Excel ({sku_actual}.xlsx)",
                            data=xl_bytes,
                            file_name=f"Consulta_Inventario_{sku_actual}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            key="btn_dl_excel_sgp"
                        )
                        
                    # 2. PDF
                    pdf_bytes = generar_pdf_consulta_reportlab("Pieza SGP", sku_actual, df_tabla_export, spec_info=spec_dict, img_local_path=img_path_local)
                    with c_dl2:
                        st.download_button(
                            label=f"📄 Descargar PDF Oficial de Impresión ({sku_actual}.pdf)",
                            data=pdf_bytes,
                            file_name=f"Reporte_SGP_Pieza_{sku_actual}.pdf",
                            mime="application/pdf",
                            key="btn_dl_pdf_sgp"
                        )

                    # 3. Borrador EML con Excel + PDF Adjuntos e Imágenes Inline (Réplica PDF)
                    inline_images_eml = {}
                    if os.path.exists("logo_sigrama.png"):
                        try:
                            with open("logo_sigrama.png", "rb") as f_l:
                                inline_images_eml['logo_sigrama_cid'] = f_l.read()
                        except Exception: pass

                    if img_path_local and os.path.exists(img_path_local):
                        try:
                            with open(img_path_local, "rb") as f_i:
                                inline_images_eml['foto_sku_cid'] = f_i.read()
                        except Exception: pass

                    has_logo_cid = 'logo_sigrama_cid' in inline_images_eml
                    has_img_cid = 'foto_sku_cid' in inline_images_eml

                    cuerpo_eml_html = generar_html_correo_sku(
                        sku_actual, spec_dict, pzs_disp, pzs_rem, pzs_tot, tar_tot, df_tabla_export,
                        has_logo_cid=has_logo_cid, has_img_cid=has_img_cid
                    )
                    adjuntos_eml = {
                        f"Reporte_Inventario_{sku_actual}.xlsx": xl_bytes,
                        f"Reporte_Impresion_{sku_actual}.pdf": pdf_bytes
                    }
                    dest_cc_sku = "bryan.mancinas@sigrama.com.mx; cruz.carreon@sigrama.com.mx; jesus.morales@sigrama.com.mx; jose.fernandez@sigrama.com.mx; luis.quintana@sigrama.com.mx"
                    eml_bytes = generar_archivo_eml(
                        dest_to="",
                        dest_cc=dest_cc_sku,
                        subject=f"Reporte de Inventario e Historial - SKU: {sku_actual} - Industria Sigrama",
                        body_html=cuerpo_eml_html,
                        adjuntos_dict=adjuntos_eml,
                        inline_images_dict=inline_images_eml
                    )
                    with c_dl3:
                        st.download_button(
                            label=f"📩 Descargar Borrador Correo (.eml) (Excel + PDF)",
                            data=eml_bytes,
                            file_name=f"Correo_Reporte_Inventario_{sku_actual}.eml",
                            mime="message/rfc822",
                            key="btn_dl_eml_sku"
                        )
            else:
                st.info("No hay datos de detalle de tarimas cargados.")

# =============================================================================
# PESTAÑA 2: BÚSQUEDA POR PROYECTO O PO
# =============================================================================
with tab_proyectos_po:
    st.markdown("### 📌 Búsqueda de Inventario por Proyecto u Orden de Compra (PO)")
    col_proj_sel, col_po_sel = st.columns(2)
    
    with col_proj_sel:
        if not df_detalle.empty and 'Proyecto' in df_detalle.columns:
            proyectos_disponibles = sorted([str(p).strip() for p in df_detalle['Proyecto'].dropna().unique() if str(p).strip() not in ['', 'nan', 'None']])
            p_sel = st.selectbox("Seleccione el Proyecto:", options=["(Seleccione...)"] + proyectos_disponibles, key="sgp_proj_sel")
            if p_sel != "(Seleccione...)":
                df_proj = df_detalle[df_detalle['Proyecto'].astype(str).str.strip() == p_sel].copy()
                if not df_proj.empty:
                    if 'Fecha_Creacion' in df_proj.columns:
                        df_proj['_dt_sort'] = pd.to_datetime(df_proj['Fecha_Creacion'], errors='coerce', dayfirst=True)
                        df_proj = df_proj.sort_values(by='_dt_sort', ascending=False)
                    st.write(f"#### Partidas del Proyecto `{p_sel}`")
                    st.dataframe(df_proj[['ID_Tarima', 'SKU', 'PO', 'Parcialidad', 'Cantidad', 'Descripcion']], use_container_width=True, hide_index=True)

    with col_po_sel:
        if not df_detalle.empty and 'PO' in df_detalle.columns:
            pos_disponibles = sorted([str(p).strip() for p in df_detalle['PO'].dropna().unique() if str(p).strip() not in ['', 'nan', 'None']])
            po_sel = st.selectbox("Seleccione la PO:", options=["(Seleccione...)"] + pos_disponibles, key="sgp_po_sel")
            if po_sel != "(Seleccione...)":
                df_po = df_detalle[df_detalle['PO'].astype(str).str.strip() == po_sel].copy()
                if not df_po.empty:
                    if 'Fecha_Creacion' in df_po.columns:
                        df_po['_dt_sort'] = pd.to_datetime(df_po['Fecha_Creacion'], errors='coerce', dayfirst=True)
                        df_po = df_po.sort_values(by='_dt_sort', ascending=False)
                    st.write(f"#### Partidas de la Orden `{po_sel}`")
                    st.dataframe(df_po[['ID_Tarima', 'SKU', 'Proyecto', 'Parcialidad', 'Cantidad', 'Descripcion']], use_container_width=True, hide_index=True)

# =============================================================================
# PESTAÑA 3: BITÁCORA E HISTORIAL GLOBAL
# =============================================================================
with tab_bitacora_global:
    st.markdown("### 📜 Bitácora e Historial Global de Movimientos")
    if not df_actividad.empty:
        q_search = st.text_input("🔍 Buscar en historial (SKU, TPM-XXXX, Remisión, Usuario, PO):", key="sgp_log_search")
        df_log_f = df_actividad.copy()
        if 'Fecha_Hora' in df_log_f.columns:
            df_log_f['_dt_sort'] = pd.to_datetime(df_log_f['Fecha_Hora'], errors='coerce', dayfirst=True)
            df_log_f = df_log_f.sort_values(by='_dt_sort', ascending=False)
        if q_search.strip():
            q = q_search.strip().lower()
            mask = df_log_f.apply(lambda row: row.astype(str).str.lower().str.contains(q).any(), axis=1)
            df_log_f = df_log_f[mask]
        st.dataframe(df_log_f, use_container_width=True, hide_index=True)

st.markdown("---")
st.markdown("<p style='text-align: center; color: #94A3B8; font-size: 12px;'>Industria SIGRAMA S.A. de C.V. — Base de Datos SGP Metales</p>", unsafe_allow_html=True)
